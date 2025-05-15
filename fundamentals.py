import streamlit as st
import datetime
# from streamlit_option_menu import option_menu
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np

import base64  # Standard Python Module
from io import StringIO, BytesIO  # Standard Python Module

#text_rgb = {'#f8ba43' : "rgb(248,186,67)", "#D6D41B":"rgb(248,186,67)", '#1959BF': "rgb(33,161,225)", '#0971C9': "rgb(33,161,225)",'#11A694': "rgb(0,253,85)", '#11A64B': "rgb(0,253,85)", }
# color_dict = {'black':{'hash':'000000','rgb':'rgb(0,0,0)'}, 'white':{'hash':'#ffffff','rgb':'rgb(255, 255, 255)'},
#               'blue1':{'hash':'#21A1E1', 'rgb':'rgb(33,161,225)'}, 'blue2':{'hash':'#5DB7D2','rgb':'rgb(93,183,210)'},
#               'blue3':{'hash':'#00A3FE','rgb':'rgb(0,163,254)'},'green1':{'hash':'#00F954','rgb':'rgb(0,249,84)'},
#               'yellow1':{'hash':'#FFFF01','rgb':'rgb(255,255,1)'},'yellow2':{'hash':'#FFFE57','rgb':'rgb(255,254,87)'},
#               'red1':{'hash':'#CC0118','rgb':'rgb(204,1,24)'},
#               }
tabs = ['Profit & Loss', 'Quarters','Balance Sheet', 'Cash Flow' ,'Data Sheet']
table = ['PROFIT & LOSS', 'Quarters', 'BALANCE SHEET' ]
# lets define the first and last table keys to extract the exact table size
DataSheet_Key_Values = ['PROFIT & LOSS', 'Dividend Amount', 'Quarters', 'Operating Profit', 'BALANCE SHEET', 'Face value',
                        'CASH FLOW:', 'Net Cash Flow','DERIVED:','Adjusted Equity Shares in Cr']

funda_keys = ['PROFIT&LOSS','BALANCE SHEET','CASH FLOW']                # used in edeveloping yearly data while reading from its CSV
funda_menu = ['Key_Data'] + ['QTR PnL'] + funda_keys # used for showing the menu in Interactivechat.py

# # color_hover = "darkgrey"
# # color_background = "grey"
# #Top 16:9 Resolutions. 640 x 360 (nHD) 854 x 480 (FWVGA) 960 x 540 (qHD) 1024 x 576 (WSVGA) 1280 x 720 (HD/WXGA) 1366 x 768 (FWXGA) 1600 x 900 (HD+) 1920 x 1080 (FHD) 2048 x 1152 (QWXGA) 2560 x 1440 (QHD) 3200 x 1800 (WQXGA+) 3840 x 2160 (UHD) 5120 x 2880 (UHD+) 7680 x 4320 (FUHD) 15360 x 8640 (QUHD) 30720 x 17280 (HHD) 61440 x 34560 (FHHD) 122880 x 69120 (QHHD)
# width_val = 1120 #1024
# height_val = 360 #574
# #aspect_ratio = 16/9  # For a 16:9 aspect ratio
# #height_val = width_val / aspect_ratio
# # TARGET_FOLDER = "C:/Users/sahaveer/OneDrive/Documents/bhavcopy/"

# #write_on_chart = "<i>@itimesalgo        </i>"
# write_on_chart = "<i>https://itimesalgo.streamlit.app/</i>"

# if 'nsecode_list' not in st.session_state:
#     import nse_bse_search
# if 'nsecode_list' not in st.session_state or 'nseISIN_list' not in st.session_state:
#     nse_data = pd.read_csv('./cm21JUN2024bhav.csv')
#     nse_data.columns = nse_data.columns.str.replace(' ', '_')
#     # nse_ISIN = nse_data["ISIN"].tolist()
#     # nse_code = nse_data["SYMBOL"].tolist()
#     # st.session_state.nseISIN_list = nse_ISIN
#     # st.session_state.nsecode_list = nse_code
#     st.session_state.nseISIN_list = nse_data["ISIN"].tolist()
#     st.session_state.nsecode_list = nse_data["SYMBOL"].tolist()

# nsecode = st.session_state.nsecode_list

# def excel_link_to_download(df,filename,label):
#     # Credit Excel: https://discuss.streamlit.io/t/how-to-add-a-Download-excel-csv-function-to-a-button/4474/5
#     towrite = BytesIO()
#     df.to_excel(towrite, encoding="utf-8", index=False, header=True)  # write to BytesIO buffer
#     towrite.seek(0)  # reset pointer
#     b64 = base64.b64encode(towrite.read()).decode()
#     #href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" Download="data_Download.xlsx">Download Excel File</a>'
#     href = f'<a href="data:file/xlsx;base64,{b64}" download="{filename}">{label}</a>'
#     return st.markdown(href, unsafe_allow_html=True)

# def generate_html_Download_link(fig):
#     # Credit Plotly: https://discuss.streamlit.io/t/Download-plotly-plot-as-html/4426/2
#     towrite = StringIO()
#     fig.write_html(towrite, include_plotlyjs="cdn")
#     towrite = BytesIO(towrite.getvalue().encode())
#     b64 = base64.b64encode(towrite.read()).decode()
#     href = f'<a href="data:text/html;charset=utf-8;base64, {b64}" Download="plot.html">Download Plot</a>'
#     return st.markdown(href, unsafe_allow_html=True)


def get_tables(datasht,file):
    # st.success("Got in get_tables in fundamentals.py")
    for i in range(1,datasht.max_row+1) :
        if datasht['A'+str(i)].value == DataSheet_Key_Values[0] :
            pnl_start_row = i
        if datasht['A' + str(i)].value == DataSheet_Key_Values[1]:
            pnl_end_row = i-1
        if datasht['A' + str(i)].value == DataSheet_Key_Values[2]:
            quarterly_start_row = i
        if datasht['A' + str(i)].value == DataSheet_Key_Values[3]:
            quarterly_end_row = i-1
        if datasht['A' + str(i)].value == DataSheet_Key_Values[4]:
            BS_start_row = i
        if datasht['A' + str(i)].value == DataSheet_Key_Values[5]:
            BS_end_row = i-1
        if datasht['A' + str(i)].value == DataSheet_Key_Values[6]:
            cash_start_row = i
        if datasht['A' + str(i)].value == DataSheet_Key_Values[7]:
            cash_end_row = i-1
        if datasht['A' + str(i)].value == DataSheet_Key_Values[8]:
            eq_start_row = i
        if datasht['A' + str(i)].value == DataSheet_Key_Values[9]:
            eq_end_row = i-1
        
    reqd_cols = "A :" + str(get_column_letter(datasht.max_column))
    try:
        if pnl_start_row is not None and pnl_end_row is not None :
            pnl = pd.read_excel(file, index_col=0, sheet_name=tabs[-1], header=pnl_start_row,usecols=reqd_cols,
                                nrows=pnl_end_row-pnl_start_row)
            pnl.fillna(0,inplace=True)
            pnl.index = pnl.index.str.strip()
            pnl.index = pnl.index.str.upper()
            pnl = pnl.loc[:, (pnl != 0).any()]

            # if eq_start_row is not None and eq_end_row is not None :
            #     noofeq = pd.read_excel(file, index_col=0, sheet_name=tabs[-1], header=eq_start_row-1,usecols=reqd_cols,
            #                         nrows=1)
            #     # replace column values with the above cash flow columns
            #     noofeq.columns = pnl.columns
            #     noofeq.fillna(0,inplace=True)
            #     noofeq.index = noofeq.index.str.strip()
            #     noofeq.index = noofeq.index.str.upper()
            #     noofeq = noofeq.loc[:, (noofeq != 0).any()]
            #     # st.dataframe(noofeq)
            #     # Append noofeq to pnl
            #     pnl1 = pd.concat([pnl, noofeq], axis=0)
            #     st.dataframe(pnl1)
            # else:
            #     pnl1 = pnl

        if quarterly_start_row is not None and quarterly_end_row is not None :
            qtr_pnl = pd.read_excel(file, index_col=0, sheet_name=tabs[-1], header=quarterly_start_row, usecols=reqd_cols,
                                    nrows=quarterly_end_row - quarterly_start_row)

            qtr_pnl.fillna(0,inplace=True)
            qtr_pnl.index = qtr_pnl.index.str.strip()
            qtr_pnl.index = qtr_pnl.index.str.upper()
            qtr_pnl = qtr_pnl.loc[:, (qtr_pnl != 0).any()]
        if BS_start_row is not None and BS_end_row is not None:
            balancesht = pd.read_excel(file, index_col=0, sheet_name=tabs[-1], header=BS_start_row,usecols=reqd_cols,
                                    nrows=BS_end_row-BS_start_row)
            #balancesht.columns = balancesht.columns.strftime('%d-%m-%Y')
            balancesht.fillna(0,inplace=True)
            balancesht.index = balancesht.index.str.strip()
            balancesht.index = balancesht.index.str.upper()
            balancesht = balancesht.loc[:, (balancesht != 0).any()]
        if cash_start_row is not None and cash_end_row is not None:
            cashflow = pd.read_excel(file, index_col=0, sheet_name=tabs[-1], header=cash_start_row,usecols=reqd_cols,
                                    nrows=cash_end_row-cash_start_row)
            #cashflow.columns = cashflow.columns.strftime('%d-%m-%Y')
            cashflow.fillna(0,inplace=True)
            cashflow.index = cashflow.index.str.strip()
            cashflow.index = cashflow.index.str.upper()
            cashflow = cashflow.loc[:, (cashflow != 0).any()]
            # st.success(cash_start_row)          #80
            # st.success(cash_end_row)            #84
            # st.success(eq_start_row)            #92
            # st.success(eq_end_row)              #92
        
        if(pnl is not None and balancesht is not None and cashflow is not None and qtr_pnl is not None):
            sht_list = [pnl,balancesht,cashflow]
            df_comp = pd.concat(sht_list,keys=funda_keys)
            # st.success(df_comp.columns)
            df_comp.index.set_names(['Section', 'Item'], inplace=True)
            # st.success(df_comp.columns)
            #df_comp = df_comp.drop(columns=[''])
            #qtr_pnl = qtr_pnl.drop(columns=[''])
        return qtr_pnl,df_comp
    except Exception as e:
        st.error(f"Error reading excel file: {e}")
        return pd.DataFrame(),pd.DataFrame()
    

# Define a custom function to apply the condition
def OPM(row):
    if row['OPERATING PROFIT'] > 0:
        return round((row['OPERATING PROFIT'] / row['SALES'])*100,1)
    else:
        return 0

def NPM(row):
    if row['NET PROFIT'] > 0:
        return round((row['NET PROFIT'] / row['SALES'])*100,1)
    else:
        return 0


def develop_data(qtr_pnl, df_comp):
    pnl = df_comp.loc['PROFIT&LOSS']
    pnl.fillna(0, inplace=True)
    pnl.index = pnl.index.str.strip()
    pnl = pnl.transpose()
    pnl['EXPENSES'] = pnl['RAW MATERIAL COST'] - pnl['CHANGE IN INVENTORY'] + pnl['POWER AND FUEL'] + pnl[
        'OTHER MFR. EXP'] + pnl['EMPLOYEE COST'] + pnl['SELLING AND ADMIN'] + pnl['OTHER EXPENSES']
    pnl = pnl.drop(
        columns=['RAW MATERIAL COST', 'CHANGE IN INVENTORY', 'POWER AND FUEL', 'OTHER MFR. EXP', 'EMPLOYEE COST',
                 'SELLING AND ADMIN', 'OTHER EXPENSES'], axis=1)
    pnl['OPERATING PROFIT'] = pnl['SALES'] - pnl['EXPENSES']
    pnl['OPM %'] = pnl.apply(OPM, axis=1)
    pnl['NPM %'] = pnl.apply(NPM, axis=1)
    # Calculate the QoQ percentage increase for SALES, NET PROFIT, and OPERATING PROFIT
    pnl['SALES_QoQ'] = pnl['SALES'].pct_change() * 100
    pnl['NET PROFIT_QoQ'] = pnl['NET PROFIT'].pct_change() * 100
    pnl['OPERATING PROFIT_QoQ'] = pnl['OPERATING PROFIT'].pct_change() * 100

    balancesht = df_comp.loc['BALANCE SHEET'].drop(index='TOTAL', errors='ignore')
    balancesht.fillna(0, inplace=True)
    balancesht.index = balancesht.index.str.strip()
    balancesht = balancesht.transpose()
    balancesht['WORKING CAPITAL'] = balancesht['OTHER ASSETS'] - balancesht['OTHER LIABILITIES']
    balancesht['DEBTOR DAYS'] = np.where(pnl['SALES'] > 0, balancesht['RECEIVABLES'] / (pnl['SALES'] / 365), 0)
    balancesht['INVENTORY TURNOVER'] = np.where(balancesht['INVENTORY'] > 0, pnl['SALES'] / balancesht['INVENTORY'],
                                                0)
    balancesht['ROCE'] = np.where(balancesht['NET BLOCK'] + balancesht['WORKING CAPITAL'] > 0, (
                (pnl['OPERATING PROFIT'] - pnl['DEPRECIATION'] - pnl['TAX']) / (
                    balancesht['NET BLOCK'] + balancesht['WORKING CAPITAL'])) * 100, 0)

    qtr_pnl.fillna(0, inplace=True)
    qtr_pnl.index = qtr_pnl.index.str.strip()
    qtr_pnl = qtr_pnl.transpose()
    qtr_pnl['OPM %'] = qtr_pnl.apply(OPM, axis=1)
    qtr_pnl['NPM %'] = qtr_pnl.apply(NPM, axis=1)
    qtr_pnl['SALES_QoQ'] = qtr_pnl['SALES'].pct_change() * 100
    qtr_pnl['NET PROFIT_QoQ'] = qtr_pnl['NET PROFIT'].pct_change() * 100
    qtr_pnl['OPERATING PROFIT_QoQ'] = qtr_pnl['OPERATING PROFIT'].pct_change() * 100

    pnl = pnl.transpose()
    pnl = pnl.round(1)
    #pnl.columns = pnl.columns.strftime('%d-%m-%Y')
    # st.dataframe(pnl)
    balancesht = balancesht.transpose()
    balancesht = balancesht.round(1)
    # st.dataframe(balancesht)
    qtr_pnl = qtr_pnl.transpose()
    qtr_pnl = qtr_pnl.round(1)
    #qtr_pnl.columns = qtr_pnl.columns.strftime('%d-%m-%Y')

    # st.dataframe(qtr_pnl)
    return pnl, balancesht, qtr_pnl

def develop_quarterly(qtr_pnl):
    try:
        qtr_pnl.columns = pd.to_datetime(qtr_pnl.columns, format='%d-%m-%Y')
    except Exception as ValueError:
        qtr_pnl.columns = pd.to_datetime(qtr_pnl.columns, format='%Y-%m-%d')
        
    qtr_pnl.fillna(0, inplace=True)
    qtr_pnl.index = qtr_pnl.index.str.strip()
    qtr_pnl = qtr_pnl.transpose()
    qtr_pnl['OPM %'] = qtr_pnl.apply(OPM, axis=1)
    qtr_pnl['NPM %'] = qtr_pnl.apply(NPM, axis=1)
    qtr_pnl['SALES_QoQ'] = qtr_pnl['SALES'].pct_change() * 100
    qtr_pnl['NET PROFIT_QoQ'] = qtr_pnl['NET PROFIT'].pct_change() * 100
    qtr_pnl['OPERATING PROFIT_QoQ'] = qtr_pnl['OPERATING PROFIT'].pct_change() * 100

    if len(qtr_pnl.columns) >= 5:  # Check if there are at least 5 rows in the DataFrame
        qtr_pnl['SALES_YoY'] = qtr_pnl['SALES'].pct_change(periods=4) * 100
        qtr_pnl['OPERATING PROFIT_YoY'] = qtr_pnl['OPERATING PROFIT'].pct_change(periods=4) * 100
        qtr_pnl['NET PROFIT_YoY'] = qtr_pnl['NET PROFIT'].pct_change(periods=4) * 100

    else:
        qtr_pnl['SALES_YoY'] = None  # Assign None or any placeholder value when insufficient data is available
        qtr_pnl['OPERATING PROFIT_YoY'] = None
        qtr_pnl['NET PROFIT_YoY'] = None

    qtr_pnl = qtr_pnl.transpose()
    qtr_pnl = qtr_pnl.round(1)
    try:
        qtr_pnl.columns = pd.to_datetime(qtr_pnl.columns, format='%d-%m-%Y')
    except Exception as ValueError:
        qtr_pnl.columns = pd.to_datetime(qtr_pnl.columns, format='%Y-%m-%d')

    # st.dataframe(qtr_pnl)
    return qtr_pnl

def develop_yearly(df_comp):
    try:
        df_comp.columns = pd.to_datetime(df_comp.columns, format='%d-%m-%Y')
    except Exception as ValueError:
        df_comp.columns = pd.to_datetime(df_comp.columns, format='%Y-%m-%d')
    # st.dataframe(df_comp)
    pnl = df_comp.loc['PROFIT&LOSS']
    pnl.fillna(0, inplace=True)
    pnl.index = pnl.index.str.strip()
    pnl = pnl.transpose()
    pnl['EXPENSES'] = pnl['RAW MATERIAL COST'] - pnl['CHANGE IN INVENTORY'] + pnl['POWER AND FUEL'] + pnl[
        'OTHER MFR. EXP'] + pnl['EMPLOYEE COST'] + pnl['SELLING AND ADMIN'] + pnl['OTHER EXPENSES']
    pnl = pnl.drop(
        columns=['RAW MATERIAL COST', 'CHANGE IN INVENTORY', 'POWER AND FUEL', 'OTHER MFR. EXP', 'EMPLOYEE COST',
                 'SELLING AND ADMIN', 'OTHER EXPENSES'], axis=1)
    pnl['OPERATING PROFIT'] = pnl['SALES'] - pnl['EXPENSES']
    pnl['OPM %'] = pnl.apply(OPM, axis=1)
    pnl['NPM %'] = pnl.apply(NPM, axis=1)
    # Calculate the QoQ percentage increase for SALES, NET PROFIT, and OPERATING PROFIT
    pnl['SALES_QoQ'] = pnl['SALES'].pct_change() * 100
    pnl['NET PROFIT_QoQ'] = pnl['NET PROFIT'].pct_change() * 100
    pnl['OPERATING PROFIT_QoQ'] = pnl['OPERATING PROFIT'].pct_change() * 100
    # Assuming 'qtr_pnl' is your DataFrame and 'SALES' column exists
    if len(pnl.columns) >= 5:  # Check if there are at least 5 rows in the DataFrame
        pnl['SALES_YoY'] = pnl['SALES'].pct_change(periods=4) * 100
        pnl['OPERATING PROFIT_YoY'] = pnl['OPERATING PROFIT'].pct_change(periods=4) * 100
        pnl['NET PROFIT_YoY'] = pnl['NET PROFIT'].pct_change(periods=4) * 100

    else:
        pnl['SALES_YoY'] = None  # Assign None or any placeholder value when insufficient data is available
        pnl['OPERATING PROFIT_YoY'] = None
        pnl['NET PROFIT_YoY'] = None

    balancesht = df_comp.loc['BALANCE SHEET'].drop(index='TOTAL', errors='ignore')

    balancesht.fillna(0, inplace=True)
    balancesht.index = balancesht.index.str.strip()
    balancesht = balancesht.transpose()
    balancesht['WORKING CAPITAL'] = balancesht['OTHER ASSETS'] - balancesht['OTHER LIABILITIES']
    balancesht['DEBTOR DAYS'] = np.where(pnl['SALES'] > 0, balancesht['RECEIVABLES'] / (pnl['SALES'] / 365), 0)
    balancesht['INVENTORY TURNOVER'] = np.where(balancesht['INVENTORY'] > 0, pnl['SALES'] / balancesht['INVENTORY'],
                                                0)
    balancesht['ROCE'] = np.where(balancesht['NET BLOCK'] + balancesht['WORKING CAPITAL'] > 0, (
                (pnl['OPERATING PROFIT'] - pnl['DEPRECIATION'] - pnl['TAX']) / (
                    balancesht['NET BLOCK'] + balancesht['WORKING CAPITAL'])) * 100, 0)
    pnl = pnl.transpose()
    pnl = pnl.round(1)
    #pnl.columns = pnl.columns.strftime('%d-%m-%Y')
    # st.dataframe(pnl)
    balancesht = balancesht.transpose()
    balancesht = balancesht.round(1)
    # st.dataframe(balancesht)


    cashflow = df_comp.loc['CASH FLOW']
    cashflow.fillna(0, inplace=True)
    cashflow.index = cashflow.index.str.strip()

    try:
        pnl.columns = pd.to_datetime(pnl.columns,'%d-%m-%Y')
        balancesht.columns = pd.to_datetime(balancesht.columns,'%d-%m-%Y')
        cashflow.columns = pd.to_datetime(cashflow.columns,'%d-%m-%Y')
    except Exception as ValueError:
        pnl.columns = pd.to_datetime(pnl.columns,'%Y-%m-%d')
        balancesht.columns = pd.to_datetime(balancesht.columns,'%Y-%m-%d')
        cashflow.columns = pd.to_datetime(cashflow.columns,'%Y-%m-%d')
        
    # st.dataframe(pnl)
    # st.dataframe(balancesht)
    # st.dataframe(cashflow)
    return pnl, balancesht, cashflow

def stmt_for_qoq(df):
    # THE FOLLOWIGN CODE CALCULATES THE GROWTH OR DEGROWTH
    df.columns = pd.to_datetime(df.columns, format='%d-%m-%Y')
    last_quarter = df.columns[-1]
    sentence = f"{datetime.datetime.strftime(last_quarter,'%d-%b-%Y')} #Result\n"
    # Create a list of metrics
    metrics = ['SALES', 'OPERATING PROFIT', 'NET PROFIT']
    # qtr_string = last_quarter.strptime(last_quarter,"%b%Y")
    if len(df.columns)>=5:
        prev_quarter = df.columns[-2]
        yoy_quarter = df.columns[-5]
        # Loop through the metrics
        for metric in metrics:
            # Retrieve the QoQ percentage change value for the last quarter
            value = df.loc[metric, last_quarter]
            value_qoq = df.loc[metric, prev_quarter]
            value_yoy = df.loc[metric, yoy_quarter]
            yoy_growth = ((value - value_yoy)/value_yoy)*100
            qoq_growth = df.at[metric + '_QoQ', last_quarter]
            if not pd.isna(qoq_growth) and not pd.isna(value):
                qoq_trend = " up by " if qoq_growth > 0 else " down by "
                yoy_trend = " up by " if yoy_growth > 0 else " down by "
                #sentence += f"{metric}: {value:.2f}cr vs {value_qoq}cr,{trend}{qoq_growth:.2f}% QoQ\n"
                #sentence += f"{metric}: {value:.2f}cr,{qoq_trend}{qoq_growth:.2f}% QoQ & {yoy_trend}{yoy_growth:.2f}% YoY\n"
                sentence += f"{value:.2f}cr, QoQ {qoq_growth:.2f}% & YoY {yoy_growth:.2f}% {metric};\n"

        # st.info(type(last_quarter))
        for metric in ['OPM %', 'NPM %']:
            value = df.loc[metric, last_quarter]
            if not pd.isna(value):
                #abs_value = abs(value)
                sentence += f"{metric}: {value:.2f}% vs {(df.loc[metric, prev_quarter])}%\n"
        #sentence += f"For more: https://www.instagram.com/itimesalgo/\nAnalyse your favourite company using https://www.itimesalgo.streamlit.app\nhttps://t.me/itimesalgo"
        # sentence += f"OPM {pnl.loc['OPM %', last_quarter]}% vs {pnl.loc['OPM %', prev_quarter]}%\n"
        # sentence += f"NPM {pnl.loc['NPM %', last_quarter]}% vs {pnl.loc['NPM %', prev_quarter]}%\n"
    elif len(df.columns)>=2:
        last_quarter = df.columns[-1]
        prev_quarter = df.columns[-2]
        # Loop through the metrics
        for metric in metrics:
            # Retrieve the QoQ percentage change value for the last quarter
            value = df.loc[metric, last_quarter]
            value_qoq = df.loc[metric, prev_quarter]
            qoq_growth = df.at[metric + '_QoQ', last_quarter]
            if not pd.isna(qoq_growth) and not pd.isna(value):
                qoq_trend = " up by " if qoq_growth > 0 else " down by "
                sentence += f"{value:.2f}cr, QoQ {qoq_growth:.2f}% {metric};\n"

        # st.info(type(last_quarter))
        for metric in ['OPM %', 'NPM %']:
            value = df.loc[metric, last_quarter]
            if not pd.isna(value):
                sentence += f"{metric}: {value:.2f}% vs {(df.loc[metric, prev_quarter])}%\n"
        #sentence += f"For more: https://www.instagram.com/itimesalgo/\nAnalyse your favourite company using https://www.itimesalgo.streamlit.app\nhttps://t.me/itimesalgo"
        # sentence += f"OPM {pnl.loc['OPM %', last_quarter]}% vs {pnl.loc['OPM %', prev_quarter]}%\n"
        # sentence += f"NPM {pnl.loc['NPM %', last_quarter]}% vs {pnl.loc['NPM %', prev_quarter]}%\n"
    else:
        last_quarter = df.columns[-1]
        sales_q = df.loc["SALES",last_quarter]
        Net_Profit_q = df.loc["NET PROFIT",last_quarter]
        sentence += f"{sales_q}\n{Net_Profit_q}"
    return sentence

def analyse_df(pnl,balancesht,qtr_pnl):
    telegram_update = False
    # st.dataframe(pnl)
    # st.dataframe(balancesht)
    # st.dataframe(qtr_pnl)
    send_metadata = {}
    send_metadata['telegram'] = []
    send_metadata['BalanceSheet_Statement'] = ""
    send_metadata['YPNL_Statement'] = ""
    send_metadata['pros'] = []
    send_metadata['cons'] = []
    send_metadata['statement'] = []
    send_metadata['BALANCE_SHEET'] = ""
    send_metadata['tags'] = []
    #make columns to datetime
    qtr_pnl.columns = pd.to_datetime(qtr_pnl.columns, format='%Y-%m-%d')

    if not pnl.empty and not balancesht.empty and pnl is not None and balancesht is not None and isinstance(pnl, pd.DataFrame) and isinstance(balancesht, pd.DataFrame):

        if not qtr_pnl.empty and qtr_pnl is not None and isinstance(qtr_pnl,pd.DataFrame) :                                                                                                                       # DEFINING THE TABLE CONTENTS AS OBJECTS
            send_metadata['QPNL_Statement'] = ""
            send_metadata['updated_results_on'] = datetime.datetime.now()
            QPNL_CQ = qtr_pnl.columns[-1]
            send_metadata['recent_quarter'] = QPNL_CQ
            if len(balancesht.columns)>3 :
                # YEARLY ANALYSIS
                BS_CY = balancesht.columns[-1]
                BS_2Y = balancesht.columns[-2]
                BS_3Y = balancesht.columns[-3]
                show_BS_CY = str(datetime.datetime.strftime(BS_CY, '%b-%Y'))
                show_BS_2Y = str(datetime.datetime.strftime(BS_2Y, '%b-%Y'))
                show_BS_3Y = str(datetime.datetime.strftime(BS_3Y, '%b-%Y'))

                eq_BS_CY = round(balancesht.loc["NO. OF EQUITY SHARES", BS_CY] / 10000000, 1)
                eq_BS_2Y = round(balancesht.loc['NO. OF EQUITY SHARES', BS_2Y] / 10000000, 1)
                FV_BS_CY = round(balancesht.loc['FACE VALUE', BS_CY])
                FV_BS_2Y = round(balancesht.loc['FACE VALUE', BS_2Y])
                ROCE_BS_CY = balancesht.loc['ROCE', BS_CY]
                ROCE_BS_2Y = balancesht.loc['ROCE', BS_2Y]
                #send_metadata['BalanceSheet_Statement'] = f"\n*****\nEquity in {str(datetime.datetime.strftime(BS_CY, '%b-%Y'))}: {str(eq_BS_CY)}cr; in {str(datetime.datetime.strftime(BS_2Y, '%b-%Y'))}: {str(eq_BS_2Y)}cr\nFaceValue in {str(datetime.datetime.strftime(BS_CY, '%b-%Y'))}: {str(FV_BS_CY)}; in {str(datetime.datetime.strftime(BS_2Y, '%b-%Y'))}: {str(FV_BS_2Y)}\nROCE in {str(datetime.datetime.strftime(BS_CY, '%b-%Y'))}: {str(ROCE_BS_CY)}%; in {str(datetime.datetime.strftime(BS_2Y, '%b-%Y'))}: {str(ROCE_BS_2Y)}%\n"

                Reserves_CY = balancesht.loc['RESERVES', BS_CY]
                Reserves_2Y = balancesht.loc['RESERVES', BS_2Y]
                Reserves_3Y = balancesht.loc['RESERVES', BS_3Y]
                Borrowings_CY = balancesht.loc['BORROWINGS', BS_CY]
                Borrowings_2Y = balancesht.loc['BORROWINGS', BS_2Y]
                Borrowings_3Y = balancesht.loc['BORROWINGS', BS_3Y]
            if len(pnl.columns)>3:
                PNL_CY = pnl.columns[-1]
                PNL_2Y = pnl.columns[-2]
                PNL_3Y = pnl.columns[-3]

                YSALES_Y = pnl.loc['SALES', PNL_CY]
                YSALES_2Y = pnl.loc['SALES', PNL_2Y]
                YSALES_3Y = pnl.loc['SALES', PNL_3Y]

                YSALES_QoQ_Y = pnl.loc['SALES_QoQ', PNL_CY]
                YSALES_QoQ_2Y = pnl.loc['SALES_QoQ', PNL_2Y]
                YSALES_QoQ_3Y = pnl.loc['SALES_QoQ', PNL_3Y]

                YPROFIT_QoQ_Y = pnl.loc['NET PROFIT_QoQ', PNL_CY]
                YPROFIT_QoQ_2Y = pnl.loc['NET PROFIT_QoQ', PNL_2Y]
                YPROFIT_QoQ_3Y = pnl.loc['NET PROFIT_QoQ', PNL_3Y]
                
                YPROFIT_Y = pnl.loc['NET PROFIT', PNL_CY]
                YPROFIT_2Y = pnl.loc['NET PROFIT', PNL_2Y]
                YPROFIT_3Y = pnl.loc['NET PROFIT', PNL_3Y]

                YNPM_CY = pnl.loc['NPM %', PNL_CY]
                YNPM_2Y = pnl.loc['NPM %', PNL_2Y]
                YNPM_3Y = pnl.loc['NPM %', PNL_3Y]
            if len(qtr_pnl.columns)>2:
                # QUARTERLY ANALYSIS
                QPNL_CQ = qtr_pnl.columns[-1]
                QPNL_2Q = qtr_pnl.columns[-2]
                QPNL_3Q = qtr_pnl.columns[-3]
                show_QPNL_CQ = str(datetime.datetime.strftime(QPNL_CQ, '%b-%Y'))
                show_QPNL_2Q = str(datetime.datetime.strftime(QPNL_2Q, '%b-%Y'))
                show_QPNL_3Q = str(datetime.datetime.strftime(QPNL_3Q, '%b-%Y'))

                QSALES_CQ = qtr_pnl.loc['SALES', QPNL_CQ]
                QSALES_2Q = qtr_pnl.loc['SALES', QPNL_2Q]
                QSALES_3Q = qtr_pnl.loc['SALES', QPNL_3Q]

                QPROFIT_CQ = qtr_pnl.loc['NET PROFIT', QPNL_CQ]
                QPROFIT_2Q = qtr_pnl.loc['NET PROFIT', QPNL_2Q]
                QPROFIT_3Q = qtr_pnl.loc['NET PROFIT', QPNL_3Q]

                QOPM_CQ = qtr_pnl.loc['OPM %',QPNL_CQ]
                QOPM_2Q = qtr_pnl.loc['OPM %',QPNL_2Q]
                QOPM_3Q = qtr_pnl.loc['OPM %',QPNL_3Q]

                QNPM_CQ = qtr_pnl.loc['NPM %', QPNL_CQ]
                QNPM_2Q = qtr_pnl.loc['NPM %', QPNL_2Q]
                QNPM_3Q = qtr_pnl.loc['NPM %', QPNL_3Q]

                                                                                                                                # ANALYSING BALANCESHEET

            if len(balancesht.columns) > 3:
                if Borrowings_CY<0.03*Reserves_CY :
                    if Borrowings_2Y<0.03*Reserves_2Y:
                        if Borrowings_3Y<0.03*Reserves_3Y:
                            send_metadata['pros'].append("DEBT-Free Company for the last 3 yrs")
                        else:
                            send_metadata['pros'].append("A DEBT-Free Company for the last 2 yrs")
                    else:
                        send_metadata['pros'].append("A DEBT-Free Company")
                if Borrowings_CY<0.85*Borrowings_2Y and Borrowings_2Y < 0.85*Borrowings_3Y:
                    borr_trend = " Declining by <15% for the last 3yrs"
                elif Borrowings_CY<Borrowings_2Y<Borrowings_3Y:
                    borr_trend = " Declining for last 3yrs"
                elif Borrowings_CY>Borrowings_2Y>Borrowings_3Y:
                    borr_trend = " increasing for the last 3yrs"
                elif Borrowings_CY>Borrowings_2Y<Borrowings_3Y:
                    borr_trend = " have increased this year"
                else:
                    borr_trend = f"are at {Borrowings_CY}"

                # EXPANSION
                CWIP_Y = balancesht.loc["CAPITAL WORK IN PROGRESS", BS_CY]
                CWIP_2Y = balancesht.loc["CAPITAL WORK IN PROGRESS", BS_2Y]
                CWIP_3Y = balancesht.loc["CAPITAL WORK IN PROGRESS", BS_3Y]

                NETBLOCK_Y = balancesht.loc["NET BLOCK", BS_CY]
                NETBLOCK_2Y = balancesht.loc["NET BLOCK", BS_2Y]
                NETBLOCK_3Y = balancesht.loc["NET BLOCK", BS_3Y]

                DEBTORDAYS_Y = balancesht.loc['DEBTOR DAYS', BS_CY]
                DEBTORDAYS_2Y = balancesht.loc['DEBTOR DAYS', BS_2Y]
                DEBTORDAYS_3Y = balancesht.loc['DEBTOR DAYS', BS_3Y]

                INVENTORYTURNOVER_Y = balancesht.loc['INVENTORY TURNOVER', BS_CY]
                INVENTORYTURNOVER_2Y = balancesht.loc['INVENTORY TURNOVER', BS_2Y]
                INVENTORYTURNOVER_3Y = balancesht.loc['INVENTORY TURNOVER', BS_3Y]

                Expansion = "EXPANSION ??"
                if CWIP_Y > NETBLOCK_2Y and CWIP_2Y > NETBLOCK_3Y:
                    Expansion += f"- EXPANSION in the last 2 consequent years. Check if it realised SALES this year"
                    send_metadata['tags'].append(f"EXPANSION {datetime.datetime.strftime(balancesht.columns[-2],'%Y')}")
                    send_metadata['tags'].append(f"EXPANSION {datetime.datetime.strftime(balancesht.columns[-1],'%Y')}")

                else:
                    if CWIP_Y > NETBLOCK_2Y:
                        Expansion += f"- EXPANSION underway in {show_BS_CY}, Better to Check the next Quarterly Results to see if the expansion is being effectively Utilised"
                        send_metadata['tags'].append(f"EXPANSION {datetime.datetime.strftime(balancesht.columns[-1],'%Y')}")
                    if CWIP_2Y > NETBLOCK_3Y and NETBLOCK_Y > NETBLOCK_2Y and CWIP_Y < NETBLOCK_2Y:
                        Expansion += f"- There was an Expansion in the {show_BS_2Y}, so got to check SALES too"
                        send_metadata['tags'].append(f"EXPANSION {datetime.datetime.strftime(balancesht.columns[-2],'%Y')}")
                if Expansion != "EXPANSION ??":
                    send_metadata['pros'].append(Expansion)

                    # with st.expander("UNDERGOING EXPANSION"):
                    #     balancesht1 = balancesht.transpose()
                    #     balancesht2 = balancesht1[
                    #         ['NET BLOCK', 'CAPITAL WORK IN PROGRESS', 'DEBTOR DAYS', 'INVENTORY TURNOVER', 'RESERVES',
                    #          'BORROWINGS']]
                    #     st.dataframe(balancesht2.transpose())
                    #     pnl1 = pnl.transpose()
                    #     pnl2 = pnl1[['SALES', 'NET PROFIT', 'OPM %', 'NPM %']]
                    #     st.dataframe(pnl2.transpose())

                if DEBTORDAYS_Y < DEBTORDAYS_2Y < DEBTORDAYS_3Y:
                    Product_Demand = "- PRODUCT DEMAND is rising since 3 years"
                elif DEBTORDAYS_Y < DEBTORDAYS_2Y:
                    Product_Demand = "- PRODUCT DEMAND is rising since 3 years"
                elif DEBTORDAYS_Y > DEBTORDAYS_2Y > DEBTORDAYS_3Y:
                    Product_Demand = "- NO PRODUCT DEMAND"
                else :
                    Product_Demand = ""

                if Product_Demand == "- NO PRODUCT DEMAND":
                    send_metadata['cons'].append(Product_Demand)
                elif Product_Demand != "":
                    send_metadata['pros'].append(Product_Demand)
                    send_metadata['tags'].append("PRODUCT DEMAND")

                Growth_metric = 15
                if ROCE_BS_CY > Growth_metric:
                    if ROCE_BS_2Y > Growth_metric:
                        YROCE_A = f"- ROCE IS GROWING AT MORE THAN {Growth_metric}% SINCE 2 YEARS.\nROCE in {datetime.datetime.strftime(BS_CY,'%b%Y')} is at {ROCE_BS_CY}%"
                        send_metadata['pros'].append(YROCE_A)
                    else:
                        YROCE_A = f"- ROCE is at {ROCE_BS_CY}% in {datetime.datetime.strftime(BS_CY,'%b%Y')}."

                if Reserves_CY > Reserves_2Y > Reserves_3Y and ((Borrowings_CY<Borrowings_2Y<Borrowings_3Y) or Borrowings_CY<1 ):
                    send_metadata['pros'].append(f"- RESERVES are increasing at above 15% in the last 3 years with Borrowings {borr_trend}")

            # CHECKING 9M with the Current Year
            if str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b')).upper() == 'DEC' and len(qtr_pnl.columns)>3 and len(pnl.columns)>3:
                PNL_CY = pnl.columns[-1]
                YSALES_Y = pnl.loc['SALES', PNL_CY]
                YPROFIT_Y = pnl.loc['NET PROFIT', PNL_CY]

                last_9m_QSALES_A = ""
                last_9m_QSALES = qtr_pnl.loc['SALES',qtr_pnl.columns[-1]] + qtr_pnl.loc['SALES',qtr_pnl.columns[-2]] + qtr_pnl.loc['SALES',qtr_pnl.columns[-3]]
                if last_9m_QSALES > 0.5*YSALES_Y:
                    if YSALES_Y>0:
                        Qsales_9m_up_by = round((last_9m_QSALES/YSALES_Y)*100)
                        last_9m_QSALES_A = f"- Last 3QSALES is about {Qsales_9m_up_by}% YSales"
                        send_metadata['pros'].append(last_9m_QSALES_A)
                last_9m_QPROFIT_A = ""
                last_9m_QPROFIT = qtr_pnl.loc['NET PROFIT',qtr_pnl.columns[-1]] + qtr_pnl.loc['NET PROFIT',qtr_pnl.columns[-2]] + qtr_pnl.loc['NET PROFIT',qtr_pnl.columns[-3]]

                if qtr_pnl.loc['SALES',qtr_pnl.columns[-1]]>1.5*qtr_pnl.loc['SALES',qtr_pnl.columns[-2]] and qtr_pnl.loc['NET PROFIT',qtr_pnl.columns[-1]]>1.5*qtr_pnl.loc['NET PROFIT',qtr_pnl.columns[-2]]:
                    if YPROFIT_Y>0 :
                        Double_QSP = f"- Double QSales and QProfit"
                        send_metadata['pros'].append(Double_QSP)
                        # telegram_update = True
                        send_metadata['telegram'].append('BestQ')
                        send_metadata['tags'].append(f"BestQ {str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b %Y'))}")

                if last_9m_QPROFIT > 0.5 * YPROFIT_Y:
                    if YPROFIT_Y>0 :
                        QProfit_9m_up_by = round((last_9m_QPROFIT/YPROFIT_Y)*100)
                        last_9m_QPROFIT_A = f"- Last 3QPROFIT is about {QProfit_9m_up_by}% of YProfit"
                        send_metadata['pros'].append(last_9m_QPROFIT_A)
                        send_metadata['telegram'].append('BestQ')
                        send_metadata['tags'].append(f"BestQ {str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b %Y'))}")

            
            # CHECKS IF NET PROFIT IS NEGATIVE
            if len(pnl.columns)>=1 :
                PNL_CY = pnl.columns[-1]
                YPROFIT_Y = pnl.loc['NET PROFIT', PNL_CY]

                if YPROFIT_Y < 0:
                    send_metadata['cons'].append(f"- NEGATIVE YProfit : {YPROFIT_Y}")
                # CHECKS IF NET PROFIT TURNED POSITIVE FROM NEGATIVE RECENTLY
                if len(pnl.columns)>=2:
                    PNL_2Y = pnl.columns[-2]
                    YPROFIT_2Y = pnl.loc['NET PROFIT', PNL_2Y]
                    if YPROFIT_2Y < 0 and YPROFIT_Y > 0:
                        send_metadata['pros'].append(f"- YProfit Turned POSITIVE this year {datetime.datetime.strftime(PNL_CY,'%b-%Y')} : {YPROFIT_Y}")
            if len(pnl.columns) > 5:
                if pnl.loc['NET PROFIT', pnl.columns[-5]] < 0:
                    if pnl.loc['NET PROFIT', pnl.columns[-4]] > 0 and pnl.loc['NET PROFIT', pnl.columns[-3]]>0 and pnl.loc['NET PROFIT', pnl.columns[-2]]>0 and pnl.loc['NET PROFIT', pnl.columns[-1]]>0:
                        send_metadata['pros'].append(f"- YPROFIT TURNED POSITIVE 4yrs back, and maintaining ever since")
                if pnl.loc['NET PROFIT', pnl.columns[-4]] < 0:
                    if pnl.loc['NET PROFIT', pnl.columns[-3]] > 0 and pnl.loc['NET PROFIT', pnl.columns[-2]] > 0 and \
                            pnl.loc['NET PROFIT', pnl.columns[-1]] > 0:
                        send_metadata['pros'].append(f"- YPROFIT TURNED POSITIVE 3yrs back, and maintaining ever since")

                        # send_metadata['pros'].append(
                        #     f"- TURNED POSITIVE : YProfit was negative in the past, but company showing NET POSITVE PROFITS since last 3 years")
                if pnl.loc['NET PROFIT', pnl.columns[-3]] < 0:
                    if pnl.loc['NET PROFIT', pnl.columns[-2]] > 0 and pnl.loc['NET PROFIT', pnl.columns[-1]] > 0 :
                        send_metadata['pros'].append(f"- YPROFIT TURNED POSITIVE 2yrs back, and maintaining ever since")
                        # send_metadata['pros'].append(
                        #     f"- TURNED POSITIVE : YProfit was negative in the past, but company showing NET POSITVE PROFITS since last 2 years")
                # if YPROFIT_3Y < 0:
                #     if YPROFIT_2Y > 0 and YPROFIT_Y>0:
                #         send_metadata['pros'].append(f"- TURNED POSITIVE : YProfit was negative in the past, but company showing NET POSITVE PROFITS since last 2 years")
                                                                                                                            # ANALYSING YEARLY PNL
            
            if str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b')).upper() == 'SEP'  and len(qtr_pnl.columns)>1 and len(pnl.columns)>1 :
                # st.success(f"Len of qtr_pnl : {len(qtr_pnl)}")
                # st.dataframe(qtr_pnl)
                PNL_CY = pnl.columns[-1]
                PNL_2Y = pnl.columns[-2]
                YSALES_Y = pnl.loc['SALES', PNL_CY]
                YSALES_2Y = pnl.loc['SALES', PNL_2Y]                
                YPROFIT_2Y = pnl.loc['NET PROFIT', PNL_2Y]              
                YPROFIT_Y = pnl.loc['NET PROFIT', PNL_CY]

                last_6m_QSALES_A = ""
                last_6m_QSALES = qtr_pnl.loc['SALES',qtr_pnl.columns[-1]] + qtr_pnl.loc['SALES',qtr_pnl.columns[-2]]
                if last_6m_QSALES > 0.5*YSALES_Y:
                    if YSALES_Y>0:
                        Qsales_6m_up_by = round((last_6m_QSALES/YSALES_Y)*100)
                        last_6m_QSALES_A = f"- Last 2QSALES is about {Qsales_6m_up_by}% YSales"
                        send_metadata['pros'].append(last_6m_QSALES_A)
                last_6m_QPROFIT_A = ""
                last_6m_QPROFIT = qtr_pnl.loc['NET PROFIT',qtr_pnl.columns[-1]] + qtr_pnl.loc['NET PROFIT',qtr_pnl.columns[-2]]
                if qtr_pnl.loc['SALES',qtr_pnl.columns[-1]]>1.5*qtr_pnl.loc['SALES',qtr_pnl.columns[-2]] and qtr_pnl.loc['NET PROFIT',qtr_pnl.columns[-1]]>1.5*qtr_pnl.loc['NET PROFIT',qtr_pnl.columns[-2]]:
                    if YPROFIT_Y>0 :
                        Double_QSP = f"- QSales and QProfit doubled"
                        send_metadata['pros'].append(Double_QSP)
                        # telegram_update = True
                        send_metadata['telegram'].append('BestQ')
                        send_metadata['tags'].append(f"BestQ {str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b %Y'))}")

                if last_6m_QPROFIT > 0.5 * YPROFIT_Y:
                    if YPROFIT_Y>0 :
                        QProfit_6m_up_by = round((last_6m_QPROFIT/YPROFIT_Y)*100)
                        last_6m_QPROFIT_A = f"- Last 2QPROFIT is about {QProfit_6m_up_by}% of YProfit"
                        send_metadata['pros'].append(last_6m_QPROFIT_A)
                        # telegram_update = True
                        send_metadata['telegram'].append('BestQ')
                        send_metadata['tags'].append(f"BestQ {str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b %Y'))}")
                        
                # # CHECKS IF NET PROFIT IS NEGATIVE
                # if YPROFIT_Y < 0:
                #     send_metadata['cons'].append(f"- NEGATIVE YProfit : {YPROFIT_Y}")
                # # CHECKS IF NET PROFIT TURNED POSITIVE FROM NEGATIVE RECENTLY
                # if YPROFIT_2Y < 0:
                #     if YPROFIT_Y > 0:
                #         send_metadata['pros'].append(f"- YProfit Turned POSITIVE this year {datetime.datetime.strftime(PNL_CY,'%b-%Y')} : {YPROFIT_Y}")

                # if len(pnl.columns) > 5:
                #     if pnl.loc['NET PROFIT', pnl.columns[-5]] < 0:
                #         if pnl.loc['NET PROFIT', pnl.columns[-4]] > 0 and pnl.loc['NET PROFIT', pnl.columns[-3]]>0 and pnl.loc['NET PROFIT', pnl.columns[-2]]>0 and pnl.loc['NET PROFIT', pnl.columns[-1]]>0:
                #             send_metadata['pros'].append(f"- YPROFIT TURNED POSITIVE 4yrs back, and maintaining ever since")
                #     if pnl.loc['NET PROFIT', pnl.columns[-4]] < 0:
                #         if pnl.loc['NET PROFIT', pnl.columns[-3]] > 0 and pnl.loc['NET PROFIT', pnl.columns[-2]] > 0 and \
                #                 pnl.loc['NET PROFIT', pnl.columns[-1]] > 0:
                #             send_metadata['pros'].append(f"- YPROFIT TURNED POSITIVE 3yrs back, and maintaining ever since")

                #             # send_metadata['pros'].append(
                #             #     f"- TURNED POSITIVE : YProfit was negative in the past, but company showing NET POSITVE PROFITS since last 3 years")
                #     if pnl.loc['NET PROFIT', pnl.columns[-3]] < 0:
                #         if pnl.loc['NET PROFIT', pnl.columns[-2]] > 0 and pnl.loc['NET PROFIT', pnl.columns[-1]] > 0 :
                #             send_metadata['pros'].append(f"- YPROFIT TURNED POSITIVE 2yrs back, and maintaining ever since")
                #             # send_metadata['pros'].append(
                #             #     f"- TURNED POSITIVE : YProfit was negative in the past, but company showing NET POSITVE PROFITS since last 2 years")
                #     # if YPROFIT_3Y < 0:
                #     #     if YPROFIT_2Y > 0 and YPROFIT_Y>0:
                #     #         send_metadata['pros'].append(f"- TURNED POSITIVE : YProfit was negative in the past, but company showing NET POSITVE PROFITS since last 2 years")
                #                                                                                                                 # ANALYSING YEARLY PNL

            if len(pnl.columns)>3:
                # PROFIT OR SALES DOUBLED IN THE LAST 3 YEARS
                if len(pnl.columns) > 3:# and len(qtr_pnl.columns)>3:
                    if YSALES_Y>2*YSALES_2Y and YPROFIT_Y>2*YPROFIT_2Y:
                        send_metadata['pros'].append(f"- YProfits and YSales DOUBLED in {datetime.datetime.strftime(PNL_CY, '%b%Y')}")
                        send_metadata['tags'].append(f"YSD YPD {str(datetime.datetime.strftime(pnl.columns[-1], '%Y'))}")
                    if YSALES_2Y>2*YSALES_3Y and YPROFIT_2Y>2*YPROFIT_3Y:
                        send_metadata['pros'].append(f"- YProfits and YSales DOUBLED in {datetime.datetime.strftime(PNL_2Y, '%b%Y')}")
                        send_metadata['tags'].append(f"YSD YPD {str(datetime.datetime.strftime(pnl.columns[-2], '%Y'))}")
                    if len(pnl.columns) > 4: #and len(qtr_pnl.columns) > 4:
                        if YSALES_3Y > 2 * pnl.loc['SALES', pnl.columns[-4]] and YPROFIT_3Y > 2 * pnl.loc['NET PROFIT', pnl.columns[-4]]:
                            send_metadata['pros'].append(f"- YProfits and YSales DOUBLED in {datetime.datetime.strftime(PNL_3Y, '%b%Y')}")
                            send_metadata['tags'].append(f"YSD YPD {str(datetime.datetime.strftime(pnl.columns[-3], '%Y'))}")
                    # CHECKS IF LAST YEAR PROFIT DOUBLED and THIS YEAR MAINTAINING
                    if YPROFIT_2Y > 2*YPROFIT_3Y and YPROFIT_Y > YPROFIT_2Y:
                        send_metadata['pros'].append(f"- LastYr YProfits Doubled and maintained in {datetime.datetime.strftime(PNL_CY, '%b%Y')}")
                        send_metadata['tags'].append(f"LYPD PM {str(datetime.datetime.strftime(pnl.columns[-1], '%Y'))}")

                # NPM IMPROVED BY 15%               # CHECK IF NET PROFITS ARE MAINTAINED EVER SINCE
                Ymargins_A = ""
                margins_metric = 15
                if YNPM_3Y > margins_metric:
                    if YPROFIT_2Y > YPROFIT_3Y and YPROFIT_Y > YPROFIT_2Y:
                        Ymargins_A = f"- YMargins improvised by more than {margins_metric}% since {datetime.datetime.strftime(PNL_3Y,'%b%Y')} and PROFITS been maintaining ever since"
                elif YNPM_2Y > margins_metric:
                    if YPROFIT_Y > YPROFIT_2Y:
                        Ymargins_A = f"- YMargins improvised by more than {margins_metric}% since {datetime.datetime.strftime(PNL_2Y,'%b%Y')} and PROFITS been maintaining ever since."
                else:
                    Ymargins_A = "- YMARGINS LESS THAN 15%"
                if Ymargins_A != "- YMARGINS LESS THAN 15%":
                    send_metadata['pros'].append(Ymargins_A)

                # CHECKS if ROCE SALES PROFIT > 15%
                Growth_metric = 15
                YTopline_Bottomline_A = ""
                if YSALES_QoQ_Y > Growth_metric and YPROFIT_QoQ_Y > Growth_metric:
                    if YSALES_QoQ_2Y > Growth_metric and YPROFIT_QoQ_2Y > Growth_metric:
                        YTopline_Bottomline_A = f"- QoQ_GROWTH IN YSALES YPROFIT IS MORE THAN 15%"
                        if YPROFIT_QoQ_Y > YSALES_QoQ_Y:
                            if YSALES_QoQ_Y > 15 and YSALES_QoQ_Y < 90 and YPROFIT_QoQ_Y>100:
                                YTopline_Bottomline_A += f"- YPROFIT DOUBLED MARGINS IMPROVED\nAs on {datetime.datetime.strftime(PNL_CY, '%b%Y')}, Sales have increased by {YSALES_QoQ_Y}% while the Profit QoQ_growth is much better than Sales itself at {YPROFIT_QoQ_Y}%"
                                send_metadata['tags'].append(f"YS15 YPD {str(datetime.datetime.strftime(pnl.columns[-1], '%Y'))}")
                            else:
                                YTopline_Bottomline_A += f"- YPROFIT GROWTH IS BETTER THAN YSALES GROWTH\nAs on {datetime.datetime.strftime(PNL_CY, '%b%Y')}, Sales have increased by {YSALES_QoQ_Y}% while the Profit QoQ_growth is much better than Sales itself at {YPROFIT_QoQ_Y}%"
                                send_metadata['tags'].append(f"PROFIT GROWTH BETTER THAN SALES GROWTH {str(datetime.datetime.strftime(pnl.columns[-1], '%Y'))}")
                        else:
                            YTopline_Bottomline_A += f"\nAs on {datetime.datetime.strftime(PNL_CY,'%b%Y')}, Sales have increased by {YSALES_QoQ_Y}% while the Profit QoQ_growth is at {YPROFIT_QoQ_Y}%"
                    else:
                        YTopline_Bottomline_A = f"\nYSALES in {datetime.datetime.strftime(PNL_CY,'%b-%Y')} is at {pnl.loc['SALES',PNL_CY]} " \
                                                f"showing a QoQ_growth of {YSALES_QoQ_Y}%. The YNETPROFIT grew by {YPROFIT_QoQ_Y}% to {YPROFIT_Y}"
                    send_metadata['pros'].append(YTopline_Bottomline_A)

                    
                # if YSALES_QoQ_Y > YSALES_QoQ_2Y and YPROFIT_QoQ_Y > YPROFIT_QoQ_2Y:
                #     YTopline_Bottomline_A = f"- QoQ_Growth in YSALES YPROFIT is GOOD.\nYSALES in {datetime.datetime.strftime(PNL_CY,'%b-%Y')} is at {pnl.loc['SALES',PNL_CY]} " \
                #                    f"showing a QoQ_growth of {YSALES_QoQ_Y}%. The YNETPROFIT grew by {YPROFIT_QoQ_Y}% to {YPROFIT_Y}"
                #     send_metadata['pros'].append(YTopline_Bottomline_A)

                                                                                                                                # SALES AND PROFIT is more than PREV_YEAR
                # if YSALES_Y > YSALES_2Y and YPROFIT_Y > YPROFIT_2Y:
                #     YPNL_2Y_A = f"- RISING YSALES & YPROFIT : \nYSALES QoQ_growth in {PNL_CY} is {YSALES_QoQ_Y}%. The YNETPROFIT QoQ_growth is {YPROFIT_QoQ_Y}%"
                #     send_metadata['pros'].append(YPNL_2Y_A)
                if YPROFIT_Y > 0 and YPROFIT_2Y>0:
                    YPNL_A = stmt_for_qoq(pnl)
                    send_metadata['YPNL_tweet'] = f"Yearly \n {YPNL_A}"
            if len(qtr_pnl.columns)==3:
                if QOPM_CQ < QOPM_2Q < QOPM_3Q and QNPM_CQ < QNPM_2Q < QNPM_3Q :
                    if len(pnl.columns)>3:
                        if YNPM_CY < YNPM_2Y:
                            send_metadata['cons'].append(f"- DECLINING QMARGINS, BOTH QOPM AND QNPM")
                    # send_metadata['cons'].append(f"- DECLINING QUARTERLY MARGINS, BOTH OPM AND NPM")
                if QSALES_CQ > QSALES_2Q and QPROFIT_CQ > QPROFIT_2Q:
                    if QPROFIT_CQ>QSALES_CQ: 
                        QPNL_2Q_A = f"- NPM seems to be improvised : The QPROFIT is showing better QoQ_growth at {qtr_pnl.loc['NET PROFIT_QoQ',QPNL_CQ]}% while QSALES QoQ_Growth is at {qtr_pnl.loc['SALES_QoQ',QPNL_CQ]}%"
                    else:
                        QPNL_2Q_A = f"- RISING QSALES & QPROFIT : QSALES QoQ_Growth in {show_QPNL_CQ} is at {qtr_pnl.loc['SALES_QoQ',QPNL_CQ]}%. The QPROFIT QoQ_growth by {qtr_pnl.loc['NET PROFIT_QoQ',QPNL_CQ]}%"
                    send_metadata['pros'].append(QPNL_2Q_A)
                if QPROFIT_CQ>0 and QPROFIT_2Q>0:
                    QPNL_A = stmt_for_qoq(qtr_pnl)
                    send_metadata['QPNL_tweet'] = f"Quarterly \n{QPNL_A}"
            # LETS TRY GETTING THE IDEAL SCRIPTS WHERE - SALES and PROFIT >90% of Prev_Year, Debtor days reducing since last 3 years
            # TO GET THE LAST 5 YRS BALANCE SHEET DATA
            if len(balancesht.columns)>=5:
                bsht_stmt = f"\nBALANCE SHEET: \t{str(datetime.datetime.strftime(balancesht.columns[-5], '%b-%Y'))}\t{str(datetime.datetime.strftime(balancesht.columns[-4], '%b-%Y'))}\t{show_BS_3Y}\t{show_BS_2Y}\t{show_BS_CY}\n"
                bsht_stmt += f"EQUITY(cr): \t{round(balancesht.loc['NO. OF EQUITY SHARES', balancesht.columns[-5]] / 10000000, 1)}\t{round(balancesht.loc['NO. OF EQUITY SHARES', balancesht.columns[-4]] / 10000000, 2)}\t{round(balancesht.loc['NO. OF EQUITY SHARES', balancesht.columns[-3]] / 10000000, 1)}\t{round(balancesht.loc['NO. OF EQUITY SHARES', balancesht.columns[-2]] / 10000000, 1)}\t{round(balancesht.loc['NO. OF EQUITY SHARES', balancesht.columns[-1]] / 10000000, 1)}\n"
                bsht_stmt += f"ROCE: \t{balancesht.loc['ROCE', balancesht.columns[-5]]}%\t{balancesht.loc['ROCE', balancesht.columns[-4]]}%\t{balancesht.loc['ROCE', balancesht.columns[-3]]}%\t{balancesht.loc['ROCE', balancesht.columns[-2]]}%\t{balancesht.loc['ROCE', balancesht.columns[-1]]}%\n"

                bsht_stmt += f"DEBTOR DAYS: \t{balancesht.loc['DEBTOR DAYS', balancesht.columns[-5]]}%\t{balancesht.loc['DEBTOR DAYS', balancesht.columns[-4]]}%\t{balancesht.loc['DEBTOR DAYS', balancesht.columns[-3]]}%\t{balancesht.loc['DEBTOR DAYS', balancesht.columns[-2]]}%\t{balancesht.loc['DEBTOR DAYS', balancesht.columns[-1]]}%\n"
                bsht_stmt += f"INVENTORY TURNOVER: \t{balancesht.loc['INVENTORY TURNOVER', balancesht.columns[-5]]}%\t{balancesht.loc['INVENTORY TURNOVER', balancesht.columns[-4]]}%\t{balancesht.loc['INVENTORY TURNOVER', balancesht.columns[-3]]}%\t{balancesht.loc['INVENTORY TURNOVER', balancesht.columns[-2]]}%\t{balancesht.loc['INVENTORY TURNOVER', balancesht.columns[-1]]}%\n"
                bsht_stmt += f"NET BLOCK: \t{balancesht.loc['NET BLOCK', balancesht.columns[-5]]}%\t{balancesht.loc['NET BLOCK', balancesht.columns[-4]]}%\t{balancesht.loc['NET BLOCK', balancesht.columns[-3]]}%\t{balancesht.loc['NET BLOCK', balancesht.columns[-2]]}%\t{balancesht.loc['NET BLOCK', balancesht.columns[-1]]}%\n"
                bsht_stmt += f"CWIP: \t{balancesht.loc['CAPITAL WORK IN PROGRESS', balancesht.columns[-5]]}%\t{balancesht.loc['CAPITAL WORK IN PROGRESS', balancesht.columns[-4]]}%\t{balancesht.loc['CAPITAL WORK IN PROGRESS', balancesht.columns[-3]]}%\t{balancesht.loc['CAPITAL WORK IN PROGRESS', balancesht.columns[-2]]}%\t{balancesht.loc['CAPITAL WORK IN PROGRESS', balancesht.columns[-1]]}%\n"

                bsht_stmt += f"RESERVES : \t{balancesht.loc['RESERVES', balancesht.columns[-5]]}%\t{balancesht.loc['RESERVES', balancesht.columns[-4]]}%\t{balancesht.loc['RESERVES', balancesht.columns[-3]]}%\t{balancesht.loc['RESERVES', balancesht.columns[-2]]}%\t{balancesht.loc['RESERVES', balancesht.columns[-1]]}%\n"
                bsht_stmt += f"BORROWINGS : \t{balancesht.loc['BORROWINGS', balancesht.columns[-5]]}%\t{balancesht.loc['BORROWINGS', balancesht.columns[-4]]}%\t{balancesht.loc['BORROWINGS', balancesht.columns[-3]]}%\t{balancesht.loc['BORROWINGS', balancesht.columns[-2]]}%\t{balancesht.loc['BORROWINGS', balancesht.columns[-1]]}%\n"
                # bsht_stmt += f"NET CASH FLOW : \t{balancesht.loc['NET CASH FLOW', balancesht.columns[-5]]}%\t{balancesht.loc['NET CASH FLOW', balancesht.columns[-4]]}%\t{balancesht.loc['NET CASH FLOW', balancesht.columns[-3]]}%\t{balancesht.loc['NET CASH FLOW', balancesht.columns[-2]]}%\t{balancesht.loc['NET CASH FLOW', balancesht.columns[-1]]}%\t"
                send_metadata['BALANCE_SHEET_Statement'] = bsht_stmt
            # TO GET THE LAST 5 YRS DATA as TABLE
            if len(pnl.columns)>=5 and len(qtr_pnl.columns)>=5:
                QPnl_stmt = f"\nQPNL: \t{str(datetime.datetime.strftime(qtr_pnl.columns[-5], '%b-%Y'))}\t\t{str(datetime.datetime.strftime(qtr_pnl.columns[-4], '%b-%Y'))}\t{str(datetime.datetime.strftime(qtr_pnl.columns[-3], '%b-%Y'))}\t{str(datetime.datetime.strftime(qtr_pnl.columns[-2], '%b-%Y'))}\t\t{str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b-%Y'))}\n"
                QPnl_stmt += f"QSALES: \t{qtr_pnl.loc['SALES', qtr_pnl.columns[-5]]}\t\t{qtr_pnl.loc['SALES', qtr_pnl.columns[-4]]}\t\t{qtr_pnl.loc['SALES', qtr_pnl.columns[-3]]}\t\t{qtr_pnl.loc['SALES', qtr_pnl.columns[-2]]}\t\t{qtr_pnl.loc['SALES', qtr_pnl.columns[-1]]}\n"
                QPnl_stmt += f"QoQ%: \t{qtr_pnl.loc['SALES_QoQ', qtr_pnl.columns[-5]]}\t\t{qtr_pnl.loc['SALES_QoQ', qtr_pnl.columns[-4]]}\t\t{qtr_pnl.loc['SALES_QoQ', qtr_pnl.columns[-3]]}\t\t{qtr_pnl.loc['SALES_QoQ', qtr_pnl.columns[-2]]}\t\t{qtr_pnl.loc['SALES_QoQ', qtr_pnl.columns[-1]]}\n"
                QPnl_stmt += f"QOPM%: \t{qtr_pnl.loc['OPM %', qtr_pnl.columns[-5]]}\t\t{qtr_pnl.loc['OPM %', qtr_pnl.columns[-4]]}\t\t{qtr_pnl.loc['OPM %', qtr_pnl.columns[-3]]}\t\t{qtr_pnl.loc['OPM %', qtr_pnl.columns[-2]]}\t\t{qtr_pnl.loc['OPM %', qtr_pnl.columns[-1]]}\n"
                QPnl_stmt += f"QNPM%: \t{qtr_pnl.loc['NPM %', qtr_pnl.columns[-5]]}\t\t{qtr_pnl.loc['NPM %', qtr_pnl.columns[-4]]}\t\t{qtr_pnl.loc['NPM %', qtr_pnl.columns[-3]]}\t\t{qtr_pnl.loc['NPM %', qtr_pnl.columns[-2]]}\t\t{qtr_pnl.loc['NPM %', qtr_pnl.columns[-1]]}\n"
                QPnl_stmt += f"QPFT: \t{qtr_pnl.loc['NET PROFIT', qtr_pnl.columns[-5]]}\t\t{qtr_pnl.loc['NET PROFIT', qtr_pnl.columns[-4]]}\t\t{qtr_pnl.loc['NET PROFIT', qtr_pnl.columns[-3]]}\t\t{qtr_pnl.loc['NET PROFIT', qtr_pnl.columns[-2]]}\t\t{qtr_pnl.loc['NET PROFIT', qtr_pnl.columns[-1]]}\n"
                QPnl_stmt += f"PQoQ%: \t{qtr_pnl.loc['NET PROFIT_QoQ', qtr_pnl.columns[-5]]}\t\t{qtr_pnl.loc['NET PROFIT_QoQ', qtr_pnl.columns[-4]]}\t\t{qtr_pnl.loc['NET PROFIT_QoQ', qtr_pnl.columns[-3]]}\t\t{qtr_pnl.loc['NET PROFIT_QoQ', qtr_pnl.columns[-2]]}\t\t{qtr_pnl.loc['NET PROFIT_QoQ', qtr_pnl.columns[-1]]}\n"
                send_metadata['QPNL_Statement']+=QPnl_stmt

                # lets try showing Sum(Quarterly) in Yearly charts
                # First Check whats the last Quarter announced, based on that, we will try checking how many to SUM-UP

                if str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b')).upper() == 'SEP':
                    SALES_Qsum_in_Y = round(qtr_pnl.loc['SALES', qtr_pnl.columns[-1]] + qtr_pnl.loc['SALES', qtr_pnl.columns[-2]],1)
                    SALES_Qsum_QoQ_in_Y = round((SALES_Qsum_in_Y / pnl.loc['SALES', pnl.columns[-1]]) * 100, 1)
                    PROFIT_Qsum_in_Y = round(qtr_pnl.loc['NET PROFIT', qtr_pnl.columns[-1]] + qtr_pnl.loc[
                        'NET PROFIT', qtr_pnl.columns[-2]],1)
                    PROFIT_Qsum_QoQ_in_Y = round((PROFIT_Qsum_in_Y / pnl.loc['NET PROFIT', pnl.columns[-1]]) * 100, 1)

                elif str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b')).upper() == 'DEC':
                    SALES_Qsum_in_Y = round(qtr_pnl.loc['SALES', qtr_pnl.columns[-1]] + qtr_pnl.loc['SALES', qtr_pnl.columns[-2]] + \
                                    qtr_pnl.loc['SALES', qtr_pnl.columns[-3]],1)
                    SALES_Qsum_QoQ_in_Y = round((SALES_Qsum_in_Y / pnl.loc['SALES', pnl.columns[-1]]) * 100, 1)
                    PROFIT_Qsum_in_Y = round(qtr_pnl.loc['NET PROFIT', qtr_pnl.columns[-1]] + qtr_pnl.loc[
                        'NET PROFIT', qtr_pnl.columns[-2]] + qtr_pnl.loc['NET PROFIT', qtr_pnl.columns[-3]],1)
                    PROFIT_Qsum_QoQ_in_Y = round((PROFIT_Qsum_in_Y / pnl.loc['NET PROFIT', pnl.columns[-1]]) * 100, 1)

                if str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b')).upper() == 'SEP' or str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b')).upper() == 'DEC':
                    Pnl_stmt = f"\nYPNL: \t{str(datetime.datetime.strftime(pnl.columns[-5], '%b-%Y'))}\t{str(datetime.datetime.strftime(pnl.columns[-4], '%b-%Y'))}\t{str(datetime.datetime.strftime(pnl.columns[-3], '%b-%Y'))}\t{str(datetime.datetime.strftime(pnl.columns[-2], '%b-%Y'))}\t{str(datetime.datetime.strftime(pnl.columns[-1], '%b-%Y'))}\tQSUM:{datetime.datetime.strftime(qtr_pnl.columns[-1], '%b-%Y')}\n"
                else:
                    Pnl_stmt = f"\nYPNL: \t{str(datetime.datetime.strftime(pnl.columns[-5], '%b-%Y'))}\t{str(datetime.datetime.strftime(pnl.columns[-4], '%b-%Y'))}\t{str(datetime.datetime.strftime(pnl.columns[-3], '%b-%Y'))}\t{str(datetime.datetime.strftime(pnl.columns[-2], '%b-%Y'))}\t{str(datetime.datetime.strftime(pnl.columns[-1], '%b-%Y'))}\n"
                if str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b')).upper() == 'SEP' or str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b')).upper() == 'DEC':
                    Pnl_stmt += f"YSALES: \t{pnl.loc['SALES', pnl.columns[-5]]}\t\t{pnl.loc['SALES', pnl.columns[-4]]}\t\t{pnl.loc['SALES', pnl.columns[-3]]}\t\t{pnl.loc['SALES', pnl.columns[-2]]}\t\t{pnl.loc['SALES', pnl.columns[-1]]}\t\t{SALES_Qsum_in_Y}\n"
                else:
                    Pnl_stmt += f"YSALES: \t{pnl.loc['SALES', pnl.columns[-5]]}\t\t{pnl.loc['SALES', pnl.columns[-4]]}\t\t{pnl.loc['SALES', pnl.columns[-3]]}\t\t{pnl.loc['SALES', pnl.columns[-2]]}\t\t{pnl.loc['SALES', pnl.columns[-1]]}\n"

                if str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b')).upper() == 'SEP' or str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b')).upper() == 'DEC':
                    Pnl_stmt += f"SQoQ%: \t{pnl.loc['SALES_QoQ', pnl.columns[-5]]}\t\t{pnl.loc['SALES_QoQ', pnl.columns[-4]]}\t\t{pnl.loc['SALES_QoQ', pnl.columns[-3]]}\t\t{pnl.loc['SALES_QoQ', pnl.columns[-2]]}\t\t{pnl.loc['SALES_QoQ', pnl.columns[-1]]}\t\t{SALES_Qsum_QoQ_in_Y}\n"
                else:
                    Pnl_stmt += f"SQoQ%: \t{pnl.loc['SALES_QoQ', pnl.columns[-5]]}\t\t{pnl.loc['SALES_QoQ', pnl.columns[-4]]}\t\t{pnl.loc['SALES_QoQ', pnl.columns[-3]]}\t\t{pnl.loc['SALES_QoQ', pnl.columns[-2]]}\t\t{pnl.loc['SALES_QoQ', pnl.columns[-1]]}\n"

                Pnl_stmt += f"YOPM%: \t{pnl.loc['OPM %', pnl.columns[-5]]}\t\t{pnl.loc['OPM %', pnl.columns[-4]]}\t\t{pnl.loc['OPM %', pnl.columns[-3]]}\t\t{pnl.loc['OPM %', pnl.columns[-2]]}\t\t{pnl.loc['OPM %', pnl.columns[-1]]}\n"
                Pnl_stmt += f"YNPM%: \t{pnl.loc['NPM %', pnl.columns[-5]]}\t\t{pnl.loc['NPM %', pnl.columns[-4]]}\t\t{pnl.loc['NPM %', pnl.columns[-3]]}\t\t{pnl.loc['NPM %', pnl.columns[-2]]}\t\t{pnl.loc['NPM %', pnl.columns[-1]]}\n"

                if str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b')).upper() == 'SEP' or str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b')).upper() == 'DEC':
                    Pnl_stmt += f"YPROFIT: \t{pnl.loc['NET PROFIT', pnl.columns[-5]]}\t\t{pnl.loc['NET PROFIT', pnl.columns[-4]]}\t\t{pnl.loc['NET PROFIT', pnl.columns[-3]]}\t\t{pnl.loc['NET PROFIT', pnl.columns[-2]]}\t\t{pnl.loc['NET PROFIT', pnl.columns[-1]]}\t\t{PROFIT_Qsum_in_Y}\n"
                else:
                    Pnl_stmt += f"YPROFIT: \t{pnl.loc['NET PROFIT', pnl.columns[-5]]}\t\t{pnl.loc['NET PROFIT', pnl.columns[-4]]}\t\t{pnl.loc['NET PROFIT', pnl.columns[-3]]}\t\t{pnl.loc['NET PROFIT', pnl.columns[-2]]}\t\t{pnl.loc['NET PROFIT', pnl.columns[-1]]}\n"
                if str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b')).upper() == 'SEP' or str(datetime.datetime.strftime(qtr_pnl.columns[-1], '%b')).upper() == 'DEC':
                    Pnl_stmt += f"PQoQ%: \t{pnl.loc['NET PROFIT_QoQ', pnl.columns[-5]]}\t\t{pnl.loc['NET PROFIT_QoQ', pnl.columns[-4]]}\t\t{pnl.loc['NET PROFIT_QoQ', pnl.columns[-3]]}\t\t{pnl.loc['NET PROFIT_QoQ', pnl.columns[-2]]}\t\t{pnl.loc['NET PROFIT_QoQ', pnl.columns[-1]]}\t\t{PROFIT_Qsum_QoQ_in_Y}\n"
                else:
                    Pnl_stmt += f"PQoQ%: \t{pnl.loc['NET PROFIT_QoQ', pnl.columns[-5]]}\t\t{pnl.loc['NET PROFIT_QoQ', pnl.columns[-4]]}\t\t{pnl.loc['NET PROFIT_QoQ', pnl.columns[-3]]}\t\t{pnl.loc['NET PROFIT_QoQ', pnl.columns[-2]]}\t\t{pnl.loc['NET PROFIT_QoQ', pnl.columns[-1]]}\n"

                send_metadata['YPNL_Statement']+=Pnl_stmt

            return send_metadata
        else:
            send_metadata['updated_results_on'] = datetime.datetime.now()
            if len(balancesht.columns)>3 :
                # YEARLY ANALYSIS
                BS_CY = balancesht.columns[-1]
                BS_2Y = balancesht.columns[-2]
                BS_3Y = balancesht.columns[-3]
                show_BS_CY = str(datetime.datetime.strftime(BS_CY, '%b-%Y'))
                show_BS_2Y = str(datetime.datetime.strftime(BS_2Y, '%b-%Y'))
                show_BS_3Y = str(datetime.datetime.strftime(BS_3Y, '%b-%Y'))

                eq_BS_CY = round(balancesht.loc["NO. OF EQUITY SHARES", BS_CY] / 10000000, 1)
                eq_BS_2Y = round(balancesht.loc['NO. OF EQUITY SHARES', BS_2Y] / 10000000, 1)
                FV_BS_CY = round(balancesht.loc['FACE VALUE', BS_CY])
                FV_BS_2Y = round(balancesht.loc['FACE VALUE', BS_2Y])
                ROCE_BS_CY = balancesht.loc['ROCE', BS_CY]
                ROCE_BS_2Y = balancesht.loc['ROCE', BS_2Y]
                #send_metadata['BalanceSheet_Statement'] = f"\n*****\nEquity in {str(datetime.datetime.strftime(BS_CY, '%b-%Y'))}: {str(eq_BS_CY)}cr; in {str(datetime.datetime.strftime(BS_2Y, '%b-%Y'))}: {str(eq_BS_2Y)}cr\nFaceValue in {str(datetime.datetime.strftime(BS_CY, '%b-%Y'))}: {str(FV_BS_CY)}; in {str(datetime.datetime.strftime(BS_2Y, '%b-%Y'))}: {str(FV_BS_2Y)}\nROCE in {str(datetime.datetime.strftime(BS_CY, '%b-%Y'))}: {str(ROCE_BS_CY)}%; in {str(datetime.datetime.strftime(BS_2Y, '%b-%Y'))}: {str(ROCE_BS_2Y)}%\n"

                Reserves_CY = balancesht.loc['RESERVES', BS_CY]
                Reserves_2Y = balancesht.loc['RESERVES', BS_2Y]
                Reserves_3Y = balancesht.loc['RESERVES', BS_3Y]
                Borrowings_CY = balancesht.loc['BORROWINGS', BS_CY]
                Borrowings_2Y = balancesht.loc['BORROWINGS', BS_2Y]
                Borrowings_3Y = balancesht.loc['BORROWINGS', BS_3Y]
            if len(pnl.columns)>3:
                PNL_CY = pnl.columns[-1]
                PNL_2Y = pnl.columns[-2]
                PNL_3Y = pnl.columns[-3]

                YSALES_Y = pnl.loc['SALES', PNL_CY]
                YSALES_2Y = pnl.loc['SALES', PNL_2Y]
                YSALES_3Y = pnl.loc['SALES', PNL_3Y]

                YSALES_QoQ_Y = pnl.loc['SALES_QoQ', PNL_CY]
                YSALES_QoQ_2Y = pnl.loc['SALES_QoQ', PNL_2Y]
                YSALES_QoQ_3Y = pnl.loc['SALES_QoQ', PNL_3Y]

                YPROFIT_QoQ_Y = pnl.loc['NET PROFIT_QoQ', PNL_CY]
                YPROFIT_QoQ_2Y = pnl.loc['NET PROFIT_QoQ', PNL_2Y]
                YPROFIT_QoQ_3Y = pnl.loc['NET PROFIT_QoQ', PNL_3Y]

                YPROFIT_Y = pnl.loc['NET PROFIT', PNL_CY]
                YPROFIT_2Y = pnl.loc['NET PROFIT', PNL_2Y]
                YPROFIT_3Y = pnl.loc['NET PROFIT', PNL_3Y]

                YNPM_CY = pnl.loc['NPM %', PNL_CY]
                YNPM_2Y = pnl.loc['NPM %', PNL_2Y]
                YNPM_3Y = pnl.loc['NPM %', PNL_3Y]
            if len(balancesht.columns) > 3:
                if Borrowings_CY<0.03*Reserves_CY :
                    if Borrowings_2Y<0.03*Reserves_2Y:
                        if Borrowings_3Y<0.03*Reserves_3Y:
                            send_metadata['pros'].append("DEBT-Free Company for the last 3 yrs")
                        else:
                            send_metadata['pros'].append("A DEBT-Free Company for the last 2 yrs")
                    else:
                        send_metadata['pros'].append("A DEBT-Free Company")
                if Borrowings_CY<0.85*Borrowings_2Y and Borrowings_2Y < 0.85*Borrowings_3Y:
                    borr_trend = " Declining by <15% for the last 3yrs"
                elif Borrowings_CY<Borrowings_2Y<Borrowings_3Y:
                    borr_trend = " Declining for last 3yrs"
                elif Borrowings_CY>Borrowings_2Y>Borrowings_3Y:
                    borr_trend = " increasing for the last 3yrs"
                elif Borrowings_CY>Borrowings_2Y<Borrowings_3Y:
                    borr_trend = " have increased this year"
                else:
                    borr_trend = f"are at {Borrowings_CY}"

                # EXPANSION
                CWIP_Y = balancesht.loc["CAPITAL WORK IN PROGRESS", BS_CY]
                CWIP_2Y = balancesht.loc["CAPITAL WORK IN PROGRESS", BS_2Y]
                CWIP_3Y = balancesht.loc["CAPITAL WORK IN PROGRESS", BS_3Y]

                NETBLOCK_Y = balancesht.loc["NET BLOCK", BS_CY]
                NETBLOCK_2Y = balancesht.loc["NET BLOCK", BS_2Y]
                NETBLOCK_3Y = balancesht.loc["NET BLOCK", BS_3Y]

                DEBTORDAYS_Y = balancesht.loc['DEBTOR DAYS', BS_CY]
                DEBTORDAYS_2Y = balancesht.loc['DEBTOR DAYS', BS_2Y]
                DEBTORDAYS_3Y = balancesht.loc['DEBTOR DAYS', BS_3Y]

                INVENTORYTURNOVER_Y = balancesht.loc['INVENTORY TURNOVER', BS_CY]
                INVENTORYTURNOVER_2Y = balancesht.loc['INVENTORY TURNOVER', BS_2Y]
                INVENTORYTURNOVER_3Y = balancesht.loc['INVENTORY TURNOVER', BS_3Y]

                Expansion = "EXPANSION ??"
                if CWIP_Y > NETBLOCK_2Y and CWIP_2Y > NETBLOCK_3Y:
                    Expansion += f"- EXPANSION in the last 2 consequent years. Check if it realised SALES this year"
                    send_metadata['tags'].append(f"EXPANSION {datetime.datetime.strftime(balancesht.columns[-2],'%Y')}")
                    send_metadata['tags'].append(f"EXPANSION {datetime.datetime.strftime(balancesht.columns[-1],'%Y')}")

                else:
                    if CWIP_Y > NETBLOCK_2Y:
                        Expansion += f"- EXPANSION underway in {show_BS_CY}, Better to Check the next Quarterly Results to see if the expansion is being effectively Utilised"
                        send_metadata['tags'].append(f"EXPANSION {datetime.datetime.strftime(balancesht.columns[-1],'%Y')}")
                    if CWIP_2Y > NETBLOCK_3Y and NETBLOCK_Y > NETBLOCK_2Y and CWIP_Y < NETBLOCK_2Y:
                        Expansion += f"- There was an Expansion in the {show_BS_2Y}, so got to check SALES too"
                        send_metadata['tags'].append(f"EXPANSION {datetime.datetime.strftime(balancesht.columns[-2],'%Y')}")
                if Expansion != "EXPANSION ??":
                    send_metadata['pros'].append(Expansion)

                    # with st.expander("UNDERGOING EXPANSION"):
                    #     balancesht1 = balancesht.transpose()
                    #     balancesht2 = balancesht1[
                    #         ['NET BLOCK', 'CAPITAL WORK IN PROGRESS', 'DEBTOR DAYS', 'INVENTORY TURNOVER', 'RESERVES',
                    #          'BORROWINGS']]
                    #     st.dataframe(balancesht2.transpose())
                    #     pnl1 = pnl.transpose()
                    #     pnl2 = pnl1[['SALES', 'NET PROFIT', 'OPM %', 'NPM %']]
                    #     st.dataframe(pnl2.transpose())

                if DEBTORDAYS_Y < DEBTORDAYS_2Y < DEBTORDAYS_3Y:
                    Product_Demand = "- PRODUCT DEMAND is rising since 3 years"
                elif DEBTORDAYS_Y < DEBTORDAYS_2Y:
                    Product_Demand = "- PRODUCT DEMAND is rising since 3 years"
                elif DEBTORDAYS_Y > DEBTORDAYS_2Y > DEBTORDAYS_3Y:
                    Product_Demand = "- NO PRODUCT DEMAND"
                else :
                    Product_Demand = ""

                if Product_Demand == "- NO PRODUCT DEMAND":
                    send_metadata['cons'].append(Product_Demand)
                elif Product_Demand != "":
                    send_metadata['pros'].append(Product_Demand)
                    send_metadata['tags'].append("PRODUCT DEMAND")

                Growth_metric = 15
                if ROCE_BS_CY > Growth_metric:
                    if ROCE_BS_2Y > Growth_metric:
                        YROCE_A = f"- ROCE IS GROWING AT MORE THAN {Growth_metric}% SINCE 2 YEARS.\nROCE in {datetime.datetime.strftime(BS_CY,'%b%Y')} is at {ROCE_BS_CY}%"
                        send_metadata['pros'].append(YROCE_A)
                    else:
                        YROCE_A = f"- ROCE is at {ROCE_BS_CY}% in {datetime.datetime.strftime(BS_CY,'%b%Y')}."

                if Reserves_CY > Reserves_2Y > Reserves_3Y and ((Borrowings_CY<Borrowings_2Y<Borrowings_3Y) or Borrowings_CY<1 ):
                    send_metadata['pros'].append(f"- RESERVES are increasing at above 15% in the last 3 years with Borrowings {borr_trend}")
            if len(pnl.columns)>3:
                # PROFIT OR SALES DOUBLED IN THE LAST 3 YEARS
                if len(pnl.columns) > 3:# and len(qtr_pnl.columns)>3:
                    if YSALES_Y>2*YSALES_2Y and YPROFIT_Y>2*YPROFIT_2Y:
                        send_metadata['pros'].append(f"- YProfits and YSales DOUBLED in {datetime.datetime.strftime(PNL_CY, '%b%Y')}")
                        send_metadata['tags'].append(f"YSD YPD {str(datetime.datetime.strftime(pnl.columns[-1], '%Y'))}")
                    if YSALES_2Y>2*YSALES_3Y and YPROFIT_2Y>2*YPROFIT_3Y:
                        send_metadata['pros'].append(f"- YProfits and YSales DOUBLED in {datetime.datetime.strftime(PNL_2Y, '%b%Y')}")
                        send_metadata['tags'].append(f"YSD YPD {str(datetime.datetime.strftime(pnl.columns[-2], '%Y'))}")
                    if len(pnl.columns) > 4: #and len(qtr_pnl.columns) > 4:
                        if YSALES_3Y > 2 * pnl.loc['SALES', pnl.columns[-4]] and YPROFIT_3Y > 2 * pnl.loc['NET PROFIT', pnl.columns[-4]]:
                            send_metadata['pros'].append(f"- YProfits and YSales DOUBLED in {datetime.datetime.strftime(PNL_3Y, '%b%Y')}")
                            send_metadata['tags'].append(f"YSD YPD {str(datetime.datetime.strftime(pnl.columns[-3], '%Y'))}")
                    # CHECKS IF LAST YEAR PROFIT DOUBLED and THIS YEAR MAINTAINING
                    if YPROFIT_2Y > 2*YPROFIT_3Y and YPROFIT_Y > YPROFIT_2Y:
                        send_metadata['pros'].append(f"- LastYr YProfits Doubled and maintained in {datetime.datetime.strftime(PNL_CY, '%b%Y')}")
                        send_metadata['tags'].append(f"LYPD PM {str(datetime.datetime.strftime(pnl.columns[-1], '%Y'))}")

                # NPM IMPROVED BY 15%               # CHECK IF NET PROFITS ARE MAINTAINED EVER SINCE
                Ymargins_A = ""
                margins_metric = 15
                if YNPM_3Y > margins_metric:
                    if YPROFIT_2Y > YPROFIT_3Y and YPROFIT_Y > YPROFIT_2Y:
                        Ymargins_A = f"- YMargins improvised by more than {margins_metric}% since {datetime.datetime.strftime(PNL_3Y,'%b%Y')} and PROFITS been maintaining ever since"
                elif YNPM_2Y > margins_metric:
                    if YPROFIT_Y > YPROFIT_2Y:
                        Ymargins_A = f"- YMargins improvised by more than {margins_metric}% since {datetime.datetime.strftime(PNL_2Y,'%b%Y')} and PROFITS been maintaining ever since."
                else:
                    Ymargins_A = "- YMARGINS LESS THAN 15%"
                if Ymargins_A != "- YMARGINS LESS THAN 15%":
                    send_metadata['pros'].append(Ymargins_A)

                # CHECKS if ROCE SALES PROFIT > 15%
                Growth_metric = 15
                YTopline_Bottomline_A = ""
                if YSALES_QoQ_Y > Growth_metric and YPROFIT_QoQ_Y > Growth_metric:
                    if YSALES_QoQ_2Y > Growth_metric and YPROFIT_QoQ_2Y > Growth_metric:
                        YTopline_Bottomline_A = f"- QoQ_GROWTH IN YSALES YPROFIT IS MORE THAN 15%"
                        if YPROFIT_QoQ_Y > YSALES_QoQ_Y:
                            if YSALES_QoQ_Y > 15 and YSALES_QoQ_Y < 90 and YPROFIT_QoQ_Y>100:
                                YTopline_Bottomline_A += f"- YPROFIT DOUBLED MARGINS IMPROVED\nAs on {datetime.datetime.strftime(PNL_CY, '%b%Y')}, Sales have increased by {YSALES_QoQ_Y}% while the Profit QoQ_growth is much better than Sales itself at {YPROFIT_QoQ_Y}%"
                                send_metadata['tags'].append(f"YS15 YPD {str(datetime.datetime.strftime(pnl.columns[-1], '%Y'))}")
                            else:
                                YTopline_Bottomline_A += f"- YPROFIT GROWTH IS BETTER THAN YSALES GROWTH\nAs on {datetime.datetime.strftime(PNL_CY, '%b%Y')}, Sales have increased by {YSALES_QoQ_Y}% while the Profit QoQ_growth is much better than Sales itself at {YPROFIT_QoQ_Y}%"
                                send_metadata['tags'].append(f"PROFIT GROWTH BETTER THAN SALES GROWTH {str(datetime.datetime.strftime(pnl.columns[-1], '%Y'))}")
                        else:
                            YTopline_Bottomline_A += f"\nAs on {datetime.datetime.strftime(PNL_CY,'%b%Y')}, Sales have increased by {YSALES_QoQ_Y}% while the Profit QoQ_growth is at {YPROFIT_QoQ_Y}%"
                    else:
                        YTopline_Bottomline_A = f"\nYSALES in {datetime.datetime.strftime(PNL_CY,'%b-%Y')} is at {pnl.loc['SALES',PNL_CY]} " \
                                                f"showing a QoQ_growth of {YSALES_QoQ_Y}%. The YNETPROFIT grew by {YPROFIT_QoQ_Y}% to {YPROFIT_Y}"
                    send_metadata['pros'].append(YTopline_Bottomline_A)

                    
                # if YSALES_QoQ_Y > YSALES_QoQ_2Y and YPROFIT_QoQ_Y > YPROFIT_QoQ_2Y:
                #     YTopline_Bottomline_A = f"- QoQ_Growth in YSALES YPROFIT is GOOD.\nYSALES in {datetime.datetime.strftime(PNL_CY,'%b-%Y')} is at {pnl.loc['SALES',PNL_CY]} " \
                #                    f"showing a QoQ_growth of {YSALES_QoQ_Y}%. The YNETPROFIT grew by {YPROFIT_QoQ_Y}% to {YPROFIT_Y}"
                #     send_metadata['pros'].append(YTopline_Bottomline_A)

                                                                                                                                # SALES AND PROFIT is more than PREV_YEAR
                # if YSALES_Y > YSALES_2Y and YPROFIT_Y > YPROFIT_2Y:
                #     YPNL_2Y_A = f"- RISING YSALES & YPROFIT : \nYSALES QoQ_growth in {PNL_CY} is {YSALES_QoQ_Y}%. The YNETPROFIT QoQ_growth is {YPROFIT_QoQ_Y}%"
                #     send_metadata['pros'].append(YPNL_2Y_A)
                if YPROFIT_Y > 0 and YPROFIT_2Y>0:
                    YPNL_A = stmt_for_qoq(pnl)
                    send_metadata['YPNL_tweet'] = f"Yearly \n {YPNL_A}"
            # TO GET THE LAST 5 YRS BALANCE SHEET DATA
            if len(balancesht.columns)>5:
                bsht_stmt = f"\nBALANCE SHEET: \t{str(datetime.datetime.strftime(balancesht.columns[-5], '%b-%Y'))}\t{str(datetime.datetime.strftime(balancesht.columns[-4], '%b-%Y'))}\t{show_BS_3Y}\t{show_BS_2Y}\t{show_BS_CY}\n"
                bsht_stmt += f"EQUITY(cr): \t{round(balancesht.loc['NO. OF EQUITY SHARES', balancesht.columns[-5]] / 10000000, 1)}\t{round(balancesht.loc['NO. OF EQUITY SHARES', balancesht.columns[-4]] / 10000000, 2)}\t{round(balancesht.loc['NO. OF EQUITY SHARES', balancesht.columns[-3]] / 10000000, 1)}\t{round(balancesht.loc['NO. OF EQUITY SHARES', balancesht.columns[-2]] / 10000000, 1)}\t{round(balancesht.loc['NO. OF EQUITY SHARES', balancesht.columns[-1]] / 10000000, 1)}\n"
                bsht_stmt += f"ROCE: \t{balancesht.loc['ROCE', balancesht.columns[-5]]}%\t{balancesht.loc['ROCE', balancesht.columns[-4]]}%\t{balancesht.loc['ROCE', balancesht.columns[-3]]}%\t{balancesht.loc['ROCE', balancesht.columns[-2]]}%\t{balancesht.loc['ROCE', balancesht.columns[-1]]}%\n"

                bsht_stmt += f"DEBTOR DAYS: \t{balancesht.loc['DEBTOR DAYS', balancesht.columns[-5]]}%\t{balancesht.loc['DEBTOR DAYS', balancesht.columns[-4]]}%\t{balancesht.loc['DEBTOR DAYS', balancesht.columns[-3]]}%\t{balancesht.loc['DEBTOR DAYS', balancesht.columns[-2]]}%\t{balancesht.loc['DEBTOR DAYS', balancesht.columns[-1]]}%\n"
                bsht_stmt += f"INVENTORY TURNOVER: \t{balancesht.loc['INVENTORY TURNOVER', balancesht.columns[-5]]}%\t{balancesht.loc['INVENTORY TURNOVER', balancesht.columns[-4]]}%\t{balancesht.loc['INVENTORY TURNOVER', balancesht.columns[-3]]}%\t{balancesht.loc['INVENTORY TURNOVER', balancesht.columns[-2]]}%\t{balancesht.loc['INVENTORY TURNOVER', balancesht.columns[-1]]}%\n"
                bsht_stmt += f"NET BLOCK: \t{balancesht.loc['NET BLOCK', balancesht.columns[-5]]}%\t{balancesht.loc['NET BLOCK', balancesht.columns[-4]]}%\t{balancesht.loc['NET BLOCK', balancesht.columns[-3]]}%\t{balancesht.loc['NET BLOCK', balancesht.columns[-2]]}%\t{balancesht.loc['NET BLOCK', balancesht.columns[-1]]}%\n"
                bsht_stmt += f"CWIP: \t{balancesht.loc['CAPITAL WORK IN PROGRESS', balancesht.columns[-5]]}%\t{balancesht.loc['CAPITAL WORK IN PROGRESS', balancesht.columns[-4]]}%\t{balancesht.loc['CAPITAL WORK IN PROGRESS', balancesht.columns[-3]]}%\t{balancesht.loc['CAPITAL WORK IN PROGRESS', balancesht.columns[-2]]}%\t{balancesht.loc['CAPITAL WORK IN PROGRESS', balancesht.columns[-1]]}%\n"

                bsht_stmt += f"RESERVES : \t{balancesht.loc['RESERVES', balancesht.columns[-5]]}%\t{balancesht.loc['RESERVES', balancesht.columns[-4]]}%\t{balancesht.loc['RESERVES', balancesht.columns[-3]]}%\t{balancesht.loc['RESERVES', balancesht.columns[-2]]}%\t{balancesht.loc['RESERVES', balancesht.columns[-1]]}%\n"
                bsht_stmt += f"BORROWINGS : \t{balancesht.loc['BORROWINGS', balancesht.columns[-5]]}%\t{balancesht.loc['BORROWINGS', balancesht.columns[-4]]}%\t{balancesht.loc['BORROWINGS', balancesht.columns[-3]]}%\t{balancesht.loc['BORROWINGS', balancesht.columns[-2]]}%\t{balancesht.loc['BORROWINGS', balancesht.columns[-1]]}%\n"
                # bsht_stmt += f"NET CASH FLOW : \t{balancesht.loc['NET CASH FLOW', balancesht.columns[-5]]}%\t{balancesht.loc['NET CASH FLOW', balancesht.columns[-4]]}%\t{balancesht.loc['NET CASH FLOW', balancesht.columns[-3]]}%\t{balancesht.loc['NET CASH FLOW', balancesht.columns[-2]]}%\t{balancesht.loc['NET CASH FLOW', balancesht.columns[-1]]}%\t"
                send_metadata['BALANCE_SHEET_Statement'] = bsht_stmt
            return send_metadata
        
def analyse_Q_df(qtr_pnl):
    send_metadata = {}
    send_metadata['pros'] = []
    send_metadata['cons'] = []
    send_metadata['statement'] = []
    send_metadata['tags'] = []
    send_metadata['QPNL_Statement'] = ""
    qtr_pnl.columns = pd.to_datetime(qtr_pnl.columns, format='%Y-%m-%d')
    
    if not qtr_pnl.empty and qtr_pnl is not None and isinstance(qtr_pnl,pd.DataFrame) :                                                                                                                       # DEFINING THE TABLE CONTENTS AS OBJECTS
        send_metadata['updated_results_on'] = datetime.datetime.now()
        QPNL_CQ = qtr_pnl.columns[-1]
        send_metadata['recent_quarter'] = QPNL_CQ

        if len(qtr_pnl.columns)>3:
            # QUARTERLY ANALYSIS
            QPNL_2Q = qtr_pnl.columns[-2]
            QPNL_3Q = qtr_pnl.columns[-3]
            show_QPNL_CQ = str(datetime.datetime.strftime(QPNL_CQ, '%b-%Y'))
            show_QPNL_2Q = str(datetime.datetime.strftime(QPNL_2Q, '%b-%Y'))
            show_QPNL_3Q = str(datetime.datetime.strftime(QPNL_3Q, '%b-%Y'))

            QSALES_CQ = qtr_pnl.loc['SALES', QPNL_CQ]
            QSALES_2Q = qtr_pnl.loc['SALES', QPNL_2Q]
            QSALES_3Q = qtr_pnl.loc['SALES', QPNL_3Q]

            QPROFIT_CQ = qtr_pnl.loc['NET PROFIT', QPNL_CQ]
            QPROFIT_2Q = qtr_pnl.loc['NET PROFIT', QPNL_2Q]
            QPROFIT_3Q = qtr_pnl.loc['NET PROFIT', QPNL_3Q]

            QOPM_CQ = qtr_pnl.loc['OPM %',QPNL_CQ]
            QOPM_2Q = qtr_pnl.loc['OPM %',QPNL_2Q]
            QOPM_3Q = qtr_pnl.loc['OPM %',QPNL_3Q]

            QNPM_CQ = qtr_pnl.loc['NPM %', QPNL_CQ]
            QNPM_2Q = qtr_pnl.loc['NPM %', QPNL_2Q]
            QNPM_3Q = qtr_pnl.loc['NPM %', QPNL_3Q]

                                                                                                                            # ANALYSING BALANCESHEET

        if len(qtr_pnl.columns)>3:
            if QOPM_CQ < QOPM_2Q < QOPM_3Q and QNPM_CQ < QNPM_2Q < QNPM_3Q :
                if len(pnl.columns)>3:
                    if YNPM_CY < YNPM_2Y:
                        send_metadata['pros'].append(f"- DECLINING QMARGINS, BOTH QOPM AND QNPM")
                # send_metadata['cons'].append(f"- DECLINING QUARTERLY MARGINS, BOTH OPM AND NPM")
            if QSALES_CQ > QSALES_2Q and QPROFIT_CQ > QPROFIT_2Q:
                if QPROFIT_CQ>QSALES_CQ: 
                    QPNL_2Q_A = f"- NPM seems to be improvised : The QPROFIT is showing better QoQ_growth at {qtr_pnl.loc['NET PROFIT_QoQ',QPNL_CQ]}% while QSALES QoQ_Growth is at {qtr_pnl.loc['SALES_QoQ',QPNL_CQ]}%"
                else:
                    QPNL_2Q_A = f"- RISING QSALES & QPROFIT : QSALES QoQ_Growth in {show_QPNL_CQ} is at {qtr_pnl.loc['SALES_QoQ',QPNL_CQ]}%. The QPROFIT QoQ_growth by {qtr_pnl.loc['NET PROFIT_QoQ',QPNL_CQ]}%"
                send_metadata['pros'].append(QPNL_2Q_A)
            if QPROFIT_CQ>0 and QPROFIT_2Q>0:
                QPNL_A = stmt_for_qoq(qtr_pnl)
                send_metadata['QPNL_tweet'] = f"Quarterly \n{QPNL_A}"
        
        return send_metadata

def analyse_Y_df(pnl,balancesht):
    send_metadata = {}
    send_metadata['BalanceSheet_Statement'] = ""
    send_metadata['pros'] = []
    send_metadata['cons'] = []
    send_metadata['statement'] = []
    send_metadata['YPNL_Statement'] = ""
    send_metadata['BALANCE_SHEET'] = ""
    send_metadata['tags'] = []
    
    if not pnl.empty and not balancesht.empty and pnl is not None and balancesht is not None and isinstance(pnl, pd.DataFrame) and isinstance(balancesht, pd.DataFrame):
        send_metadata['updated_results_on'] = datetime.datetime.now()
        if len(balancesht.columns)>3 :
            # YEARLY ANALYSIS
            BS_CY = balancesht.columns[-1]
            BS_2Y = balancesht.columns[-2]
            BS_3Y = balancesht.columns[-3]
            show_BS_CY = str(datetime.datetime.strftime(BS_CY, '%b-%Y'))
            show_BS_2Y = str(datetime.datetime.strftime(BS_2Y, '%b-%Y'))
            show_BS_3Y = str(datetime.datetime.strftime(BS_3Y, '%b-%Y'))

            eq_BS_CY = round(balancesht.loc["NO. OF EQUITY SHARES", BS_CY] / 10000000, 1)
            eq_BS_2Y = round(balancesht.loc['NO. OF EQUITY SHARES', BS_2Y] / 10000000, 1)
            FV_BS_CY = round(balancesht.loc['FACE VALUE', BS_CY])
            FV_BS_2Y = round(balancesht.loc['FACE VALUE', BS_2Y])
            ROCE_BS_CY = balancesht.loc['ROCE', BS_CY]
            ROCE_BS_2Y = balancesht.loc['ROCE', BS_2Y]
            #send_metadata['BalanceSheet_Statement'] = f"\n*****\nEquity in {str(datetime.datetime.strftime(BS_CY, '%b-%Y'))}: {str(eq_BS_CY)}cr; in {str(datetime.datetime.strftime(BS_2Y, '%b-%Y'))}: {str(eq_BS_2Y)}cr\nFaceValue in {str(datetime.datetime.strftime(BS_CY, '%b-%Y'))}: {str(FV_BS_CY)}; in {str(datetime.datetime.strftime(BS_2Y, '%b-%Y'))}: {str(FV_BS_2Y)}\nROCE in {str(datetime.datetime.strftime(BS_CY, '%b-%Y'))}: {str(ROCE_BS_CY)}%; in {str(datetime.datetime.strftime(BS_2Y, '%b-%Y'))}: {str(ROCE_BS_2Y)}%\n"

            Reserves_CY = balancesht.loc['RESERVES', BS_CY]
            Reserves_2Y = balancesht.loc['RESERVES', BS_2Y]
            Reserves_3Y = balancesht.loc['RESERVES', BS_3Y]
            Borrowings_CY = balancesht.loc['BORROWINGS', BS_CY]
            Borrowings_2Y = balancesht.loc['BORROWINGS', BS_2Y]
            Borrowings_3Y = balancesht.loc['BORROWINGS', BS_3Y]
        if len(pnl.columns)>3:
            PNL_CY = pnl.columns[-1]
            PNL_2Y = pnl.columns[-2]
            PNL_3Y = pnl.columns[-3]

            YSALES_Y = pnl.loc['SALES', PNL_CY]
            YSALES_2Y = pnl.loc['SALES', PNL_2Y]
            YSALES_3Y = pnl.loc['SALES', PNL_3Y]

            YSALES_QoQ_Y = pnl.loc['SALES_QoQ', PNL_CY]
            YSALES_QoQ_2Y = pnl.loc['SALES_QoQ', PNL_2Y]
            YSALES_QoQ_3Y = pnl.loc['SALES_QoQ', PNL_3Y]

            YPROFIT_QoQ_Y = pnl.loc['NET PROFIT_QoQ', PNL_CY]
            YPROFIT_QoQ_2Y = pnl.loc['NET PROFIT_QoQ', PNL_2Y]
            YPROFIT_QoQ_3Y = pnl.loc['NET PROFIT_QoQ', PNL_3Y]

            YPROFIT_Y = pnl.loc['NET PROFIT', PNL_CY]
            YPROFIT_2Y = pnl.loc['NET PROFIT', PNL_2Y]
            YPROFIT_3Y = pnl.loc['NET PROFIT', PNL_3Y]

            YNPM_CY = pnl.loc['NPM %', PNL_CY]
            YNPM_2Y = pnl.loc['NPM %', PNL_2Y]
            YNPM_3Y = pnl.loc['NPM %', PNL_3Y]
        if len(balancesht.columns) > 3:
            if Borrowings_CY<0.03*Reserves_CY :
                if Borrowings_2Y<0.03*Reserves_2Y:
                    if Borrowings_3Y<0.03*Reserves_3Y:
                        send_metadata['pros'].append("DEBT-Free Company for the last 3 yrs")
                    else:
                        send_metadata['pros'].append("A DEBT-Free Company for the last 2 yrs")
                else:
                    send_metadata['pros'].append("A DEBT-Free Company")
            if Borrowings_CY<0.85*Borrowings_2Y and Borrowings_2Y < 0.85*Borrowings_3Y:
                borr_trend = " Declining by <15% for the last 3yrs"
            elif Borrowings_CY<Borrowings_2Y<Borrowings_3Y:
                borr_trend = " Declining for last 3yrs"
            elif Borrowings_CY>Borrowings_2Y>Borrowings_3Y:
                borr_trend = " increasing for the last 3yrs"
            elif Borrowings_CY>Borrowings_2Y<Borrowings_3Y:
                borr_trend = " have increased this year"
            else:
                borr_trend = f"are at {Borrowings_CY}"

            # EXPANSION
            CWIP_Y = balancesht.loc["CAPITAL WORK IN PROGRESS", BS_CY]
            CWIP_2Y = balancesht.loc["CAPITAL WORK IN PROGRESS", BS_2Y]
            CWIP_3Y = balancesht.loc["CAPITAL WORK IN PROGRESS", BS_3Y]

            NETBLOCK_Y = balancesht.loc["NET BLOCK", BS_CY]
            NETBLOCK_2Y = balancesht.loc["NET BLOCK", BS_2Y]
            NETBLOCK_3Y = balancesht.loc["NET BLOCK", BS_3Y]

            DEBTORDAYS_Y = balancesht.loc['DEBTOR DAYS', BS_CY]
            DEBTORDAYS_2Y = balancesht.loc['DEBTOR DAYS', BS_2Y]
            DEBTORDAYS_3Y = balancesht.loc['DEBTOR DAYS', BS_3Y]

            INVENTORYTURNOVER_Y = balancesht.loc['INVENTORY TURNOVER', BS_CY]
            INVENTORYTURNOVER_2Y = balancesht.loc['INVENTORY TURNOVER', BS_2Y]
            INVENTORYTURNOVER_3Y = balancesht.loc['INVENTORY TURNOVER', BS_3Y]

            Expansion = "EXPANSION ??"
            if CWIP_Y > NETBLOCK_2Y and CWIP_2Y > NETBLOCK_3Y:
                Expansion += f"- EXPANSION in the last 2 consequent years. Check if it realised SALES this year"
                send_metadata['tags'].append(f"EXPANSION {datetime.datetime.strftime(balancesht.columns[-2],'%Y')}")
                send_metadata['tags'].append(f"EXPANSION {datetime.datetime.strftime(balancesht.columns[-1],'%Y')}")

            else:
                if CWIP_Y > NETBLOCK_2Y:
                    Expansion += f"- EXPANSION underway in {show_BS_CY}, Better to Check the next Quarterly Results to see if the expansion is being effectively Utilised"
                    send_metadata['tags'].append(f"EXPANSION {datetime.datetime.strftime(balancesht.columns[-1],'%Y')}")
                if CWIP_2Y > NETBLOCK_3Y and NETBLOCK_Y > NETBLOCK_2Y and CWIP_Y < NETBLOCK_2Y:
                    Expansion += f"- There was an Expansion in the {show_BS_2Y}, so got to check SALES too"
                    send_metadata['tags'].append(f"EXPANSION {datetime.datetime.strftime(balancesht.columns[-2],'%Y')}")
            if Expansion != "EXPANSION ??":
                send_metadata['pros'].append(Expansion)

                # with st.expander("UNDERGOING EXPANSION"):
                #     balancesht1 = balancesht.transpose()
                #     balancesht2 = balancesht1[
                #         ['NET BLOCK', 'CAPITAL WORK IN PROGRESS', 'DEBTOR DAYS', 'INVENTORY TURNOVER', 'RESERVES',
                #          'BORROWINGS']]
                #     st.dataframe(balancesht2.transpose())
                #     pnl1 = pnl.transpose()
                #     pnl2 = pnl1[['SALES', 'NET PROFIT', 'OPM %', 'NPM %']]
                #     st.dataframe(pnl2.transpose())

            if DEBTORDAYS_Y < DEBTORDAYS_2Y < DEBTORDAYS_3Y:
                Product_Demand = "- PRODUCT DEMAND is rising since 3 years"
            elif DEBTORDAYS_Y < DEBTORDAYS_2Y:
                Product_Demand = "- PRODUCT DEMAND is rising since 3 years"
            elif DEBTORDAYS_Y > DEBTORDAYS_2Y > DEBTORDAYS_3Y:
                Product_Demand = "- NO PRODUCT DEMAND"
            else :
                Product_Demand = ""

            if Product_Demand == "- NO PRODUCT DEMAND":
                send_metadata['cons'].append(Product_Demand)
            elif Product_Demand != "":
                send_metadata['pros'].append(Product_Demand)
                send_metadata['tags'].append("PRODUCT DEMAND")

            Growth_metric = 15
            if ROCE_BS_CY > Growth_metric:
                if ROCE_BS_2Y > Growth_metric:
                    YROCE_A = f"- ROCE IS GROWING AT MORE THAN {Growth_metric}% SINCE 2 YEARS.\nROCE in {datetime.datetime.strftime(BS_CY,'%b%Y')} is at {ROCE_BS_CY}%"
                    send_metadata['pros'].append(YROCE_A)
                else:
                    YROCE_A = f"- ROCE is at {ROCE_BS_CY}% in {datetime.datetime.strftime(BS_CY,'%b%Y')}."

            if Reserves_CY > Reserves_2Y > Reserves_3Y and ((Borrowings_CY<Borrowings_2Y<Borrowings_3Y) or Borrowings_CY<1 ):
                send_metadata['pros'].append(f"- RESERVES are increasing at above 15% in the last 3 years with Borrowings {borr_trend}")
        if len(pnl.columns)>3:
            # PROFIT OR SALES DOUBLED IN THE LAST 3 YEARS
            if len(pnl.columns) > 3:# and len(qtr_pnl.columns)>3:
                if YSALES_Y>2*YSALES_2Y and YPROFIT_Y>2*YPROFIT_2Y:
                    send_metadata['pros'].append(f"- YProfits and YSales DOUBLED in {datetime.datetime.strftime(PNL_CY, '%b%Y')}")
                    send_metadata['tags'].append(f"YSD YPD {str(datetime.datetime.strftime(pnl.columns[-1], '%Y'))}")
                if YSALES_2Y>2*YSALES_3Y and YPROFIT_2Y>2*YPROFIT_3Y:
                    send_metadata['pros'].append(f"- YProfits and YSales DOUBLED in {datetime.datetime.strftime(PNL_2Y, '%b%Y')}")
                    send_metadata['tags'].append(f"YSD YPD {str(datetime.datetime.strftime(pnl.columns[-2], '%Y'))}")
                if len(pnl.columns) > 4: #and len(qtr_pnl.columns) > 4:
                    if YSALES_3Y > 2 * pnl.loc['SALES', pnl.columns[-4]] and YPROFIT_3Y > 2 * pnl.loc['NET PROFIT', pnl.columns[-4]]:
                        send_metadata['pros'].append(f"- YProfits and YSales DOUBLED in {datetime.datetime.strftime(PNL_3Y, '%b%Y')}")
                        send_metadata['tags'].append(f"YSD YPD {str(datetime.datetime.strftime(pnl.columns[-3], '%Y'))}")
                # CHECKS IF LAST YEAR PROFIT DOUBLED and THIS YEAR MAINTAINING
                if YPROFIT_2Y > 2*YPROFIT_3Y and YPROFIT_Y > YPROFIT_2Y:
                    send_metadata['pros'].append(f"- LastYr YProfits Doubled and maintained in {datetime.datetime.strftime(PNL_CY, '%b%Y')}")
                    send_metadata['tags'].append(f"LYPD PM {str(datetime.datetime.strftime(pnl.columns[-1], '%Y'))}")

            # NPM IMPROVED BY 15%               # CHECK IF NET PROFITS ARE MAINTAINED EVER SINCE
            Ymargins_A = ""
            margins_metric = 15
            if YNPM_3Y > margins_metric:
                if YPROFIT_2Y > YPROFIT_3Y and YPROFIT_Y > YPROFIT_2Y:
                    Ymargins_A = f"- YMargins improvised by more than {margins_metric}% since {datetime.datetime.strftime(PNL_3Y,'%b%Y')} and PROFITS been maintaining ever since"
            elif YNPM_2Y > margins_metric:
                if YPROFIT_Y > YPROFIT_2Y:
                    Ymargins_A = f"- YMargins improvised by more than {margins_metric}% since {datetime.datetime.strftime(PNL_2Y,'%b%Y')} and PROFITS been maintaining ever since."
            else:
                Ymargins_A = "- YMARGINS LESS THAN 15%"
            if Ymargins_A != "- YMARGINS LESS THAN 15%":
                send_metadata['pros'].append(Ymargins_A)

            # CHECKS if ROCE SALES PROFIT > 15%
            Growth_metric = 15
            YTopline_Bottomline_A = ""
            if YSALES_QoQ_Y > Growth_metric and YPROFIT_QoQ_Y > Growth_metric:
                if YSALES_QoQ_2Y > Growth_metric and YPROFIT_QoQ_2Y > Growth_metric:
                    YTopline_Bottomline_A = f"- QoQ_GROWTH IN YSALES YPROFIT IS MORE THAN 15%"
                    if YPROFIT_QoQ_Y > YSALES_QoQ_Y:
                        if YSALES_QoQ_Y > 15 and YSALES_QoQ_Y < 90 and YPROFIT_QoQ_Y>100:
                            YTopline_Bottomline_A += f"- YPROFIT DOUBLED MARGINS IMPROVED\nAs on {datetime.datetime.strftime(PNL_CY, '%b%Y')}, Sales have increased by {YSALES_QoQ_Y}% while the Profit QoQ_growth is much better than Sales itself at {YPROFIT_QoQ_Y}%"
                            send_metadata['tags'].append(f"YS15 YPD {str(datetime.datetime.strftime(pnl.columns[-1], '%Y'))}")
                        else:
                            YTopline_Bottomline_A += f"- YPROFIT GROWTH IS BETTER THAN YSALES GROWTH\nAs on {datetime.datetime.strftime(PNL_CY, '%b%Y')}, Sales have increased by {YSALES_QoQ_Y}% while the Profit QoQ_growth is much better than Sales itself at {YPROFIT_QoQ_Y}%"
                            send_metadata['tags'].append(f"PROFIT GROWTH BETTER THAN SALES GROWTH {str(datetime.datetime.strftime(pnl.columns[-1], '%Y'))}")
                    else:
                        YTopline_Bottomline_A += f"\nAs on {datetime.datetime.strftime(PNL_CY,'%b%Y')}, Sales have increased by {YSALES_QoQ_Y}% while the Profit QoQ_growth is at {YPROFIT_QoQ_Y}%"
                else:
                    YTopline_Bottomline_A = f"\nYSALES in {datetime.datetime.strftime(PNL_CY,'%b-%Y')} is at {pnl.loc['SALES',PNL_CY]} " \
                                            f"showing a QoQ_growth of {YSALES_QoQ_Y}%. The YNETPROFIT grew by {YPROFIT_QoQ_Y}% to {YPROFIT_Y}"
                send_metadata['pros'].append(YTopline_Bottomline_A)

                
            # if YSALES_QoQ_Y > YSALES_QoQ_2Y and YPROFIT_QoQ_Y > YPROFIT_QoQ_2Y:
            #     YTopline_Bottomline_A = f"- QoQ_Growth in YSALES YPROFIT is GOOD.\nYSALES in {datetime.datetime.strftime(PNL_CY,'%b-%Y')} is at {pnl.loc['SALES',PNL_CY]} " \
            #                    f"showing a QoQ_growth of {YSALES_QoQ_Y}%. The YNETPROFIT grew by {YPROFIT_QoQ_Y}% to {YPROFIT_Y}"
            #     send_metadata['pros'].append(YTopline_Bottomline_A)

                                                                                                                            # SALES AND PROFIT is more than PREV_YEAR
            # if YSALES_Y > YSALES_2Y and YPROFIT_Y > YPROFIT_2Y:
            #     YPNL_2Y_A = f"- RISING YSALES & YPROFIT : \nYSALES QoQ_growth in {PNL_CY} is {YSALES_QoQ_Y}%. The YNETPROFIT QoQ_growth is {YPROFIT_QoQ_Y}%"
            #     send_metadata['pros'].append(YPNL_2Y_A)
            if YPROFIT_Y > 0 and YPROFIT_2Y>0:
                YPNL_A = stmt_for_qoq(pnl)
                send_metadata['YPNL_tweet'] = f"Yearly \n {YPNL_A}"
        # TO GET THE LAST 5 YRS BALANCE SHEET DATA
        if len(balancesht.columns)>5:
            bsht_stmt = f"\nBALANCE SHEET: \t{str(datetime.datetime.strftime(balancesht.columns[-5], '%b-%Y'))}\t{str(datetime.datetime.strftime(balancesht.columns[-4], '%b-%Y'))}\t{show_BS_3Y}\t{show_BS_2Y}\t{show_BS_CY}\n"
            bsht_stmt += f"EQUITY(cr): \t{round(balancesht.loc['NO. OF EQUITY SHARES', balancesht.columns[-5]] / 10000000, 1)}\t{round(balancesht.loc['NO. OF EQUITY SHARES', balancesht.columns[-4]] / 10000000, 2)}\t{round(balancesht.loc['NO. OF EQUITY SHARES', balancesht.columns[-3]] / 10000000, 1)}\t{round(balancesht.loc['NO. OF EQUITY SHARES', balancesht.columns[-2]] / 10000000, 1)}\t{round(balancesht.loc['NO. OF EQUITY SHARES', balancesht.columns[-1]] / 10000000, 1)}\n"
            bsht_stmt += f"ROCE: \t{balancesht.loc['ROCE', balancesht.columns[-5]]}%\t{balancesht.loc['ROCE', balancesht.columns[-4]]}%\t{balancesht.loc['ROCE', balancesht.columns[-3]]}%\t{balancesht.loc['ROCE', balancesht.columns[-2]]}%\t{balancesht.loc['ROCE', balancesht.columns[-1]]}%\n"

            bsht_stmt += f"DEBTOR DAYS: \t{balancesht.loc['DEBTOR DAYS', balancesht.columns[-5]]}%\t{balancesht.loc['DEBTOR DAYS', balancesht.columns[-4]]}%\t{balancesht.loc['DEBTOR DAYS', balancesht.columns[-3]]}%\t{balancesht.loc['DEBTOR DAYS', balancesht.columns[-2]]}%\t{balancesht.loc['DEBTOR DAYS', balancesht.columns[-1]]}%\n"
            bsht_stmt += f"INVENTORY TURNOVER: \t{balancesht.loc['INVENTORY TURNOVER', balancesht.columns[-5]]}%\t{balancesht.loc['INVENTORY TURNOVER', balancesht.columns[-4]]}%\t{balancesht.loc['INVENTORY TURNOVER', balancesht.columns[-3]]}%\t{balancesht.loc['INVENTORY TURNOVER', balancesht.columns[-2]]}%\t{balancesht.loc['INVENTORY TURNOVER', balancesht.columns[-1]]}%\n"
            bsht_stmt += f"NET BLOCK: \t{balancesht.loc['NET BLOCK', balancesht.columns[-5]]}%\t{balancesht.loc['NET BLOCK', balancesht.columns[-4]]}%\t{balancesht.loc['NET BLOCK', balancesht.columns[-3]]}%\t{balancesht.loc['NET BLOCK', balancesht.columns[-2]]}%\t{balancesht.loc['NET BLOCK', balancesht.columns[-1]]}%\n"
            bsht_stmt += f"CWIP: \t{balancesht.loc['CAPITAL WORK IN PROGRESS', balancesht.columns[-5]]}%\t{balancesht.loc['CAPITAL WORK IN PROGRESS', balancesht.columns[-4]]}%\t{balancesht.loc['CAPITAL WORK IN PROGRESS', balancesht.columns[-3]]}%\t{balancesht.loc['CAPITAL WORK IN PROGRESS', balancesht.columns[-2]]}%\t{balancesht.loc['CAPITAL WORK IN PROGRESS', balancesht.columns[-1]]}%\n"

            bsht_stmt += f"RESERVES : \t{balancesht.loc['RESERVES', balancesht.columns[-5]]}%\t{balancesht.loc['RESERVES', balancesht.columns[-4]]}%\t{balancesht.loc['RESERVES', balancesht.columns[-3]]}%\t{balancesht.loc['RESERVES', balancesht.columns[-2]]}%\t{balancesht.loc['RESERVES', balancesht.columns[-1]]}%\n"
            bsht_stmt += f"BORROWINGS : \t{balancesht.loc['BORROWINGS', balancesht.columns[-5]]}%\t{balancesht.loc['BORROWINGS', balancesht.columns[-4]]}%\t{balancesht.loc['BORROWINGS', balancesht.columns[-3]]}%\t{balancesht.loc['BORROWINGS', balancesht.columns[-2]]}%\t{balancesht.loc['BORROWINGS', balancesht.columns[-1]]}%\n"
            # bsht_stmt += f"NET CASH FLOW : \t{balancesht.loc['NET CASH FLOW', balancesht.columns[-5]]}%\t{balancesht.loc['NET CASH FLOW', balancesht.columns[-4]]}%\t{balancesht.loc['NET CASH FLOW', balancesht.columns[-3]]}%\t{balancesht.loc['NET CASH FLOW', balancesht.columns[-2]]}%\t{balancesht.loc['NET CASH FLOW', balancesht.columns[-1]]}%\t"
            send_metadata['BALANCE_SHEET_Statement'] = bsht_stmt
        return send_metadata
