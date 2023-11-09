import streamlit as st
from streamlit_option_menu import option_menu
import streamlit.components.v1 as html
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import base64  # Standard Python Module
from io import StringIO, BytesIO  # Standard Python Module
import os
import fundamentals
import json
from streamlit_lottie import st_lottie

st.set_page_config(page_title="Upload", page_icon=":bar_chart:", layout="wide",initial_sidebar_state="expanded",)

#color_dict = {'Yellow_Lite':"#f8ba43",'Yellow_Dark':"#D6D41B",'Blue_Lite':"#1959BF",'Blue_Dark':"#0971C9",'Green_Lite':"#11A694",'Green_Dark':"#11A64B",}   #"Purple_Lite":"#7019BF",'Purple_Dark':"#9319BF"}
color_dict = {'blue3':{'hash':'#00A3FE','rgb':'rgb(0,163,254)'},
              'yellow1':{'hash':'#FFFF01','rgb':'rgb(255,255,1)'},
              'blue1':{'hash':'#21A1E1', 'rgb':'rgb(33,161,225)'},
              'yellow2':{'hash':'#FFFE57','rgb':'rgb(255,254,87)'},
              'blue2':{'hash':'#5DB7D2','rgb':'rgb(93,183,210)'},
              'green1':{'hash':'#00F954','rgb':'rgb(0,249,84)'},
              'red1':{'hash':'#CC0118','rgb':'rgb(204,1,24)'},
              'black': {'hash': '000000', 'rgb': 'rgb(0,0,0)'},
              'white': {'hash': '#ffffff', 'rgb': 'rgb(255, 255, 255)'},
              }

def load_lottiefile(filepath: str):
    with open(filepath, "r") as f:
        return json.load(f)
def load_lottieurl(url: str):
    r = requests.get(url)
    if r.status_code != 200:
        return None
    return r.json()
lottie_data_analytics = load_lottiefile("./lottie/data-analytics.json")

with st.sidebar:
    uploaded_file = st.file_uploader("🔗 Upload here :", type=['xlsx', 'xlsm'],accept_multiple_files=True)  # Only accepts xlsx,xlsm file format
    st.markdown(""" <style> .font {
    font-size:22px ; font-family: 'Cooper Black'; color: #FF9633;} 
    </style> """, unsafe_allow_html=True)
    # Add a file uploader to allow users to upload their csv file

    st.markdown('<p class="font">Upload One or Two xlsx/xlsm FILES from screener.in </p>',
                unsafe_allow_html=True)
    st_lottie(
        lottie_data_analytics,
        speed=0.7,
        reverse=False,
        loop=True,
        quality="low",  # medium ; high
        height=None,
        width=None,
        key="barchart", )

st.title("📈 Don't See the Latest Charts?")

if uploaded_file is not None:
    if len(uploaded_file)==1:
        comp_Name = uploaded_file[0].name.split('.xlsx')[0]
        comp_Name = comp_Name.split('.xlsm')[0]
        sentence = f"{comp_Name}: "
        col1, col2, col3, col4 = st.columns([2, 2, 1, 1])
        with col1:
            st.subheader('📊 ' + comp_Name)
        with col4:
            color_key = st.selectbox("Bar Color",color_dict.keys())
            #color_bar = st.color_picker("Bar Color", value="#ECE80F")  # blueshades"#0971C9""ECE80F"   #yellowshades"#f8ba43"
        st.write("____")
        book = openpyxl.load_workbook(uploaded_file[0])
        #comp_name = book['Data Sheet']['B1'].value
        qtr_pnl,df_comp = fundamentals.get_tables(book[fundamentals.tabs[-1]], uploaded_file[0])  # send a sheet(not whole workbook)
        #pnl, balancesht,qtr_pnl = fundamentals.develop_data(qtr_pnl,df_comp)
        pnl, balancesht = fundamentals.develop_yearly(df_comp)
        qtr_pnl = fundamentals.develop_quarterly(qtr_pnl)

        try:
            df_comp.columns = df_comp.columns.strftime('%d-%m-%Y')
        except Exception as AttributeError:
            pass
        try:
            qtr_pnl.columns = qtr_pnl.columns.strftime('%d-%m-%Y')
        except Exception as AttributeError:
            pass
        # **************************************************************************************************

        with col2:
            sub_choose = st.selectbox("Fundamentals", fundamentals.funda_menu)

        if sub_choose == fundamentals.funda_menu[0]:            # YEARLY PNL
            sentence += fundamentals.stmt_for_qoq(pnl)
            coltw1, coltw2 = st.columns([4, 1])
            with coltw1:
                st.text_area(label="👉 INSIGHTS", value=sentence, height=75)
            Ykeydata,YSales, YOtherIncome,YExpenses,YOperatingProfit,YNetProfit,Ytable = st.tabs(['Key Data','SALES','OTHER INCOME','EXPENSES','OPERATING PROFIT','NET PROFIT','Table'])

            with Ytable:
                # THE FOLLOWIGN CODE CALCULATES THE GROWTH OR DEGROWTH
                st.dataframe(pnl.style.format(formatter="{:.1f}"))
            with Ykeydata:
                fundamentals.group_2_bars(pnl, "SALES", "PROFIT BEFORE TAX","NET PROFIT",comp_Name)
                fundamentals.both_lines(pnl, "OPM %", "NPM %", color_dict['red1']['hash'], color_dict['green1']['hash'], comp_Name)
            with YSales:
                fundamentals.qoq_growth(pnl, 'SALES', color_dict[color_key]['hash'], comp_Name)
            with YOtherIncome:
                fundamentals.go_bar(pnl, 'OTHER INCOME', color_dict[color_key]['hash'], comp_Name)
            with YExpenses:
                fundamentals.go_bar(pnl, 'EXPENSES', color_dict[color_key]['hash'], comp_Name)
            with YOperatingProfit:
                fundamentals.bar_line(pnl, 'OPERATING PROFIT','OPM %', color_dict[color_key]['hash'], comp_Name)
            with YNetProfit:
                fundamentals.bar_line(pnl,'NET PROFIT','NPM %',color_dict[color_key]['hash'],comp_Name)

        if sub_choose == fundamentals.funda_menu[3]:                #QUARTERLY PNL
            sentence += fundamentals.stmt_for_qoq(qtr_pnl)
            coltw1, coltw2 = st.columns([4, 1])
            with coltw1:
                st.text_area(label="TWITTER POST", value=sentence, height=75)
            Qkeydata, QSales, QOtherIncome, QExpenses, QOperatingProfit, QNetProfit, Qtable = st.tabs(
                ['Key Data', 'SALES', 'OTHER INCOME', 'EXPENSES', 'OPERATING PROFIT', 'NET PROFIT', 'Table'])
            with Qtable:
                st.dataframe(qtr_pnl)
                # Replace the first row with NaN for the QoQ columns
                #df.loc[0, ['SALES_QoQ', 'NET PROFIT_QoQ', 'OPERATING PROFIT_QoQ']] = np.nan
                #df = df.transpose()
                #st.dataframe(qtr_pnl.style.format(formatter="{:.1f}"))

            with Qkeydata:
                fundamentals.group_2_bars(qtr_pnl, "SALES",  "PROFIT BEFORE TAX","NET PROFIT",comp_Name)
                fundamentals.both_lines(qtr_pnl, "OPM %", "NPM %", color_dict['red1']['hash'], color_dict['green1']['hash'],comp_Name)
            with QSales:
                fundamentals.qoq_growth(qtr_pnl,"SALES", color_dict[color_key]['hash'],comp_Name)
            with QOtherIncome:
                fundamentals.go_bar(qtr_pnl, 'OTHER INCOME', color_dict[color_key]['hash'], comp_Name)
            with QExpenses:
                fundamentals.go_bar(qtr_pnl, 'EXPENSES', color_dict[color_key]['hash'], comp_Name)
            with QOperatingProfit:
                fundamentals.bar_line(qtr_pnl, 'OPERATING PROFIT','OPM %', color_dict[color_key]['hash'], comp_Name)
            with QNetProfit:
                fundamentals.bar_line(qtr_pnl,'NET PROFIT','NPM %',color_dict[color_key]['hash'],comp_Name)


        if sub_choose == fundamentals.funda_menu[1]:        #YEARLY BALANCE SHEET
            BSKeyData, BStable, BSReserves, BSBorrowings, BSOtherAssets, BSOtherLiabilities, BSReceivables, BSInventory, BSCWIP = st.tabs(['KeyData','table','Reserves','Borrowings','OtherAssets','OtherLiabilities','Receivables','Inventory','CWIP'])
            with BSKeyData:
                fundamentals.bar_line(balancesht,"RESERVES","BORROWINGS",color_dict[color_key]['hash'],comp_Name)
                fundamentals.bar_line(balancesht,"RECEIVABLES","INVENTORY",color_dict[color_key]['hash'],comp_Name)
            with BStable:
                st.dataframe(balancesht.style.format(formatter="{:.1f}"))
            with BSCWIP:
                fundamentals.go_bar(balancesht, "CAPITAL WORK IN PROGRESS", color_dict[color_key]['hash'],comp_Name)
            with BSInventory:
                fundamentals.go_bar(balancesht, "INVENTORY", color_dict[color_key]['hash'],comp_Name)
            with BSReserves:
                fundamentals.qoq_growth(balancesht, "RESERVES", color_dict[color_key]['hash'],comp_Name)
            with BSBorrowings:
                fundamentals.qoq_growth(balancesht, "BORROWINGS", color_dict[color_key]['hash'],comp_Name)
            with BSReceivables:
                fundamentals.qoq_growth(balancesht, "RECEIVABLES", color_dict[color_key]['hash'],comp_Name)
            with BSOtherAssets:
                fundamentals.qoq_growth(balancesht, "OTHER ASSETS", color_dict[color_key]['hash'],comp_Name)
            with BSOtherLiabilities:
                fundamentals.qoq_growth(balancesht, "OTHER LIABILITIES", color_dict[color_key]['hash'],comp_Name)


        if sub_choose == fundamentals.funda_menu[2]:        # YEARLY CASH FLOWS
            CF, CFTab, CFop, CFinv, CFfin, NetCF = st.tabs(["KeyData","Table","Operating Cash","Investing Cash","Financing Cash","Net Cash Flow"])
            with CFTab:
                st.dataframe(df_comp.loc[sub_choose].style.format(formatter="{:.1f}"))
            with CF:
                fundamentals.go_group_bar(df_comp.loc[sub_choose], "cash_flows", color_dict[color_key]['hash'])
            with CFop:
                fundamentals.qoq_growth(df_comp.loc[sub_choose], "CASH FROM OPERATING ACTIVITY", color_dict[color_key]['hash'],comp_Name)
            with CFfin:
                fundamentals.go_bar(df_comp.loc[sub_choose], "CASH FROM FINANCING ACTIVITY", color_dict[color_key]['hash'], comp_Name)
            with CFinv:
                fundamentals.go_bar(df_comp.loc[sub_choose], "CASH FROM INVESTING ACTIVITY", color_dict[color_key]['hash'],comp_Name)
            with NetCF:
                fundamentals.go_bar(df_comp.loc[sub_choose], "NET CASH FLOW", color_dict[color_key]['hash'], comp_Name)

    if len(uploaded_file)==2:
        comp1_Name = uploaded_file[0].name.split('.')[0]
        comp2_Name = uploaded_file[1].name.split('.')[0]
        col1, col2, col3 = st.columns([0.5, 0.3, 0.2])
        with col1:
            st.subheader('📈 ' + comp1_Name + '📈 ' + comp2_Name )
        with col2:
            qoq_checked = st.checkbox("Sequential_Growth_%")
        with col3:
            color_bar = st.color_picker("Bar Color",
                                        value="#ECE80F")  # blueshades"#0971C9""ECE80F"   #yellowshades"#f8ba43"
        st.write("____")
        book1 = openpyxl.load_workbook(uploaded_file[0])
        qtr1_pnl, df1 = fundamentals.get_tables(book1[fundamentals.tabs[-1]], uploaded_file[0])  # send a sheet(not whole workbook)
        book2 = openpyxl.load_workbook(uploaded_file[1])
        qtr2_pnl, df2 = fundamentals.get_tables(book2[fundamentals.tabs[-1]], uploaded_file[1])  # send a sheet(not whole workbook)
        book1 = openpyxl.load_workbook(uploaded_file[0])

        qtr1_pnl, df1 = fundamentals.get_tables(book1[fundamentals.tabs[-1]], uploaded_file[0])  # send a sheet(not whole workbook)
        book2 = openpyxl.load_workbook(uploaded_file[1])
        qtr2_pnl, df2 = fundamentals.get_tables(book2[fundamentals.tabs[-1]], uploaded_file[1])  # send a sheet(not whole workbook)
        main_menu = st.sidebar.selectbox("Chose", fundamentals.funda_keys)
        sub_menu = st.sidebar.selectbox("SubChose", list(df1.loc[main_menu].index))
        new_df = [df1.loc[main_menu].loc[sub_menu], df2.loc[main_menu].loc[sub_menu]]
        df_new = pd.concat(new_df, keys=[comp1_Name, comp2_Name], axis=1)
        fundamentals.peer_bar(df_new, sub_menu,comp1_Name,comp2_Name)
    if len(uploaded_file) >= 3:
        st.error("Please, upload only 2 excel files")




st.write("____")
st.write('made with :green_heart: to my Indian Stock Investors')


