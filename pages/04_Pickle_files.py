import streamlit as st
from time import sleep
import glob,os
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os.path
import win32com.client as win32
from win32com.client import Dispatch
from selenium.common.exceptions import NoSuchElementException
from selenium import webdriver
from selenium.webdriver.common.by import By

import fundamentals
st.set_page_config(page_title="Pickling", page_icon=":bar_chart:", layout="wide",initial_sidebar_state="expanded",)
download_excel_xpath = "/html/body/main/div[3]/div[1]/form/button"
login_button_xpath = "/html/body/main/div[2]/div[2]/form/button"

#C:\Users\sahaveer\PycharmProjects\webapps\Scripts\itimes\pages\05_Pickle_files.py:124: DeprecationWarning:
#find_elements_by_xpath is deprecated. Please use find_elements(by=By.XPATH, value=xpath) instead


def get_tables(datasht,file):
    for i in range(1,datasht.max_row+1) :
        if datasht['A'+str(i)].value == DataSheet_Key_Values[0] :
            pnl_start_row = i
        if datasht['A' + str(i)].value == DataSheet_Key_Values[1]:
            pnl_end_row = i-1
        if datasht['A' + str(i)].value == DataSheet_Key_Values[2]:
            quarterly_start_row = i
        if datasht['A' + str(i)].value == DataSheet_Key_Values[3]:
            quarterly_end_row = i-1
        if datasht['A' + str(i)].value == DataSheet_Key_Values[4]:
            BS_start_row = i
        if datasht['A' + str(i)].value == DataSheet_Key_Values[5]:
            BS_end_row = i-1
        if datasht['A' + str(i)].value == DataSheet_Key_Values[6]:
            cash_start_row = i
        if datasht['A' + str(i)].value == DataSheet_Key_Values[7]:
            cash_end_row = i-1
    reqd_cols = "A :" + str(get_column_letter(datasht.max_column))

    if pnl_start_row is not None and pnl_end_row is not None :
        pnl = pd.read_excel(file, index_col=0, sheet_name=tabs[-1], header=pnl_start_row,usecols=reqd_cols,
                            nrows=pnl_end_row-pnl_start_row )
        pnl.index = pnl.index.str.upper()
        # ADDING SOME MORE PARAMS
        Expenses = (pnl.loc['RAW MATERIAL COST'] + pnl.loc['POWER AND FUEL'] + pnl.loc['OTHER MFR. EXP'] + pnl.loc['EMPLOYEE COST'] + pnl.loc['SELLING AND ADMIN'] + pnl.loc['OTHER EXPENSES'] + (-1 * pnl.loc['CHANGE IN INVENTORY']))
        Expenses.name = "EXPENSES"
        pnl = pnl.append(Expenses)
        Op_profit = pnl.loc['SALES'] - pnl.loc['EXPENSES']
        Op_profit.name = "OPERATING PROFIT"
        pnl = pnl.append(Op_profit)
        OPM = (pnl.loc['OPERATING PROFIT'] / pnl.loc['SALES'])*100
        OPM.name = "OPM"
        pnl = pnl.append(OPM)
        npm_df = (pnl.loc['NET PROFIT'] / pnl.loc['SALES'])*100     # if NetProfit>0, else 0
        npm_df.name = 'NPM'
        pnl = pnl.append(npm_df)
        div_payout = (pnl.loc['DIVIDEND AMOUNT']/pnl.loc['NET PROFIT'])*100
        div_payout.name = 'DIV_PAYOUT'
        pnl = pnl.append(div_payout)

    if BS_start_row is not None and BS_end_row is not None:
        balancesht = pd.read_excel(file, index_col=0, sheet_name=tabs[-1], header=BS_start_row,usecols=reqd_cols,
                                   nrows=BS_end_row-BS_start_row )
        balancesht.index = balancesht.index.str.upper()
        W_C = balancesht.loc['OTHER ASSETS']-balancesht.loc['OTHER LIABILITIES']
        W_C.name = 'WORKING CAPITAL'
        balancesht = balancesht.append(W_C)
        Debtor_Days = balancesht.loc['RECEIVABLES'] / ( pnl.loc['SALES']/365 )   #if sales>0, else 0
        Debtor_Days.name = 'DEBTOR DAYS'
        balancesht = balancesht.append(Debtor_Days)
        Inv_Turn = pnl.loc['SALES'] / balancesht.loc['INVENTORY']    # if inv>0, else 0
        Inv_Turn.name = 'INVENTORY TURNOVER'
        balancesht = balancesht.append(Inv_Turn)
        ROE = (pnl.loc['NET PROFIT'] / (balancesht.loc['EQUITY SHARE CAPITAL']+balancesht.loc['RESERVES']))*100
        # if (Equity+Reserves)>0, else Null
        ROE.name = "ROE"
        balancesht = balancesht.append(ROE)
        ROCE = (pnl.loc['OPERATING PROFIT']-pnl.loc['DEPRECIATION'] - pnl.loc['TAX']) / (balancesht.loc['NET BLOCK'] + balancesht.loc['WORKING CAPITAL'])*100   #if Net Block+Working Capital>0, else Null
        ROCE.name = "ROCE"
        balancesht = balancesht.append(ROCE)

    if cash_start_row is not None and cash_end_row is not None:
        cashflow = pd.read_excel(file, index_col=0, sheet_name=tabs[-1], header=cash_start_row,usecols=reqd_cols,
                                   nrows=cash_end_row-cash_start_row)
        cashflow.index = cashflow.index.str.upper()
    if(pnl is not None and balancesht is not None and cashflow is not None):
        sht_list = [pnl,balancesht,cashflow]
        df_comp = pd.concat(sht_list,keys=funda_keys)
        try:
            df_comp.columns = df_comp.columns.strftime('%d-%m-%Y')
        except Exception as AttributeError:
            pass
    return df_comp

def login_screener(driver):
    driver.get('https://www.screener.in/login/')
    #if driver.find_elements_by_id("id_username"):
    if driver.find_elements(by=By.ID, value="id_username"):
        pass
    else:
        return
    #useridform = driver.find_element_by_id('id_username')
    useridform = driver.find_element(by=By.ID,value='id_username')
    pswrdform = driver.find_element(by=By.ID,value='id_password')
    #pswrdform = driver.find_element_by_id('id_password')
    usermail = "sahaveer@gmail.com"
    pswrd = "Qwerty@123"
    useridform.send_keys(usermail)
    pswrdform.send_keys(pswrd)
    #if driver.find_elements_by_xpath(login_button_xpath):
    if driver.find_elements(by=By.XPATH, value=login_button_xpath):
        #driver.find_element_by_xpath(login_button_xpath).click()
        driver.find_element(by=By.XPATH, value=login_button_xpath).click()

    else:
        st.error("seems like Screener has changed the Xpath")

def search_screener(driver, code):
    #if driver.find_elements_by_xpath('/html/body/nav/div[1]/div/div[1]/div/div[3]/div[1]'):
    if driver.find_elements(by=By.XPATH, value='/html/body/nav/div[1]/div/div[1]/div/div[3]/div[1]'):
        pass
    else:
        login_screener(driver)

    code_list = [code]
    print(code_list)
    #link_consolidated = 'https://www.screener.in/company/'+code+'/consolidated/'
    #link_standalone =  'https://www.screener.in/company/'+code+'/'
    for i in range(len(code_list)):
        #st.info(f"Inside the search_screener function, we are searching for {code_list[i]}")
        driver.get('https://www.screener.in/company/' + code_list[i] + '/consolidated/')
        sleep(3)
        #if driver.find_elements_by_xpath(download_excel_xpath):
        if driver.find_elements(by=By.XPATH, value=download_excel_xpath):
            #clic = driver.find_element_by_xpath(download_excel_xpath)
            # downloading the excel file
            #clic.click()
            driver.find_element(by=By.XPATH, value=download_excel_xpath).click()
            sleep(3)
            latest_file = recently_downloaded_file(path_download, type_file)
            book = openpyxl.load_workbook(latest_file)
            comp_name = book['Data Sheet']['B1'].value
            qtr_pnl, df_comp = fundamentals.get_tables(book[fundamentals.tabs[-1]],latest_file)  # send a sheet(not whole workbook)

            return df_comp, qtr_pnl
        else:
            driver.get('https://www.screener.in/company/' + code_list[i] + '/')
            sleep(3)
            #if driver.find_elements_by_xpath(download_excel_xpath):
            if driver.find_elements(by=By.XPATH, value=download_excel_xpath):
                #clic = driver.find_element_by_xpath(download_excel_xpath)
                # downloading the excel file
                #clic.click()
                driver.find_element(by=By.XPATH, value=download_excel_xpath).click()
                #st.write('Downloading ' + code_list[i])
                sleep(5)
                latest_file = recently_downloaded_file(path_download, type_file)
                #df_comp, comp_name = create_df(latest_file)
                book = openpyxl.load_workbook(latest_file)
                comp_name = book['Data Sheet']['B1'].value
                qtr_pnl, df_comp = fundamentals.get_tables(book[fundamentals.tabs[-1]],latest_file)  # send a sheet(not whole workbook)
                return df_comp, qtr_pnl
            else:
                return None, None

def recently_downloaded_file(path_download, type_file):
    sendfile = ''
    #grabs the last created file in a specified folder
    pick_xlsx = glob.glob(path_download+type_file)
    last_created_file = max(pick_xlsx, key=os.path.getctime)
    send_file = last_created_file.replace("\\","/")
    return send_file

def create_df(send_file):
    # ****************************************************
    # NOW ADD HERE A FUNCTION TO OPEN EXCEL SHEET, CREATE DF AND THEN CREATE PICKLE FILE OUT OF IT
    book = openpyxl.load_workbook(send_file)
    df_comp = get_tables(book['Data Sheet'], send_file)
    comp_name = book['Data Sheet']['B1'].value
    st.write(comp_name)
    #excel_file_name = send_file.split("/")[-1]
    #excel_name = excel_file_name.split('.xlsx')[0]
    return df_comp,comp_name

path_download = "C:/Users/sahaveer/Downloads/"
path_bhav = 'C:/Users/sahaveer/OneDrive/Documents/bhavcopy/'
path_to_save = path_download+"Results/"
type_file = "*.xlsx"


tabs = ['Profit & Loss', 'Quarters','Balance Sheet', 'Cash Flow' ,'Data Sheet']
DataSheet_Key_Values = ['PROFIT & LOSS', 'Dividend Amount', 'Quarters', 'Operating Profit', 'BALANCE SHEET', 'Cash & Bank',
                        'CASH FLOW:', 'Net Cash Flow']
funda_keys = ['PROFIT&LOSS','BALANCE SHEET','CASH FLOW']    # dont change the order of this list as it will affect the keys used in Yearly df
# HERE WE MUST BE ABLE TO OPEN FILES FROM STRATEGY CONVERTER FILE


compCode = st.text_input(label='Comp Code')
if st.button('PICKLE'):
    driver = webdriver.Edge(r"C://Users/sahaveer/PycharmProjects/msedgedriver.exe", )
    driver.maximize_window()
    login_screener(driver)
    sleep(2)
    yr_df, qtr_df = search_screener(driver, compCode)
    sleep(1)
    # SORT FILES ALPHABETICALLY
    first_letter = pickle_name[0].upper()  # Get the first letter and convert it to uppercase
    alphabetic_folder = os.path.join("./pickl/", first_letter)
    # Create the folder if it doesn't exist
    if not os.path.exists(alphabetic_folder):
        os.makedirs(alphabetic_folder)
    st.info(f"{alphabetic_folder}")
    #if yr_df != 1:
    if isinstance(yr_df, pd.DataFrame):
        save_pickl_as = alphabetic_folder + '/' + pickle_name + " Yearly.pkl"
        yr_df.to_pickle(save_pickl_as)
    #if qtr_df != 1:
    if isinstance(qtr_df, pd.DataFrame):
        save_pickl_as = alphabetic_folder + '/' + pickle_name + " Quarterly.pkl"
        qtr_df.to_pickle(save_pickl_as)
    st.info(f"saved pickl file {pickle_name} in working directory pickle folder ")
st.write("____")

st.write("Download pickle files from entire NSE List")
if st.button("PICKLE all NSE"):
    st.info("reading list from " + path_bhav+'nselist.txt')
    driver = webdriver.Edge(r"C://Users/sahaveer/PycharmProjects/msedgedriver.exe", )
    driver.maximize_window()
    login_screener(driver)
    sleep(2)
    with open(path_bhav+'nselist.txt','r') as file:
        for line in file:
            comp_Code = str(line)
            pickle_name = comp_Code.split('\n')[0]
            #st.info(f"Comp code is {comp_Code}")
            #st.info(f"Pickle Name is {pickle_name}")
            yr_df, qtr_df = search_screener(driver, comp_Code)
            sleep(1)
            # SORT FILES ALPHABETICALLY
            first_letter = pickle_name[0].upper()  # Get the first letter and convert it to uppercase
            alphabetic_folder = os.path.join("./pickl/", first_letter)
            # Create the folder if it doesn't exist
            if not os.path.exists(alphabetic_folder):
                os.makedirs(alphabetic_folder)
            #st.info(f"{alphabetic_folder}")
            #if yr_df != 1:
            if isinstance(yr_df, pd.DataFrame):
                save_pickl_as = alphabetic_folder + '/' + pickle_name + " Yearly.pkl"
                yr_df.to_pickle(save_pickl_as)
            #if qtr_df != 1:
            if isinstance(qtr_df, pd.DataFrame):
                save_pickl_as = alphabetic_folder + '/' + pickle_name + " Quarterly.pkl"
                qtr_df.to_pickle(save_pickl_as)

    st.info("saved pickl file in working directory pickle folder")

st.write("____")
st.write("Download from Downloaded path of Computer")
if st.button("Downloaded Excel files"):
    st.info("reading list from " + path_download+' path')
    for each_file in glob.glob(path_download+'*.xlsx', recursive=False):
        xl_file = each_file.replace('\\', '/')
        df_comp, comp_name = create_df(xl_file)
        pickle_name = comp_name
        #if df_comp != 1:
        if isinstance(df_comp, pd.DataFrame):
            df_comp.to_pickle(path_download+"pickl/" + pickle_name + ".pkl")
    st.info("saved pickl file in working directory pickle folder")

st.write("____")
st.write("Download pickle files from entire BSE List")
if st.button("BSE files"):
    st.info("reading list from " + path_bhav+'bselist.txt')
    driver = webdriver.Edge(r"C://Users/sahaveer/PycharmProjects/msedgedriver.exe", )
    driver.maximize_window()
    login_screener(driver)
    sleep(2)
    with open(path_bhav+'bselist.txt','r') as file:
        for line in file:
            comp_Code = str(line)
            pickle_name = comp_Code.split('\n')[0]
            #df_comp,comp_name = search_screener(driver, comp_Code)
            #sleep(1)
            yr_df, qtr_df = search_screener(driver, comp_Code)
            sleep(1)
            # Create the folder if it doesn't exist
            first_letter = pickle_name[0].upper()  # Get the first letter and convert it to uppercase
            alphabetic_folder = os.path.join("./pickl/", first_letter)
            if not os.path.exists(alphabetic_folder):
                os.makedirs(alphabetic_folder)
            #st.info(f"{alphabetic_folder}")
            #if yr_df != 1:
            if isinstance(yr_df, pd.DataFrame):
                save_pickl_as = alphabetic_folder + '/' + pickle_name + " Yearly.pkl"
                yr_df.to_pickle(save_pickl_as)
            #if qtr_df != 1:
            if isinstance(qtr_df, pd.DataFrame):
                save_pickl_as = alphabetic_folder + '/' + pickle_name + " Quarterly.pkl"
                qtr_df.to_pickle(save_pickl_as)
            st.info(f"saved pickl file {pickle_name} in working directory pickle folder ")
