import csv
import pandas as pd
import streamlit as st
from streamlit_option_menu import option_menu
import openpyxl
from openpyxl.utils import get_column_letter

import nse_bse_search
import fromyahoo
import lastdayprice
import numpy as np
import concurrent.futures
import time
import datetime
from datetime import timedelta
import threading
import fundamentals

#pledging_path = 'C:/Users/sahaveer/PycharmProjects/webapps/Scripts/itimes local/zerodha pledging.xlsx'
pledging_path = './zerodha pledging.xlsx'

st.set_page_config(page_title="Portfolio", page_icon=":bar_chart:", layout="wide",initial_sidebar_state="collapsed",)
st.title('Portfolio Proficiency Analyzer 💸')
#🌟

# Get today's date
today = datetime.datetime.now()
# Check if today is a weekend (Saturday or Sunday)
if today.weekday() >= 5:  # 5 represents Saturday, 6 represents Sunday
    # Calculate how many days to subtract to get the last weekday (Friday)
    days_to_subtract = (today.weekday() + 1) % 5  # Add 1 to shift Sunday to Monday
    # Subtract the appropriate number of days
    last_weekday = today - timedelta(days=days_to_subtract)
else:
    # If today is a weekday, use today's date
    last_weekday = today
dd = last_weekday.strftime('%d')
mm = last_weekday.strftime('%m')
mmm = last_weekday.strftime('%b')
yyyy = last_weekday.strftime('%Y')
yy = last_weekday.strftime('%y')
live_prices = {}

#tab1, tab2, tab3 = st.tabs(["Portfolio", "Collateral"])
portfolio_option = option_menu("", ["Portfolio", "Collateral"],icons=['cash', 'cash'], menu_icon="cast", default_index=0, orientation="horizontal")
if portfolio_option == "Portfolio":
    col1,col2 = st.columns([2,2])
    col2.subheader("💡 Key Benefits:")
    col2.write("✅ Calculate Open & Closed Portfolio")
    col2.write("📊 Analyze Your Past Actions")
    col2.write("💰 Free Shares Calculation")
    col2.write("💹 SIP vs Portfolio")
    col1.subheader("🔍 How it works:")
    #tradebook = st.file_uploader("upload TradeBook from Zerodha", type= ['xlsx'])
    tradebook_url = 'https://console.zerodha.com/reports/tradebook'
    col2.markdown(f"[***ZERODHA TRADEBOOK***]({tradebook_url})", unsafe_allow_html=True)
    tradebook = col1.file_uploader("upload TradeBooks from Zerodha", type= ['xlsx'],accept_multiple_files = True)
    st.info(f'https://www.bseindia.com/download/BhavCopy/Equity/EQ' + dd + mm + yy + '_CSV.ZIP')
    st.info(f"https://archives.nseindia.com/products/content/sec_bhavdata_full_" + dd + mm + yyyy + ".csv")
@st.cache_data
def tradebook_perday(xl):
    if isinstance(xl,pd.DataFrame):
        tradebook_day = pd.DataFrame()
        #symbol_isin = {}
        for each in xl['ISIN'].unique():
            # get the first group of SCRIPT
            xl_symbol = xl.groupby("ISIN").get_group(each)
            #phase-2      All transactions per each day into one
            df_grouped = xl_symbol[["Symbol","ISIN", "Trade Date", "Trade Type", "Quantity", "Price"]].set_index("ISIN")
            #TO REPLACE ISIN WITH THE SYMBOL NAMES
            #symbol_isin[each] = (df_grouped['Symbol'].iloc[0])
            # CALCULATE THE TRADE VAL and then sums up the BUYS/SELLS which happened on the same DATES
            df_grouped['Investment'] = df_grouped['Quantity'] * df_grouped['Price']
            temp_df_grouped = df_grouped.groupby(["ISIN", "Trade Date", "Trade Type"])[['Quantity', 'Investment']].sum()  # .reset_index()
                    # temp_df_grouped = pd.DataFrame(df_grouped).groupby(["Symbol","Trade Date","Trade Type"]).aggregate({'Quantity':'sum','Investment':sum})
            temp_df_grouped["avg_price"] = temp_df_grouped["Investment"] / temp_df_grouped["Quantity"]

            #HERE WE GET A DATAFRAME grouped_df WHICH IS MORE LIKE A DAILY TRADEBOOK
            new_df = (temp_df_grouped.sort_values('Trade Date', ascending=True))
            # converts the multiline index to normal DF
            grouped_df = new_df.reset_index()
            grouped_df.drop("Investment", axis=1, inplace=True)
            tradebook_day = pd.concat([tradebook_day,grouped_df])

        return tradebook_day

# Create a function to map YCODE to live prices with a default value
def map_ycode_to_live_price(ycode):
    return live_prices.get(ycode, 0)

# Create a function to update live prices
def update_live_prices(final_pf):
    while True:
        # Fetch live prices and update the DataFrame
        for index, row in final_pf.iterrows():
            ycode = row['YCODE']
            live_price = fromyahoo.liveprice(ycode)
            final_pf.at[index, 'LTP'] = live_price
        # Sleep for a specified interval (e.g., every 5 seconds)
        time.sleep(5)

@st.cache_data
def createpf(xl):
    if isinstance(xl,pd.DataFrame):
        show_only = {}
        show_only['Symbol'] = []
        show_only['Quantity'] = []
        show_only['Price'] = []

        # DEFINING the dataframes ENDRESULT - its column names
        open_pf = pd.DataFrame(columns=['ISIN','Trade Date', 'Trade Type', 'Quantity', 'avg_price'])  # EMPTY DF
        closed_pf = pd.DataFrame({'ISIN' : [] ,'Buy Date': [], 'Sell Date': [], 'Quantity': [], 'Buy Price': [], 'Sell Price': [], 'PnL': []})
        only_sell_pf = pd.DataFrame(columns=['ISIN','Trade Date', 'Trade Type', 'Quantity', 'avg_price'])
        sold_data = []
        buy_queue = []  # to save all buy transactions temporarily, {} values in a list
        symbol_isin = {}
        # NOW LETS GROUP EVERY DF BY 'ISIN' instead of 'SYMBOL'

        for each in xl['ISIN'].unique():
            grouped_df = xl.groupby("ISIN").get_group(each)
            buy_queue = []
            remaining_sell_quantity = 0
            # we will try to get 2 seperate dataframes from the grouped_df : 1 is sold_df and one is current_df

            for index, row in grouped_df.iterrows():
                if (grouped_df['Trade Type'] == 'sell').all() or grouped_df['Trade Type'].iloc[0] == 'sell':
                    only_sell_pf = pd.concat([only_sell_pf,grouped_df],ignore_index=True)
                else:  # ensure the first row trate_type is not "sell" and the entire grouped df doesnt have only sell orders
                    trade_date, trade_type, quantity, avg_price = row['Trade Date'], row['Trade Type'], row['Quantity'], row['avg_price']
                    if trade_type == 'buy':
                        # Add 'buy' transactions to the queue
                        buy_queue.append(
                            {'ISIN': each, 'Trade Date': trade_date, 'Trade Type': trade_type, 'Quantity': quantity,
                             'avg_price': avg_price})
                    elif trade_type == 'sell':
                        # Process 'sell' transactions
                        remaining_sell_quantity = quantity
                        i = 0  # whenever sell comes, it shud start from the first index of buy_queue
                        # Check if there are 'buy' transactions in the queue to offset the 'sell' quantity
                        while remaining_sell_quantity > 0 and (buy_queue):
                            # get data from queue
                            buy_transaction = buy_queue[0]  # Get the 'buy' transaction at index i
                            buy_date = buy_transaction['Trade Date']
                            buy_qty = buy_transaction['Quantity']
                            buy_price = buy_transaction['avg_price']
                            # get current sell data
                            sell_date = trade_date
                            sell_qty = quantity
                            sell_price = avg_price

                            if buy_qty == remaining_sell_quantity:
                                # print(f"SELL {remaining_sell_quantity} quantity @{sell_price} ")
                                # manage sold quantity
                                # sold_data = {'Symbol':each,'Buy Date':buy_date,'Sell Date':sell_date, 'Quantity': sell_qty, 'Buy Price': buy_price, 'Sell Price':sell_price, 'PnL':(sell_price-buy_price)*sell_qty}
                                # closed_pf = closed_pf.append(sold_data, ignore_index=True)
                                sold_data.append(
                                    {'ISIN': each, 'Buy Date': buy_date, 'Sell Date': sell_date, 'Quantity': sell_qty,
                                     'Buy Price': buy_price, 'Sell Price': sell_price,
                                     'PnL': (sell_price - buy_price) * sell_qty})
                                # closed_pf = pd.DataFrame(sold_data)
                                # manage buy_queue
                                remaining_sell_quantity -= buy_qty
                                buy_queue.pop(0)

                            elif buy_qty > remaining_sell_quantity:  # 445>336
                                # print(f"SELL {remaining_sell_quantity} quantity @{sell_price} ")
                                # If the 'buy' quantity is greater or equal to the remaining 'sell' quantity, adjust the 'buy' quantity
                                # manage sold quantity
                                # sold_data = {'Symbol':each,'Buy Date': buy_date, 'Sell Date': sell_date, 'Quantity': remaining_sell_quantity,'Buy Price': buy_price, 'Sell Price': sell_price, 'PnL': (sell_price - buy_price)*remaining_sell_quantity}
                                sold_data.append({'ISIN': each, 'Buy Date': buy_date, 'Sell Date': sell_date,
                                                  'Quantity': remaining_sell_quantity, 'Buy Price': buy_price,
                                                  'Sell Price': sell_price,
                                                  'PnL': (sell_price - buy_price) * remaining_sell_quantity})
                                # closed_pf = pd.DataFrame(sold_data)

                                # manage buy_queue
                                buy_queue[0][
                                    'Quantity'] -= remaining_sell_quantity  # this checks the remaining buy_qty and update in the buy_queue
                                remaining_sell_quantity = 0  # checks remaining sell qty to ensure the while loop continues

                            else:  # if buy_qty<sell_qty               114,800 < 802
                                # print(f"SELL {remaining_sell_quantity} quantity @{sell_price} ")
                                # print("$$$$$$$$$$ BUY QUEUE IS : $$$$$$$$$$$$$$$")
                                # print(buy_queue[0])
                                # If the 'buy' quantity is less than the remaining 'sell' quantity, adjust the 'sell' quantity
                                # sold_data = {'Symbol':each,'Buy Date': buy_date, 'Sell Date': sell_date, 'Quantity': buy_qty,'Buy Price': buy_price, 'Sell Price': sell_price, 'PnL': (sell_price-buy_price)*buy_qty}
                                # closed_pf = closed_pf.append(sold_data, ignore_index=True)
                                sold_data.append(
                                    {'ISIN': each, 'Buy Date': buy_date, 'Sell Date': sell_date, 'Quantity': buy_qty,
                                     'Buy Price': buy_price, 'Sell Price': sell_price,
                                     'PnL': (sell_price - buy_price) * buy_qty})
                                # Manage buy_queue now to get the next queue
                                buy_queue.pop(0)
                                remaining_sell_quantity -= buy_qty

            # Append any remaining 'buy' transactions to the result DataFrame
            open_pf = pd.concat([open_pf,pd.DataFrame(buy_queue)],ignore_index=True)
            closed_pf = pd.DataFrame(sold_data)

        # open_pf is the dataframe which needs to be worked upon if addition of Stock ttrades is happening
        open_pf['Investment'] = open_pf['Quantity'] * open_pf['avg_price']
        temp_open_pf = open_pf.groupby(["ISIN","Trade Type"])[['Quantity', 'Investment']].sum()  # .reset_index()
        temp_open_pf["Price"] = temp_open_pf["Investment"] / temp_open_pf["Quantity"]
        # converts the multiline index to normal DF
        final_pf = temp_open_pf.reset_index()
        # Create a dictionary mapping 'Symbol' values to 'PnL' values from 'closed_pf'
        symbol_pnl_mapping = closed_pf.set_index('ISIN')['PnL'].to_dict()
        # Use the map function to assign 'PnL' values to 'final_pf' based on 'Symbol' matching
        final_pf['prev_pnl'] = final_pf['ISIN'].map(symbol_pnl_mapping)
        final_pf['prev_pnl'].fillna('0',inplace=True)
        final_pf['prev_pnl'] = pd.to_numeric(final_pf['prev_pnl'], errors='coerce')

        #LETS COMBINE ISIN AND CODE PROVIDED BY ZERODHA
        final_pf['CODE'] = final_pf['ISIN'].apply(nse_bse_search.isin_to_code)
        final_pf['YCODE'] = final_pf['ISIN'].apply(nse_bse_search.isin_to_ycode)

        #final_pf['LTP'] = final_pf['YCODE'].apply(fromyahoo.liveprice)
        final_pf['CLosingPrice'] = final_pf['CODE'].apply(lastdayprice.getltp)
        final_pf['FreeShares'] = final_pf['prev_pnl'] / final_pf['CLosingPrice']

        closed_pf['CODE'] = closed_pf['ISIN'].apply(nse_bse_search.isin_to_code)
        closed_pf['YCODE'] = closed_pf['ISIN'].apply(nse_bse_search.isin_to_ycode)
        #closed_pf['LTP'] = closed_pf['YCODE'].apply(fromyahoo.liveprice)
        closed_pf['CLosingPrice'] = closed_pf['CODE'].apply(lastdayprice.getltp)
        closed_pf['FreeShares'] = closed_pf['PnL']/closed_pf['CLosingPrice']
        closed_pf['PnL_toDate'] = ((closed_pf['CLosingPrice'] - closed_pf['Buy Price']) * closed_pf['Quantity'])
        #final_pf['LTP'] = ['Loading'] * len(final_pf)
        ## THIS IS WORKING CODE TO GET THE PRICES FROM YAHOO CODE
        #final_pf_copy = final_pf.dropna(subset=['YCODE'])
        #closed_pf_copy = closed_pf.dropna(subset=['YCODE'])
        #unique_ycode = np.unique(np.concatenate([final_pf_copy['YCODE'].unique(), closed_pf_copy['YCODE'].unique()]))
        ## Use parallel processing to fetch live prices for unique scripts
        #with concurrent.futures.ThreadPoolExecutor() as executor:
        #    results = executor.map(fromyahoo.liveprice, unique_ycode)
        #    for script, price in zip(unique_ycode, results):
        #        live_prices[script] = price

        ## Create a new DataFrame with live prices, handling missing values
        #final_pf['LTP'] = final_pf['YCODE'].map(map_ycode_to_live_price)
        ## Assuming 'final_pf' is your initial DataFrame
        #closed_pf['LTP'] = closed_pf['YCODE'].map(map_ycode_to_live_price)

        #final_pf['FreeShares'] = final_pf['prev_pnl'] / final_pf['LTP']
        #final_pf = final_pf.round()

        #closed_pf['PnL_toDate'] = ((closed_pf['LTP'] - closed_pf['Buy Price'])*closed_pf['Quantity']) - closed_pf['PnL']

        #closedpf_pnl = closed_pf['PnL'].sum()
        #st.info(f'Total Profit of Loss for CLOSED POrtfolio is {closedpf_pnl}')

        #only_sell_pf['CODE'] = only_sell_pf['ISIN'].apply(nse_bse_search.isin_to_code)
        ##only_sell_pf['LTP'] = only_sell_pf['CODE'].apply(lastdayprice.getltp)


        return final_pf,closed_pf,only_sell_pf

# Define a custom function to calculate 'Allowed' based on 'Broker limit reached'
def calculate_allowed(row):
    if row['Broker limit reached'] == 'No':
        return (100 - (row['Haircut %']))
    else:
        return 0

subcol1,subcol2 = st.columns([8,2])

if portfolio_option == "Portfolio":
    #loc = "./TRADEBOOK.xlsx"
    if tradebook is not None:
        start_time = time.time()
        # PLEDGING EXCEL FILE - PART1
        ISIN_haircut_mapping = None
        # PROCESSING THE PLEDGING FILE FROM ZERODHA
        pledging_xl = pd.read_excel(pledging_path)
        pledging_xl['Allowed'] = pledging_xl.apply(calculate_allowed, axis=1)
        # st.dataframe(pledging_xl)
        # Create a dictionary mapping 'Symbol' values to 'PnL' values from 'closed_pf'
        ISIN_haircut_mapping = pledging_xl.set_index('ISIN')['Allowed'].to_dict()

        i=0
        orig_xl = pd.DataFrame()
        for each_xl in tradebook:
            start_row = None
            book = openpyxl.load_workbook(each_xl)
            datasht = book["Equity"]
            #st.info(datasht['C7'].value)
            for i in range(1, datasht.max_row + 1):
                if datasht['B' + str(i)].value == "Symbol":
                    start_row = i-1
                    break
            reqd_cols = "B :" + str(get_column_letter(datasht.max_column))
            if start_row is not None:
                #st.info(datasht['B'+str(start_row)].value)
                # Load each file into a DataFrame
                df = pd.read_excel(each_xl,header=start_row,usecols=reqd_cols)
                # Concatenate the cleaned data to the combined DataFrame
                orig_xl = pd.concat([orig_xl, df])
                orig_xl = orig_xl.dropna(axis=1)
                #st.dataframe(df)
        # Remove duplicates based on all columns
        orig_xl = orig_xl.drop_duplicates()

        if 'ISIN' in orig_xl.columns:
            # Continue with your processing logic
            orig_xl = orig_xl.sort_values('Trade Date', ascending=True)
            # THIS needs to be SAVED in DATABASE in the Signed-IN Username
            #           ("Original DF from all the uploaded files is stored as orig_xl dataframe in the code")
            #st.dataframe(orig_xl)

            # Get a tradebook where mulitple exeuction on same day is combined to one
            tradebook_daily = tradebook_perday(orig_xl)

            #this download_tradebook is especially for formatting the tradebook in downloadable format
            download_tradebook = tradebook_daily.copy()
            download_tradebook['YCODE'] = download_tradebook['ISIN'].apply(nse_bse_search.isin_to_ycode)
            #st.dataframe(download_tradebook)
            #final_pf['YCODE'] = final_pf['ISIN'].apply(nse_bse_search.isin_to_ycode)

            show_pf, show_closed_pf, show_only_sell_pf =createpf(tradebook_daily)
            # this is to only SHOW the users by replacing ISIN to Symbol Name
            Open_Portfolio, CLosed_Portfolio,Make_SIP = st.tabs(["Open Position", "Closed Position","Make SIP"])
            #st.info("OPEN PORTFOLIO : ")

            with subcol2:
                fundamentals.excel_link_to_download(tradebook_daily, "Download History.xlsx", "Download Tradebook")
                if st.button("Download Holdings Txt"):
                    pf_for_txt = show_pf.copy().sort_values('Investment', ascending=False)
                    with open(f"./tradebook {today.strftime('%d%b%Y')}.txt",'w') as w:
                        for entry in pf_for_txt["YCODE"]:
                            w.write(f"{entry}\n")
                    with open(f"./tradebook {today.strftime('%d%b%Y')}.txt", 'r') as file:
                        txt_data = file.read()
                        st.download_button(label="Download_Now", data=txt_data, file_name=f"holdings_{today.strftime('%d%b%Y')}.txt", mime="text/plain")


            with Open_Portfolio:
                # Use the map function to assign 'PnL' values to 'final_pf' based on 'Symbol' matching
                #show_pf["Investment"] = show_pf["Quantity"] * show_pf["Price"]
                show_pf['Allowed %'] = show_pf['ISIN'].map(ISIN_haircut_mapping)  ##'Broker limit reached' == 'No'
                show_pf['ValuationNow'] = show_pf['Quantity'] * show_pf['CLosingPrice']
                show_pf['CanPledge'] = (show_pf['Allowed %'] / 100) * show_pf['ValuationNow']

                show_pf1 = show_pf[["CODE","YCODE","Quantity","Price","Investment","CLosingPrice","ValuationNow","FreeShares","CanPledge"]].copy()
                st.dataframe(show_pf1.set_index('CODE').sort_values('Investment',ascending=False), use_container_width=True)
                st.info("INVALID ENTRIES : ")
                st.dataframe(show_only_sell_pf, use_container_width=True)
                # show_only_sell_pf['Symbol'] = show_only_sell_pf['Symbol'].replace(symbol_isin)
                # print(show_only_sell_pf.set_index('Symbol'))
                # st.dataframe(show_only_sell_pf.set_index('Symbol'))
                Total_Collateral = show_pf1['CanPledge'].sum()
                st.info(f'You can pledge a total of {round(Total_Collateral)}rs')
                if st.button("Download OpenHoldings Txt"):
                    pf_for_txt = show_pf.copy().sort_values('Investment', ascending=False)
                    with open(f"./openholdings {today.strftime('%d%b%Y')}.txt",'w') as w:
                        for entry in pf_for_txt["YCODE"]:
                            w.write(f"{entry}\n")
                    with open(f"./openholdings {today.strftime('%d%b%Y')}.txt", 'r') as file:
                        txt_data = file.read()
                        st.download_button(label="Download_Now", data=txt_data, file_name=f"holdings_{today.strftime('%d%b%Y')}.txt", mime="text/plain")

                download_tradebook = download_tradebook[['YCODE','Quantity','Trade Date','avg_price','Trade Type']]
                download_tradebook['Trade Date'] = pd.to_datetime(download_tradebook['Trade Date'])
                download_tradebook['Trade Date'] = download_tradebook['Trade Date'].dt.strftime('%d-%m-%Y')
                #st.dataframe(download_tradebook)
                download_tradebook = download_tradebook.rename(
                    columns={'YCODE': 'Symbol', 'Trade Date': 'BuyDate', 'Trade Type': 'Type', 'avg_price': 'BuyPrice'})
                fundamentals.excel_link_to_download(download_tradebook, "Tradebook Marketsmith.xlsx", "Download MarketSmith Format")

            #st.info("CLosed Portfolio")
            with CLosed_Portfolio:
                st.dataframe(show_closed_pf.set_index('CODE').sort_values('Sell Date',ascending=True), use_container_width=True)
                closedpf_pnl = round(show_closed_pf['PnL'].sum()/100000,1)
                closedpf_pnl_open = round(show_closed_pf['PnL_toDate'].sum()/100000,1)
                st.info(f'Realised Profit/Loss is {closedpf_pnl}Laks')
                st.info(f'If held all your closed positions till now, PnL would have been {closedpf_pnl_open}Laks')
                if (closedpf_pnl_open>closedpf_pnl):
                    st.error(f"So you would have made more {round(closedpf_pnl_open-closedpf_pnl,1)}Laks if positions were kept Open.")
                    st.info("You should even consider the fact that, non-rotating cash amoung ur stocks mean, You need to do SIP at your entry levels")
                else:
                    st.succes("You made better returns by doing Positional Investment. Keep Going")
                # show_closed_pf['Symbol'] = show_closed_pf['Symbol'].replace(symbol_isin)
                # print(show_closed_pf.set_index('Symbol'))
                # st.dataframe(show_closed_pf.set_index('Symbol'))

            with Make_SIP:
                sip_investment = st.slider(label="What if you SIPped on your closed Portfolio ? Chose your SIP amount per stock :", min_value=1000, max_value=100000, value=2000, step=1000)
                #st.text_input(label="Enter Principal per stock to know your SIP value now")
                if st.button('Show SIP'):
                    sip_pf = show_closed_pf[['CODE','Buy Date','Buy Price','CLosingPrice']].copy()
                    #sip_pf['Invested'] = sip_investment
                    sip_pf.loc[:, 'Invested'] = sip_investment
                    sip_pf['PnL'] = sip_pf['CLosingPrice'] * (sip_pf['Invested'] / sip_pf['Buy Price'])
                    sip_pf = sip_pf.drop_duplicates(subset=['CODE', 'Buy Date'])
                    st.dataframe(sip_pf.sort_values('Buy Date',ascending=True))
                    SIP_totCapital = sip_pf['Invested'].sum()
                    SIP_PnL = sip_pf['PnL'].sum()
                    return_on_SIP = round(((SIP_PnL - SIP_totCapital) / SIP_totCapital)*100)
                    st.success(f"Your SIP generated Profit/Loss of {round(SIP_PnL/100000,2)}lak a return of ({return_on_SIP}%) on your Total SIP Investment of {round(SIP_totCapital/100000,2)}lak")
                    # Group by year and month and calculate total investment
                    result=pd.DataFrame()
                    sip_pf['Buy Date'] = pd.to_datetime(sip_pf['Buy Date'])
                    result['Year'] = sip_pf['Buy Date'].dt.year
                    result['Month'] = sip_pf['Buy Date'].dt.month
                    result['Invested'] = sip_pf['Invested']
                    result = result.groupby([ 'Year' , 'Month' ])['Invested'].sum().reset_index()
                    # Rename the columns for clarity
                    #result = result.rename(columns={'Buy Date': 'Year', 'Buy Date': 'Month', 'Invested': 'Total Investment'})
                    st.error(f"By the way : MAX SIP amount per month went upto {result['Invested'].max()} and minimum of {result['Invested'].min()}. Average SIP amount would be {round(result['Invested'].mean())}")
                    with st.expander("See Your Total Monthly SIP Amount PER MONTH"):
                        st.dataframe(result)

            # IMPROVEMENTS
            #any new upload of the excel sheet shud only append the initial dataframe xl
            end_time = time.time() - start_time
            st.info(f"Downloaded in {end_time} sec")
        else:
            # Handle the case where 'ISIN' is not present in the DataFrame
            st.success("Make Sure to Upload from the Start to avoid Malfunctioning")

        # Use a separate thread to continuously update live prices
        #thread = threading.Thread(target=update_live_prices, args=(show_pf,))
        #thread.start()

if portfolio_option == "Collateral":
    colx,coly = st.columns([2,2])
    coly.subheader("💡 Key Benefits:")
    coly.write("✅ Discover which of your stocks are eligible for collateral")
    coly.write("💰 Maximize your trading potential")
    coly.write("💹 Leverage your existing investments")
    colx.subheader("🔍 How it works:")
    holding_url = 'https://console.zerodha.com/portfolio/holdings'
    st.markdown(f"[***ZERODHA HOLDINGS***]({holding_url})", unsafe_allow_html=True)
    holding = colx.file_uploader("Upload your holdings Excel file", type= ['xlsx'])
    ISIN_haircut_mapping = None
    if holding is not None:
                    # PROCESSING THE PLEDGING FILE FROM ZERODHA
        pledging_xl = pd.read_excel(pledging_path)
        pledging_xl['Allowed'] = pledging_xl.apply(calculate_allowed, axis=1)
        #st.dataframe(pledging_xl)
        # Create a dictionary mapping 'Symbol' values to 'PnL' values from 'closed_pf'
        ISIN_haircut_mapping = pledging_xl.set_index('ISIN')['Allowed'].to_dict()
                    # PROCESSING THE UPLOADED FILE

        i=0
        orig_xl = pd.DataFrame()
        start_row = None
        book = openpyxl.load_workbook(holding)
        datasht = book["Equity"]
        for i in range(1, datasht.max_row + 1):
            if datasht['B' + str(i)].value == "Symbol":
                start_row = i-1
                break
        reqd_cols = "B :" + str(get_column_letter(datasht.max_column))
        if start_row is not None:
            #st.info(datasht['B'+str(start_row)].value)
            # Load each file into a DataFrame
            orig_xl = pd.read_excel(holding,header=start_row,usecols=reqd_cols)
            # Concatenate the cleaned data to the combined DataFrame
            orig_xl = orig_xl.dropna(axis=1)
        # Remove duplicates based on all columns
        orig_xl = orig_xl.drop_duplicates()
                    #ProcessPortfolio.createpf(xl)
        if 'ISIN' in orig_xl.columns:
            # Use the map function to assign 'PnL' values to 'final_pf' based on 'Symbol' matching
            orig_xl['Allowed %'] = orig_xl['ISIN'].map(ISIN_haircut_mapping)                  ##'Broker limit reached' == 'No'
            orig_xl['Valuation'] = orig_xl['Quantity Available']*orig_xl['Previous Closing Price']
            orig_xl['CanPledge'] = (orig_xl['Allowed %']/100)*orig_xl['Valuation']
            # Check if 'Valuation' is greater than 50000
            #mask = orig_xl['Valuation'] > 50000
            # Apply the operation only where the condition is met
            #orig_xl.loc[mask, 'CanPledge'] = (1 - (orig_xl['Haircut'] / 100)) * orig_xl.loc[mask, 'Valuation']
            show_pledging = orig_xl[['Symbol', 'CanPledge']].copy()
            # Drop rows where 'CanPledge' is None (missing values)
            show_pledging = show_pledging.dropna(subset=['CanPledge'])
            st.dataframe(show_pledging)
            Total_Collateral = orig_xl['CanPledge'].sum()
            st.info(f'You can pledge a total of {round(Total_Collateral)}rs')
        else:
            # Handle the case where 'ISIN' is not present in the DataFrame
            st.warning("Upload from Zerodha Console")

