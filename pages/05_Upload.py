import streamlit as st
from streamlit_option_menu import option_menu
import streamlit.components.v1 as html
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import base64  # Standard Python Module
from io import StringIO, BytesIO  # Standard Python Module
import os
import datetime
import fundamentals
import json
from streamlit_lottie import st_lottie
from telegram import Bot
import nse_bse_search

global bsecodenum_codename
global bsecodename_codenum
# bsecodenum_codename, bsecodename_codenum = nse_bse_search.bsecodenum_bsecodename()
bsecodenum_codename, bsecodename_codenum, bsecodenum_fullname, bsecodename_fullname, bsefullname_codenum, bsefullname_codename = nse_bse_search.bsecodenum_bsecodename()
# This gets us the BSE NAME from the DAILY BHAVCOPY THAT WE ARE DOWNLOADING
bsesccode_scname,bsescname_sccode = nse_bse_search.bseSCNAME_SCCODE()

#with st.expander("BSESCCODE_SCNAME"):
#st.info(bsesccode_scname.keys())

#gets nse codes as list
nse_code_list = []
with open('./nselist.txt', 'r') as file:
    nse_code_list = [line.strip() for line in file.readlines()]
    #nse_code_list = file.readlines()

st.set_page_config(page_title="Upload", page_icon=":bar_chart:", layout="wide",initial_sidebar_state="expanded",)
token_jarvis = "1698319688:AAG5X-bmCzGqWHIyaksIUfBG_rxZRE3tUvI"                     # JarvisPOSTME
chat_id = "@itimesAlgo_d"
#chat_id = "@itimesalgo"
bot = Bot(token=token_jarvis)
TELEGRAM_API_URL = f"https://api.telegram.org/bot{token_jarvis}/sendMessage"

#color_dict = {'Yellow_Lite':"#f8ba43",'Yellow_Dark':"#D6D41B",'Blue_Lite':"#1959BF",'Blue_Dark':"#0971C9",'Green_Lite':"#11A694",'Green_Dark':"#11A64B",}   #"Purple_Lite":"#7019BF",'Purple_Dark':"#9319BF"}
color_dict = {'blue3':{'hash':'#00A3FE','rgb':'rgb(0,163,254)'},
              'yellow1':{'hash':'#FFFF01','rgb':'rgb(255,255,1)'},
              'blue1':{'hash':'#21A1E1', 'rgb':'rgb(33,161,225)'},
              'yellow2':{'hash':'#FFFE57','rgb':'rgb(255,254,87)'},
              'blue2':{'hash':'#5DB7D2','rgb':'rgb(93,183,210)'},
              'green1':{'hash':'#00F954','rgb':'rgb(0,249,84)'},
              'red1':{'hash':'#CC0118','rgb':'rgb(204,1,24)'},
              'black': {'hash': '000000', 'rgb': 'rgb(0,0,0)'},
              'white': {'hash': '#ffffff', 'rgb': 'rgb(255, 255, 255)'},
              }

def load_lottiefile(filepath: str):
    with open(filepath, "r") as f:
        return json.load(f)
def load_lottieurl(url: str):
    r = requests.get(url)
    if r.status_code != 200:
        return None
    return r.json()
lottie_data_analytics = load_lottiefile("./lottie/data-analytics.json")

with st.sidebar:
    uploaded_file = st.file_uploader("🔗 Upload here :", type=['xlsx','xlsm'],accept_multiple_files = True)  # Only accepts xlsx,xlsm file format

    st.markdown(""" <style> .font {
    font-size:22px ; font-family: 'Cooper Black'; color: #FF9633;} 
    </style> """, unsafe_allow_html=True)
    # Add a file uploader to allow users to upload their csv file

    st.markdown('<p class="font">Upload One or Two xlsx/xlsm FILES from screener.in </p>',
                unsafe_allow_html=True)
    st_lottie(
        lottie_data_analytics,
        speed=0.7,
        reverse=False,
        loop=True,
        quality="low",  # medium ; high
        height=None,
        width=None,
        key="barchart", )

def process_code(code,code_name):
    if code.isdigit():
        code = int(code)
        # for sure we have code in bsecodenum_name
        if code in bsesccode_scname.keys():
            # st.info(f"{code} available in SCCODE-SCNAME")
            if bsecodenum_codename[code] in nse_code_list:
                # st.info(f"{code} BSECODE given - Found in nselist also")
                code_names = [code, bsesccode_scname[code].strip(), bsecodenum_codename[code].strip()]
                st.info(f"{code_names}")
            else:
                # st.info(f"{code} not found in scname")
                code_names = [code, bsesccode_scname[code].strip()]
                st.info(f"BSECODE given - not in NSE {code_names}")
        elif code in bsecodenum_codename.keys():
            code_names = [code, bsecodenum_codename[code].strip()]
            st.info(f"BSECODE given - Notfound in BSESCCODE Found in BSECODENUM_CODENAME {code_names}")
        else:
            code_names = [code, code_name.strip()]
            st.info(f"BSECODE given - Notfound anywhere so sticking to code and code_name {code_names}")
    else:
        if code in bsecodename_codenum.keys():
            if bsecodename_codenum[code] in bsesccode_scname.keys():
                code_names = [code, str(bsecodename_codenum[code]).strip(),
                              bsesccode_scname[bsecodename_codenum[code]].strip()]
                st.info(f"NSECODE given - Found in BSECODENAME_CODENUM and BSESCCODE as well : {code_names}")
            else:
                code_names = [code, str(bsecodename_codenum[code]).strip()]
                st.info(f"NSECODE given - Found in BSECODENAME_CODENUM and not found in BSESCCODE : {code_names}")
        else:
            code_names = [code]
            st.info(f"NSECODE given - not in BSE : {code_names}")
        code = code.upper()
    return code_names

st.title("📈 Don't See the Latest Charts?")
if uploaded_file is not None:
    if len(uploaded_file)==1:
        sentence = ""
        sentence1 = ""
        statement = ""
        statement1 = ""
        Yearly_sentence_in_Quarterly = ""
        comp_Name = uploaded_file[0].name.split('.xlsx')[0]
        comp_Name = comp_Name.split('.xlsm')[0]

        sentence = f"{comp_Name}: "
        col1, col2, col3, col4 = st.columns([2, 2, 1, 1])
        with col1:
            st.subheader('📊 ' + comp_Name)
        with col4:
            color_key = st.selectbox("Bar Color",color_dict.keys())
            #color_bar = st.color_picker("Bar Color", value="#ECE80F")  # blueshades"#0971C9""ECE80F"   #yellowshades"#f8ba43"
        st.write("____")
        book = openpyxl.load_workbook(uploaded_file[0])
        #comp_name = book['Data Sheet']['B1'].value
        qtr_pnl,df_comp = fundamentals.get_tables(book[fundamentals.tabs[-1]], uploaded_file[0])  # send a sheet(not whole workbook)
        #pnl, balancesht,qtr_pnl = fundamentals.develop_data(qtr_pnl,df_comp)
        pnl, balancesht = fundamentals.develop_yearly(df_comp)
        qtr_pnl = fundamentals.develop_quarterly(qtr_pnl)
        try:
            df_comp.columns = df_comp.columns.strftime('%d-%m-%Y')
        except Exception as AttributeError:
            pass
        try:
            qtr_pnl.columns = qtr_pnl.columns.strftime('%d-%m-%Y')
        except Exception as AttributeError:
            pass
        # **************************************************************************************************
        colx, coly,colw,colz = st.columns([2,1,2,1])
        with colx:
            save_code1 = st.text_input(label='Comp Code')
        with coly:
            if st.button('Pickle') and save_code1 :
                if save_code1.isdigit():
                    company_code = int(save_code1)
                    save_code = bsecodenum_codename[company_code]
                else:
                    save_code = str(save_code1.upper())
                first_letter = save_code[0].upper()
                alphabetic_folder = os.path.join("./pickl/", first_letter)
                # Create the folder if it doesn't exist
                if not os.path.exists(alphabetic_folder):
                    os.makedirs(alphabetic_folder)
                df_comp.to_pickle("./pickl/" + first_letter + '/' + save_code + " Yearly.pkl")
                qtr_pnl.to_pickle("./pickl/" + first_letter + '/' + save_code + " Quarterly.pkl")
                st.info(f"Saved your pickle file as {save_code}")

                #SAVE TEXT FILES AND CSV FILES
                code_names = process_code(save_code1,save_code)
                if len(code_names) == 1:
                    sentence1 += f"{str(code_names[0])} "
                else:
                    for each in code_names:
                        sentence1 += f"{str(each)} "
                if not df_comp.empty and df_comp is not None:
                    yr_df = df_comp.copy()
                    yr_df.columns = pd.to_datetime(yr_df.columns, format='%d-%m-%Y')
                    # TO SAVE IN AMIBROKER CSV
                    developed_df = pd.concat([pnl, balancesht, yr_df.loc["CASH FLOW", :]], axis=0)
                    for each in code_names:
                        each = str(each)
                        sentence = sentence1
                        amibroker_txt = "./amibroker/notes/" + each + ".txt"
                        amibroker_csv = "./amibroker/csv/" + each + " Yearly.csv"
                        try:
                            sentence += f"\n**********YEARLY**********\n"
                            if len(developed_df.columns)>=2:
                                last_year = developed_df.columns[-1]
                                prev_year = developed_df.columns[-2]
                                get_from_yearly =["SALES","NET PROFIT","NPM %","NO. OF EQUITY SHARES","FACE VALUE","DEBTOR DAYS","INVENTORY TURNOVER","ROCE","RESERVES","BORROWINGS"]
                                eq_last_year = round(developed_df.loc["NO. OF EQUITY SHARES", last_year]/10000000,2)
                                eq_prev_year = round(developed_df.loc['NO. OF EQUITY SHARES', prev_year]/10000000,2)
                                FV_last_year = round(developed_df.loc['FACE VALUE', last_year])
                                FV_prev_year = round(developed_df.loc['FACE VALUE', prev_year])
                                ROCE_last_year = developed_df.loc['ROCE', last_year]
                                ROCE_prev_year = developed_df.loc['ROCE', prev_year]
                                Yearly_sentence_in_Quarterly = f"Equity in {str(datetime.datetime.strftime(last_year,'%b-%Y'))}: {str(eq_last_year)}cr; in {str(datetime.datetime.strftime(prev_year,'%b-%Y'))}: {str(eq_prev_year)}cr\nFaceValue in {str(datetime.datetime.strftime(last_year,'%b-%Y'))}: {str(FV_last_year)}; in {str(datetime.datetime.strftime(prev_year,'%b-%Y'))}: {str(FV_prev_year)}\nROCE in {str(datetime.datetime.strftime(last_year,'%b-%Y'))}: {str(ROCE_last_year)}%; in {str(datetime.datetime.strftime(prev_year,'%b-%Y'))}: {str(ROCE_prev_year)}%\n"
                                st.info(Yearly_sentence_in_Quarterly)

                            sentence += fundamentals.stmt_for_qoq(pnl)
                            #st.info("trying to open amiboker txt file")
                            with open(amibroker_txt, "a") as f:
                                f.write(sentence)
                                f.write("\n")
                            # st.success(f"Written in {amibroker_txt} file")
                        except Exception as e:
                            st.error(f"Getting error {e} while trying to write txt file")

                        try:
                            # st.dataframe(developed_df)
                            # trying to save the data in CSV file to load in AMIBROKER
                            #yr_df2.to_csv(amibroker_csv)
                            developed_df.to_csv(amibroker_csv)
                        except Exception as e:
                            st.error(f"Getting error {e} while trying to write csv files ")

                if not qtr_pnl.empty and isinstance(qtr_pnl, pd.DataFrame) and qtr_pnl is not None:
                    qtr_df = qtr_pnl.copy()
                    qtr_pnl.columns = pd.to_datetime(qtr_pnl.columns, format='%d-%m-%Y')
                    for each in code_names:
                        each = str(each)
                        sentence = sentence1
                        amibroker_txt = "./amibroker/notes/" + each + ".txt"
                        amibroker_csv = "./amibroker/csv/" + each + " Quarterly.csv"
                        try:
                            sentence += f"\n**********QUARTERLY**********\n"
                            sentence += f"Update Date: {datetime.datetime.strftime(datetime.datetime.now(), '%d-%b-%Y')}\n"
                            sentence += Yearly_sentence_in_Quarterly
                            sentence += f"********************\n"
                            sentence += fundamentals.stmt_for_qoq(qtr_pnl)
                            with open(amibroker_txt, "a") as f:
                                f.write(sentence)
                                f.write("\n")
                            # st.success(f"Written in {amibroker_txt} file")
                        except Exception as e:
                            st.error(f"Getting error {e} while trying to write txt file")
                        try:
                            # trying to save the data in CSV file to load in AMIBROKER
                            qtr_df.to_csv(amibroker_csv)
                        except Exception as e:
                            st.error(f"Getting error {e} while trying to write csv files ")
        with col2:
            sub_choose = st.selectbox("Fundamentals", fundamentals.funda_menu)

        if sub_choose == 'Key_Data':
            key_data = str("""<!-- TradingView Widget BEGIN -->
                                    <div class="tradingview-widget-container">
                                      <div class="tradingview-widget-container__widget"></div>
                                      <div class="tradingview-widget-copyright"><a href="https://www.tradingview.com/symbols/NASDAQ-AAPL/financials-overview/" rel="noopener" target="_blank"><span class="blue-text">Fundamental Data</span></a> by TradingView</div>
                                      <script type="text/javascript" src="https://s3.tradingview.com/external-embedding/embed-widget-financials.js" async>
                                      {
                                      "colorTheme": "dark",
                                      "isTransparent": false,
                                      "largeChartUrl": "",
                                      "displayMode": "regular",
                                      "width": "100%",
                                      "height": 880,
                                      "symbol": "xx",
                                      "locale": "en"
                                      }
                                      </script>
                                    </div>
                                    <!-- TradingView Widget END -->
                                """)

            comp_profile = str("""
                                    <!-- TradingView Widget BEGIN -->
                                    <div class="tradingview-widget-container">
                                      <div class="tradingview-widget-container__widget"></div>
                                      <div class="tradingview-widget-copyright"><a href="https://www.tradingview.com/symbols/NASDAQ-AAPL/" rel="noopener" target="_blank"><span class="blue-text"> Profile</span></a> by TradingView</div>
                                      <script type="text/javascript" src="https://s3.tradingview.com/external-embedding/embed-widget-symbol-profile.js" async>
                                      {
                                      "width": "100%",
                                      "height": 880,
                                      "colorTheme": "dark",
                                      "isTransparent": false,
                                      "symbol": "xxyy",
                                      "locale": "en"
                                    }
                                      </script>
                                    </div>
                                    <!-- TradingView Widget END -->
                                    """)
            with st.expander(label="TRADINGVIEW DATA"):
                colx, coly = st.columns([1.5, 1])
                with colx:
                    components.html(key_data.replace("xx", comp_Name), height=1080)
                with coly:
                    components.html(comp_profile.replace("xxyy", comp_Name), height=1080)
            with st.expander(label='BALANCE SHEET'):
                st.dataframe(balancesht)
            with st.expander(label='YEARLY PNL'):
                st.dataframe(pnl)
            with st.expander(label='QUARTERLY PNL'):
                st.dataframe(qtr_pnl)

            keydata_col1, keydata_col2 = st.columns([1,1])
            with keydata_col1:
                fundamentals.bar_line(balancesht, "DEBTOR DAYS", "INVENTORY TURNOVER", color_dict[color_key]['hash'],comp_Name, "Yearly")
            with keydata_col2:
                fundamentals.bar_line(balancesht, "NET BLOCK", "CAPITAL WORK IN PROGRESS", color_dict[color_key]['hash'], comp_Name,"Yearly")
            with keydata_col1:
                fundamentals.bar_line(balancesht, "RESERVES", "BORROWINGS", color_dict[color_key]['hash'], comp_Name,"Yearly")
            with keydata_col2:
                fundamentals.bar_line(balancesht, "WORKING CAPITAL", "CASH & BANK", color_dict[color_key]['hash'],comp_Name,"Yearly")

            with keydata_col1:
                st.title("QUARTERLY")
                fundamentals.group_2_bars(qtr_pnl, "SALES", "OTHER INCOME", comp_Name, "Quarterly")

            with keydata_col2:
                st.title("YEARLY")
                fundamentals.group_2_bars(pnl, "SALES", "OTHER INCOME", comp_Name, "Yearly")

            with keydata_col1:
                fundamentals.both_lines(qtr_pnl, "OPM %", "NPM %", color_dict['red1']['hash'], color_dict['green1']['hash'],
                                        comp_Name, "Quarterly")
                fundamentals.bar_line(qtr_pnl, 'NET PROFIT', 'NPM %', color_dict[color_key]['hash'], comp_Name, "Quarterly")

            with keydata_col2:
                fundamentals.both_lines(pnl, "OPM %", "NPM %", color_dict['red1']['hash'], color_dict['green1']['hash'],
                                        comp_Name, "Yearly")
                fundamentals.bar_line(pnl, 'NET PROFIT', 'NPM %', color_dict[color_key]['hash'], comp_Name, "Yearly")

            fundamentals.go_group_bar(df_comp.loc['CASH FLOW'], "cash_flows", color_dict[color_key]['hash'], "Yearly")

        # for each in variables.metadata.keys():
        #     st.info(each)

        if sub_choose == "PROFIT&LOSS":            # YEARLY PNL
            sentence += fundamentals.stmt_for_qoq(pnl)
            coltw1, coltw2 = st.columns([4, 1])
            with coltw1:
                st.text_area(label="👉 INSIGHTS", value=sentence, height=100)
            with coltw2:
                if st.button('Send Telegram'):
                    bot.send_message(chat_id=chat_id, text=sentence)

            Ykeydata,YSales, YOtherIncome,YExpenses,YOperatingProfit,YNetProfit,Ytable = st.tabs(['Key Data','SALES','OTHER INCOME','EXPENSES','OPERATING PROFIT','NET PROFIT','Table'])

            with Ytable:
                # THE FOLLOWIGN CODE CALCULATES THE GROWTH OR DEGROWTH
                st.dataframe(pnl.style.format(formatter="{:.1f}"))
            with Ykeydata:
                fundamentals.group_2_bars(pnl,"SALES","OTHER INCOME",comp_Name, "Yearly")
                fundamentals.group_2_bars(pnl,"PROFIT BEFORE TAX","NET PROFIT",comp_Name, "Yearly")
                fundamentals.group_3_bars(pnl, "SALES", "PROFIT BEFORE TAX","NET PROFIT",comp_Name,"Yearly")
                fundamentals.both_lines(pnl, "OPM %", "NPM %", color_dict['red1']['hash'], color_dict['green1']['hash'], comp_Name,"Yearly")
            with YSales:
                fundamentals.qoq_growth(pnl, 'SALES', color_dict[color_key]['hash'], comp_Name,"Yearly")
            with YOtherIncome:
                fundamentals.go_bar(pnl, 'OTHER INCOME', color_dict[color_key]['hash'], comp_Name,"Yearly")
            with YExpenses:
                fundamentals.go_bar(pnl, 'EXPENSES', color_dict[color_key]['hash'], comp_Name,"Yearly")
            with YOperatingProfit:
                fundamentals.bar_line(pnl, 'OPERATING PROFIT','OPM %', color_dict[color_key]['hash'], comp_Name,"Yearly")
            with YNetProfit:
                fundamentals.bar_line(pnl,'NET PROFIT','NPM %',color_dict[color_key]['hash'],comp_Name,"Yearly")

        if sub_choose == "QTR PnL":                #QUARTERLY PNL
            sentence += fundamentals.stmt_for_qoq(qtr_pnl)
            coltw1, coltw2 = st.columns([4, 1])
            with coltw1:
                st.text_area(label="TWITTER POST", value=sentence, height=75)
            with coltw2:
                if st.button('Send Telegram'):
                    bot.send_message(chat_id=chat_id, text=sentence)

            Qkeydata, QSales, QOtherIncome, QExpenses, QOperatingProfit, QNetProfit, Qtable = st.tabs(
                ['Key Data', 'SALES', 'OTHER INCOME', 'EXPENSES', 'OPERATING PROFIT', 'NET PROFIT', 'Table'])
            with Qtable:
                st.dataframe(qtr_pnl)
                # Replace the first row with NaN for the QoQ columns
                #df.loc[0, ['SALES_QoQ', 'NET PROFIT_QoQ', 'OPERATING PROFIT_QoQ']] = np.nan
                #df = df.transpose()
                #st.dataframe(qtr_pnl.style.format(formatter="{:.1f}"))

            with Qkeydata:
                fundamentals.group_2_bars(qtr_pnl, "SALES", "OTHER INCOME", comp_Name, "Quarterly")
                fundamentals.group_2_bars(qtr_pnl, "PROFIT BEFORE TAX", "NET PROFIT", comp_Name, "Quarterly")
                #fundamentals.group_3_bars(qtr_pnl, "SALES",  "PROFIT BEFORE TAX","NET PROFIT",comp_Name,"Quarterly")
                fundamentals.both_lines(qtr_pnl, "OPM %", "NPM %", color_dict['red1']['hash'], color_dict['green1']['hash'],comp_Name,"Quarterly")
            with QSales:
                fundamentals.qoq_growth(qtr_pnl,"SALES", color_dict[color_key]['hash'],comp_Name,"Quarterly")
            with QOtherIncome:
                fundamentals.go_bar(qtr_pnl, 'OTHER INCOME', color_dict[color_key]['hash'], comp_Name,"Quarterly")
            with QExpenses:
                fundamentals.go_bar(qtr_pnl, 'EXPENSES', color_dict[color_key]['hash'], comp_Name,"Quarterly")
            with QOperatingProfit:
                fundamentals.bar_line(qtr_pnl, 'OPERATING PROFIT','OPM %', color_dict[color_key]['hash'], comp_Name,"Quarterly")
            with QNetProfit:
                fundamentals.bar_line(qtr_pnl,'NET PROFIT','NPM %',color_dict[color_key]['hash'],comp_Name,"Quarterly")


        if sub_choose == "BALANCE SHEET":        #YEARLY BALANCE SHEET
            BSKeyData, BStable, BSReserves, BSBorrowings, BSOtherAssets, BSOtherLiabilities, BSReceivables, BSInventory, BSCWIP = st.tabs(['KeyData','table','Reserves','Borrowings','OtherAssets','OtherLiabilities','Receivables','Inventory','CWIP'])
            with BSKeyData:
                fundamentals.bar_line(balancesht,"RESERVES","BORROWINGS",color_dict[color_key]['hash'],comp_Name,"Yearly")
                fundamentals.bar_line(balancesht,"RECEIVABLES","INVENTORY",color_dict[color_key]['hash'],comp_Name,"Yearly")
            with BStable:
                st.dataframe(balancesht.style.format(formatter="{:.1f}"))
            with BSCWIP:
                fundamentals.go_bar(balancesht, "CAPITAL WORK IN PROGRESS", color_dict[color_key]['hash'],comp_Name,"Yearly")
            with BSInventory:
                fundamentals.go_bar(balancesht, "INVENTORY", color_dict[color_key]['hash'],comp_Name,"Yearly")
            with BSReserves:
                fundamentals.qoq_growth(balancesht, "RESERVES", color_dict[color_key]['hash'],comp_Name,"Yearly")
            with BSBorrowings:
                fundamentals.qoq_growth(balancesht, "BORROWINGS", color_dict[color_key]['hash'],comp_Name,"Yearly")
            with BSReceivables:
                fundamentals.qoq_growth(balancesht, "RECEIVABLES", color_dict[color_key]['hash'],comp_Name,"Yearly")
            with BSOtherAssets:
                fundamentals.qoq_growth(balancesht, "OTHER ASSETS", color_dict[color_key]['hash'],comp_Name,"Yearly")
            with BSOtherLiabilities:
                fundamentals.qoq_growth(balancesht, "OTHER LIABILITIES", color_dict[color_key]['hash'],comp_Name,"Yearly")


        if sub_choose == "CASH FLOW":        # YEARLY CASH FLOWS
            CF, CFTab, CFop, CFinv, CFfin, NetCF = st.tabs(["KeyData","Table","Operating Cash","Investing Cash","Financing Cash","Net Cash Flow"])
            with CFTab:
                st.dataframe(df_comp.loc[sub_choose].style.format(formatter="{:.1f}"))
            with CF:
                fundamentals.go_group_bar(df_comp.loc[sub_choose], "cash_flows", color_dict[color_key]['hash'],"Yearly")
            with CFop:
                fundamentals.qoq_growth(df_comp.loc[sub_choose], "CASH FROM OPERATING ACTIVITY", color_dict[color_key]['hash'],comp_Name,"Yearly")
            with CFfin:
                fundamentals.go_bar(df_comp.loc[sub_choose], "CASH FROM FINANCING ACTIVITY", color_dict[color_key]['hash'], comp_Name,"Yearly")
            with CFinv:
                fundamentals.go_bar(df_comp.loc[sub_choose], "CASH FROM INVESTING ACTIVITY", color_dict[color_key]['hash'],comp_Name,"Yearly")
            with NetCF:
                fundamentals.go_bar(df_comp.loc[sub_choose], "NET CASH FLOW", color_dict[color_key]['hash'], comp_Name,"Yearly")

    if len(uploaded_file)==2:
        comp1_Name = uploaded_file[0].name.split('.')[0]
        comp2_Name = uploaded_file[1].name.split('.')[0]
        col1, col2, col3 = st.columns([0.5, 0.3, 0.2])
        with col1:
            st.subheader('📈 ' + comp1_Name + '📈 ' + comp2_Name )
        with col2:
            qoq_checked = st.checkbox("Sequential_Growth_%")
        with col3:
            color_bar = st.color_picker("Bar Color",
                                        value="#ECE80F")  # blueshades"#0971C9""ECE80F"   #yellowshades"#f8ba43"
        st.write("____")
        book1 = openpyxl.load_workbook(uploaded_file[0])
        qtr1_pnl, df1 = fundamentals.get_tables(book1[fundamentals.tabs[-1]], uploaded_file[0])  # send a sheet(not whole workbook)
        book2 = openpyxl.load_workbook(uploaded_file[1])
        qtr2_pnl, df2 = fundamentals.get_tables(book2[fundamentals.tabs[-1]], uploaded_file[1])  # send a sheet(not whole workbook)
        book1 = openpyxl.load_workbook(uploaded_file[0])

        qtr1_pnl, df1 = fundamentals.get_tables(book1[fundamentals.tabs[-1]], uploaded_file[0])  # send a sheet(not whole workbook)
        book2 = openpyxl.load_workbook(uploaded_file[1])
        qtr2_pnl, df2 = fundamentals.get_tables(book2[fundamentals.tabs[-1]], uploaded_file[1])  # send a sheet(not whole workbook)
        main_menu = st.sidebar.selectbox("Chose", fundamentals.funda_keys)
        sub_menu = st.sidebar.selectbox("SubChose", list(df1.loc[main_menu].index))
        new_df = [df1.loc[main_menu].loc[sub_menu], df2.loc[main_menu].loc[sub_menu]]
        df_new = pd.concat(new_df, keys=[comp1_Name, comp2_Name], axis=1)
        fundamentals.peer_bar(df_new, sub_menu,comp1_Name,comp2_Name)
    if len(uploaded_file) >= 3:
        st.error("Please, upload only 2 excel files")




st.write("____")
st.write('made with :green_heart: to my Indian Stock Investors')


