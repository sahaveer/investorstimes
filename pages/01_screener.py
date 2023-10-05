import streamlit as st
from streamlit_option_menu import option_menu
import streamlit.components.v1 as html
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import base64  # Standard Python Module
from io import StringIO, BytesIO  # Standard Python Module
import os
import fundamentals
import json
from streamlit_lottie import st_lottie

st.set_page_config(
        page_title="Data Visualisations from SCREENER",
        #page_icon=":hammer_and_wrench:",
        layout="wide"
    )

color_dict = {'Yellow_Lite':"#f8ba43",'Yellow_Dark':"#D6D41B",'Blue_Lite':"#1959BF",'Blue_Dark':"#0971C9",'Green_Lite':"#11A694",'Green_Dark':"#11A64B","Purple_Lite":"#7019BF",'Purple_Dark':"#9319BF"}
#color_list = ["#D6D41B","#f8ba43","#0971C9","#1959BF","#11A694","#11A64B","#7019BF","#9319BF"]
def load_lottiefile(filepath: str):
    with open(filepath, "r") as f:
        return json.load(f)
def load_lottieurl(url: str):
    r = requests.get(url)
    if r.status_code != 200:
        return None
    return r.json()
lottie_data_analytics = load_lottiefile("./lottie/data-analytics.json")

with st.sidebar:
    st.markdown(""" <style> .font {
    font-size:22px ; font-family: 'Cooper Black'; color: #FF9633;} 
    </style> """, unsafe_allow_html=True)
    # Add a file uploader to allow users to upload their csv file
    
    uploaded_file = st.file_uploader("", type=['xlsx','xlsm'],accept_multiple_files = True)  # Only accepts xlsx,xlsm file format
    st.markdown('<p class="font">Upload One or Two xlsx/xlsm FILES from screener.in </p>',
                unsafe_allow_html=True)
    st_lottie(
        lottie_data_analytics,
        speed=0.7,
        reverse=False,
        loop=True,
        quality="low",  # medium ; high
        height=None,
        width=None,
        key="barchart", )
if uploaded_file is not None:
    if len(uploaded_file)==1:
        comp_Name = uploaded_file[0].name.split('.xlsx')[0]
        comp_Name = comp_Name.split('.xlsm')[0]

        col1, col2, col3 = st.columns([0.5, 0.3, 0.2])
        with col1:
            st.subheader('📈 ' + comp_Name)
        with col3:
            color_key = st.selectbox("Bar Color",color_dict.keys())
            #color_bar = st.color_picker("Bar Color", value="#ECE80F")  # blueshades"#0971C9""ECE80F"   #yellowshades"#f8ba43"
        st.write("____")
        book = openpyxl.load_workbook(uploaded_file[0])
        qtr_pnl,df_comp = fundamentals.get_tables(book[fundamentals.tabs[-1]], uploaded_file[0])  # send a sheet(not whole workbook)
        try:
            df_comp.columns = df_comp.columns.strftime('%d-%m-%Y')
        except Exception as AttributeError:
            pass
        try:
            qtr_pnl.columns = qtr_pnl.columns.strftime('%d-%m-%Y')
        except Exception as AttributeError:
            pass
        # **************************************************************************************************
        if os.path.isdir('./pickl'):
            df_comp.to_pickle("./pickl/"+comp_Name + ".pkl")
        sub_choose = st.sidebar.selectbox("Fundamentals", fundamentals.funda_menu)
        if sub_choose == fundamentals.funda_menu[0]:
            index_list = ["key_params"] + list(df_comp.loc[sub_choose].index)
            param = st.sidebar.selectbox("SubChose", index_list)
            if param == "key_params":
                with st.expander("YEARLY PROFIT & LOSS DATA"):
                    st.dataframe(df_comp.loc[sub_choose].style.format(formatter="{:.1f}"))
                fundamentals.qoq_growth(df_comp.loc[sub_choose], "SALES", color_dict[color_key],comp_Name)
                fundamentals.group_2_bars(df_comp.loc[sub_choose],"PROFIT BEFORE TAX","NET PROFIT",comp_Name)
            else:
                with col2:
                    qoq_checked = st.checkbox("Sequential_Growth_%")
                if qoq_checked:
                    fundamentals.qoq_growth(df_comp.loc[sub_choose], param, color_dict[color_key],comp_Name)
                else:
                    fundamentals.go_bar(df_comp.loc[sub_choose], param, color_dict[color_key],comp_Name)
        if sub_choose == fundamentals.funda_menu[3]:
            index_list = ["key_params"] + list(qtr_pnl.index)
            param = st.sidebar.selectbox("SubChose", index_list)
            if param == "key_params":
                with st.expander("QUARTERLY PROFIT & LOSS DATA"):
                    st.dataframe(qtr_pnl.style.format(formatter="{:.1f}"))
                fundamentals.qoq_growth(qtr_pnl, "SALES", color_dict[color_key],comp_Name)
                fundamentals.group_2_bars(qtr_pnl, "PROFIT BEFORE TAX","NET PROFIT",comp_Name)
                fundamentals.qoq_growth(qtr_pnl, "PROFIT BEFORE TAX", color_dict[color_key],comp_Name)
                fundamentals.qoq_growth(qtr_pnl, "NET PROFIT", color_dict[color_key],comp_Name)
            else:
                with col2:
                    qoq_checked = st.checkbox("Sequential_Growth_%")
                if qoq_checked:
                    fundamentals.qoq_growth(qtr_pnl, param, color_dict[color_key],comp_Name)
                else:
                    fundamentals.go_bar(qtr_pnl, param, color_dict[color_key],comp_Name)
    
        if sub_choose == fundamentals.funda_menu[1]:
            index_list = ["key_params"] + list(df_comp.loc[sub_choose].index)
            param = st.sidebar.selectbox("SubChose", index_list)
            if param == "key_params":
                with st.expander("YEARLY BALANCE SHEET DATA"):
                    st.dataframe(df_comp.loc[sub_choose].style.format(formatter="{:.1f}"))
                fundamentals.bar_line(df_comp.loc[sub_choose],"RESERVES","BORROWINGS",color_dict[color_key],comp_Name)
                fundamentals.bar_line(df_comp.loc[sub_choose],"RECEIVABLES","INVENTORY",color_dict[color_key],comp_Name)
                fundamentals.go_bar(df_comp.loc[sub_choose], "CAPITAL WORK IN PROGRESS", color_dict[color_key],comp_Name)
                fundamentals.go_bar(df_comp.loc[sub_choose], "CASH & BANK", color_dict[color_key],comp_Name)
            else:
                with col2:
                    qoq_checked = st.checkbox("Sequential_Growth_%")
                if qoq_checked:
                    fundamentals.qoq_growth(df_comp.loc[sub_choose], param, color_dict[color_key],comp_Name)
                else:
                    fundamentals.go_bar(df_comp.loc[sub_choose], param, color_dict[color_key],comp_Name)
        if sub_choose == fundamentals.funda_menu[2]:
            index_list = ["key_params"] + list(df_comp.loc[sub_choose].index)
            param = st.sidebar.selectbox("SubChose", index_list)
            if param == "key_params":
                with st.expander("YEARLY CASH FLOWS DATA"):
                    st.dataframe(df_comp.loc[sub_choose].style.format(formatter="{:.1f}"))
                fundamentals.go_group_bar(df_comp.loc[sub_choose], "cash_flows", color_dict[color_key])
            else:
                with col2:
                    qoq_checked = st.checkbox("Sequential_Growth_%")
                if qoq_checked:
                    fundamentals.qoq_growth(df_comp.loc[sub_choose], param, color_dict[color_key],comp_Name)
                else:
                    fundamentals.go_bar(df_comp.loc[sub_choose], param, color_dict[color_key], comp_Name)
    if len(uploaded_file)==2:
        comp1_Name = uploaded_file[0].name.split('.')[0]
        comp2_Name = uploaded_file[1].name.split('.')[0]
        col1, col2, col3 = st.columns([0.5, 0.3, 0.2])
        with col1:
            st.subheader('📈 ' + comp1_Name + '📈 ' + comp2_Name )
        with col2:
            qoq_checked = st.checkbox("Sequential_Growth_%")
        with col3:
            color_bar = st.color_picker("Bar Color",
                                        value="#ECE80F")  # blueshades"#0971C9""ECE80F"   #yellowshades"#f8ba43"
        st.write("____")
        book1 = openpyxl.load_workbook(uploaded_file[0])
        qtr1_pnl, df1 = fundamentals.get_tables(book1[fundamentals.tabs[-1]], uploaded_file[0])  # send a sheet(not whole workbook)
        book2 = openpyxl.load_workbook(uploaded_file[1])
        qtr2_pnl, df2 = fundamentals.get_tables(book2[fundamentals.tabs[-1]], uploaded_file[1])  # send a sheet(not whole workbook)
        book1 = openpyxl.load_workbook(uploaded_file[0])

        qtr1_pnl, df1 = fundamentals.get_tables(book1[fundamentals.tabs[-1]], uploaded_file[0])  # send a sheet(not whole workbook)
        book2 = openpyxl.load_workbook(uploaded_file[1])
        qtr2_pnl, df2 = fundamentals.get_tables(book2[fundamentals.tabs[-1]], uploaded_file[1])  # send a sheet(not whole workbook)
        main_menu = st.sidebar.selectbox("Chose", fundamentals.funda_keys)
        sub_menu = st.sidebar.selectbox("SubChose", list(df1.loc[main_menu].index))
        new_df = [df1.loc[main_menu].loc[sub_menu], df2.loc[main_menu].loc[sub_menu]]
        df_new = pd.concat(new_df, keys=[comp1_Name, comp2_Name], axis=1)
        fundamentals.peer_bar(df_new, sub_menu,comp1_Name,comp2_Name)
    if len(uploaded_file) >= 3:
        st.error("Please, upload only 2 excel files")

st.write("____")
st.write('made with :green_heart: to my Indian Stock Investors')
