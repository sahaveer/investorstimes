import streamlit as st
from streamlit_option_menu import option_menu
import streamlit.components.v1 as html
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import base64  # Standard Python Module
from io import StringIO, BytesIO  # Standard Python Module
import os

st.set_page_config(page_title="InteractiveCharts",page_icon=":bar_chart:",layout="wide")
st.title('Excel Plotter 📈')
st.subheader('Upload your Excel file and have fun')

tabs = ['Profit & Loss', 'Quarters','Balance Sheet', 'Cash Flow' ,'Data Sheet']
table = ['PROFIT & LOSS', 'Quarters', 'BALANCE SHEET' ]
# lets define the first and last table keys to extract the exact table size
DataSheet_Key_Values = ['PROFIT & LOSS', 'Dividend Amount', 'Quarters', 'Operating Profit', 'BALANCE SHEET', 'Cash & Bank',
                        'CASH FLOW:', 'Net Cash Flow']
funda_keys = ['PROFIT&LOSS','BALANCE SHEET','CASH FLOW']    # dont change the order of this list as it will affect the keys used in Yearly df
funda_menu = funda_keys + ['QTR PnL']
color_hover = "darkgrey"
color_background = "grey"

def generate_excel_download_link(df):
    # Credit Excel: https://discuss.streamlit.io/t/how-to-add-a-download-excel-csv-function-to-a-button/4474/5
    towrite = BytesIO()
    df.to_excel(towrite, encoding="utf-8", index=False, header=True)  # write to BytesIO buffer
    towrite.seek(0)  # reset pointer
    b64 = base64.b64encode(towrite.read()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="data_download.xlsx">Download Excel File</a>'
    return st.markdown(href, unsafe_allow_html=True)

def generate_html_download_link(fig):
    # Credit Plotly: https://discuss.streamlit.io/t/download-plotly-plot-as-html/4426/2
    towrite = StringIO()
    fig.write_html(towrite, include_plotlyjs="cdn")
    towrite = BytesIO(towrite.getvalue().encode())
    b64 = base64.b64encode(towrite.read()).decode()
    href = f'<a href="data:text/html;charset=utf-8;base64, {b64}" download="plot.html">Download Plot</a>'
    return st.markdown(href, unsafe_allow_html=True)

#def sns_bar(df1, row_name,color_bar,color_line,comp):
    #df = df1.transpose()
    #fig = plt.figure(figsize=(10,4))
    #sns.barplot(y= df1.loc['Sales'],x=df1.columns, data = df,color="salmon", saturation=.5) #palette="Blues_d")
    #st.pyplot(fig)

def go_bar(df, row_name,color_bar,color_line,comp):
    #fig = ff.create_table(df)
    #st.plotly_chart(fig)
    #st.write(df)                                                   # this is givng data conversion error sometimes
    fig = go.Figure(data = [go.Bar(x = df.columns,y = df.loc[row_name],
                                   text = df.loc[row_name], textposition = 'auto')])
    # go.bar has another attribute - hovertext = ['27% market share', '24% market share', '19% market share']
    fig.update_traces(marker_color=color_bar, marker_line_color='rgb(8,48,107)',
                      marker_line_width=1.5, opacity=1, texttemplate='%{text:.2s}', textposition='outside')
    fig.update_layout(height=600,width=900,
                      title={'text': "<b>" + comp_Name.upper() + "</b> : <i>" + row_name + ' Report</i>',
                             'y': 0.9,'x': 0.5,'xanchor': 'center','yanchor': 'bottom'},
                      xaxis_tickfont_size=14, xaxis_tickangle=-45,
                      xaxis=dict(showgrid=False),
                      yaxis=dict(showgrid=False, title='INR (cr)',titlefont_size=16,tickfont_size=1),
                      legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                      bargap=0.15)     #legend=dict(x=0,y=1.0,bgcolor='rgba(255, 255, 255, 0)',bordercolor='rgba(255, 255, 255, 0)'),
    st.markdown("---")
    st.plotly_chart(fig,use_container_width=True)
    #st.subheader('Downloads:')
    #generate_html_download_link(fig)

def go_bar_line(df,row_name,color_bar,color_line):
    temp_df = df.loc[row_name]
    df_qoq = (temp_df.pct_change() * 100)
    df_qoq.name = row_name + '_QoQ'
    df2 = df.append(df_qoq)
    # Create figure with secondary y-axis
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace( go.Bar(x=df2.columns, y=df2.loc[row_name], name=row_name, textposition = 'auto',text = df.loc[row_name]),
                   secondary_y=False)
    fig.update_traces(marker_color=color_bar, marker_line_color='rgb(8,48,107)',
                      marker_line_width=1.5, opacity=1, texttemplate='%{text:.2s}', textposition='outside')
    fig.add_trace( go.Scatter(x=df2.columns, y=df2.iloc[-1], name=row_name + " QoQ", text = df.iloc[-1]),
                   secondary_y=True)
    # Add figure title
    fig.update_layout( height=600, width=900,
                       title={'text': "<b>" + comp_Name.upper() + "</b> : <i>" + row_name + ' Report</i>',
                              'y': 0.9, 'x': 0.5, 'xanchor': 'center', 'yanchor': 'top'},
                       xaxis_tickfont_size=14, xaxis_tickangle=-45,
                       yaxis=dict(title='INR (cr)', titlefont_size=16, tickfont_size=1,),
                       legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                       bargap=0.15)#legend=dict(yanchor="top",y=0.99,xanchor="left",x=0.01),bargap = 0.15)
    # Set y-axes titles
    fig.update_yaxes(title_text="<b>" + row_name + "</b> in cr ", secondary_y=False,showgrid=False)
    fig.update_yaxes(title_text="<b> QoQ </b> in % ", secondary_y=True,showgrid=False)
    st.markdown("---")
    st.plotly_chart(fig,use_container_width=True)
    #st.subheader('Downloads:')
    #generate_excel_download_link(df2)
    #generate_html_download_link(fig)

def group_bar(df,sub_menu):   #this has 2 series concatinated with key names
    bar_list = list(df.columns)
    fig = go.Figure(data=[go.Bar(name=bar_list[0], x=df.index, y=df[bar_list[0]]),
                          go.Bar(name=bar_list[1], x=df.index, y=df[bar_list[1]])])
    # Change the bar mode
    fig.update_layout(barmode='group', bargroupgap=0.1,
                      height=600, width=900,
                      title={'text': sub_menu + ' Report',
                             'y': 0.9, 'x': 0.5, 'xanchor': 'center', 'yanchor': 'top'},
                      xaxis_tickfont_size=14, xaxis_tickangle=-45,
                      yaxis=dict(title='INR (cr)', titlefont_size=16, tickfont_size=14, ),
                      legend=dict(yanchor="top",y=0.99,xanchor="left",x=0.01),)   # this barmode = 'group | stack | 'relative''
    st.plotly_chart(fig)

def go_group_bar(df, row_name,color_bar,color_line):
    bar_list = list(df.index)
    fig = go.Figure(data=[ go.Bar(name=bar_list[0], x=df.columns, y=df.iloc[0], textposition = 'auto', text = df.iloc[0]),
                           go.Bar(name=bar_list[1],  x=df.columns, y=df.iloc[1], textposition = 'auto', text = df.iloc[1]),
                           go.Bar(name=bar_list[2], x=df.columns, y=df.iloc[2], textposition='auto', text=df.iloc[2]),
                           go.Bar(name=bar_list[3], x=df.columns, y=df.iloc[3], textposition='auto', text=df.iloc[3])
                           ])
    # Change the bar mode
    fig.update_layout(barmode='group', bargroupgap=0.1,
                      height=600, width=900,
                      title={'text': row_name + ' Report',
                             'y': 0.9, 'x': 0.5, 'xanchor': 'center', 'yanchor': 'top'},
                      xaxis_tickfont_size=14, xaxis_tickangle=-45,
                      yaxis=dict(title='INR (cr)', titlefont_size=16, tickfont_size=14, ),
                      legend=dict(yanchor="top",y=0.99,xanchor="left",x=0.01),)   # this barmode = 'group | stack | 'relative''
    #fig.update_layout(barmode='stack', xaxis={'categoryorder':'category ascending'}) # WHILE USING STACK
    st.plotly_chart(fig)

def get_tables(datasht,file):
    for i in range(1,datasht.max_row+1) :
        if datasht['A'+str(i)].value == DataSheet_Key_Values[0] :
            pnl_start_row = i
        if datasht['A' + str(i)].value == DataSheet_Key_Values[1]:
            pnl_end_row = i-1
        if datasht['A' + str(i)].value == DataSheet_Key_Values[2]:
            quarterly_start_row = i
        if datasht['A' + str(i)].value == DataSheet_Key_Values[3]:
            quarterly_end_row = i-1
        if datasht['A' + str(i)].value == DataSheet_Key_Values[4]:
            BS_start_row = i
        if datasht['A' + str(i)].value == DataSheet_Key_Values[5]:
            BS_end_row = i-1
        if datasht['A' + str(i)].value == DataSheet_Key_Values[6]:
            cash_start_row = i
        if datasht['A' + str(i)].value == DataSheet_Key_Values[7]:
            cash_end_row = i-1
    reqd_cols = "A :" + str(get_column_letter(datasht.max_column))

    if pnl_start_row is not None and pnl_end_row is not None :
        pnl = pd.read_excel(file, index_col=0, sheet_name=tabs[-1], header=pnl_start_row,usecols=reqd_cols,
                            nrows=pnl_end_row-pnl_start_row )
        pnl.columns = pnl.columns.strftime('%d-%m-%Y')
        pnl.index = pnl.index.str.upper()
    if quarterly_start_row is not None and quarterly_end_row is not None :
        qtr_pnl = pd.read_excel(file, index_col=0, sheet_name=tabs[-1], header=quarterly_start_row, usecols=reqd_cols,
                                nrows=quarterly_end_row - quarterly_start_row)
        qtr_pnl.columns = qtr_pnl.columns.strftime('%d-%m-%Y')
        qtr_pnl.index = qtr_pnl.index.str.upper()

    if BS_start_row is not None and BS_end_row is not None:
        balancesht = pd.read_excel(file, index_col=0, sheet_name=tabs[-1], header=BS_start_row,usecols=reqd_cols,
                                   nrows=BS_end_row-BS_start_row )
        balancesht.columns = balancesht.columns.strftime('%d-%m-%Y')
        balancesht.index = balancesht.index.str.upper()
    if cash_start_row is not None and cash_end_row is not None:
        cashflow = pd.read_excel(file, index_col=0, sheet_name=tabs[-1], header=cash_start_row,usecols=reqd_cols,
                                   nrows=cash_end_row-cash_start_row)
        cashflow.columns = cashflow.columns.strftime('%d-%m-%Y')
        cashflow.index = cashflow.index.str.upper()
    if(pnl is not None and balancesht is not None and cashflow is not None):
        sht_list = [pnl,balancesht,cashflow]
        df_comp = pd.concat(sht_list,keys=funda_keys)
    return qtr_pnl,df_comp

with st.sidebar:
    st.markdown(""" <style> .font {
    font-size:22px ; font-family: 'Cooper Black'; color: #FF9633;} 
    </style> """, unsafe_allow_html=True)
    # Add a file uploader to allow users to upload their csv file
    st.markdown('<p class="font">Upload your xlsx/xlsm from screener.in </p>',
                unsafe_allow_html=True)
    uploaded_file = st.file_uploader("", type=['xlsx', 'xlsm'],accept_multiple_files = True)  # Only accepts xlsx,xlsm file format

if uploaded_file is not None:
    if len(uploaded_file)==1:
        comp_Name = uploaded_file[0].name.split('.')[0]
        book = openpyxl.load_workbook(uploaded_file[0])
        qtr_pnl,df_comp = get_tables(book[tabs[-1]], uploaded_file[0])  # send a sheet(not whole workbook)
        col1, col2 = st.columns([0.8, 0.2])
        with col1:
            sub_choose = option_menu("Fundamentals", funda_menu,
                                     styles={"container": {"padding": "5!important"},
                                             "icon": {"color": "yellow", "font-size": "18px"},
                                             "nav-link": {"font-size": "16px", "text-align": "left", "margin": "0px",
                                                          "--hover-color": color_hover},
                                             "nav-link-selected": {"background-color": color_background}},
                                     menu_icon="cast", default_index=0, orientation="horizontal")
            with col2:
                color_bar = st.color_picker("Bar", value="#0f7eec")
                color_line = st.color_picker("Line", value="#D60A10")
        if sub_choose == funda_menu[0]:
            index_list = ["key_params"] + list(df_comp.loc[funda_menu[0]].index)
            with st.sidebar:
                param = option_menu(sub_choose, index_list,
                                    styles={"container": {"padding": "5!important"},
                                            "icon": {"color": "orange", "font-size": "25px"},
                                            "nav-link": {"font-size": "16px", "text-align": "left", "margin": "0px",
                                                         "--hover-color": color_hover},
                                            "nav-link-selected": {"background-color": color_background}},
                                    menu_icon="cast", default_index=0, orientation="vertical")
            if param == "key_params":
                st.dataframe(df_comp.loc[funda_menu[0]].style.format(formatter="{:.1f}"))
                #sns_bar(df_comp.loc[funda_menu[0]], "Sales", color_bar, color_line,comp_Name)
                go_bar_line(df_comp.loc[funda_menu[0]], "SALES", color_bar, color_line)
                go_bar_line(df_comp.loc[funda_menu[0]], "PROFIT BEFORE TAX", color_bar, color_line)
                go_bar_line(df_comp.loc[funda_menu[0]], "NET PROFIT", color_bar, color_line)
            else:
                if st.checkbox("Sequential_Growth_%"):
                    go_bar_line(df_comp.loc[funda_menu[0]], param, color_bar, color_line)
                else:
                    go_bar(df_comp.loc[funda_menu[0]], param, color_bar, color_line,comp_Name)

        if sub_choose == funda_menu[3]:
            # param = st.selectbox("Select  column", df_comp.loc[funda_menu[3]].index)
            index_list = ["key_params"] + list(qtr_pnl.index)
            with st.sidebar:
                param = option_menu(sub_choose, index_list,
                                    styles={"container": {"padding": "5!important"},
                                            "icon": {"color": "orange", "font-size": "25px"},
                                            "nav-link": {"font-size": "16px", "text-align": "left", "margin": "0px",
                                                         "--hover-color": color_hover},
                                            "nav-link-selected": {"background-color": color_background}},
                                    menu_icon="cast", default_index=0, orientation="vertical")
            if param == "key_params":
                st.dataframe(qtr_pnl.style.format(formatter="{:.1f}"))
                go_bar_line(qtr_pnl, "SALES", color_bar, color_line)
                go_bar_line(qtr_pnl, "PROFIT BEFORE TAX", color_bar, color_line)
                go_bar_line(qtr_pnl, "NET PROFIT", color_bar, color_line)
            else:
                if st.checkbox("Sequential_Growth_%"):
                    go_bar_line(qtr_pnl, param, color_bar, color_line)
                else:
                    go_bar(qtr_pnl, param, color_bar, color_line,comp_Name)
    
        if sub_choose == funda_menu[1]:
            #st.info(funda_menu[1])
            # param = st.selectbox("Select  column", df_comp.loc[funda_menu[2]].index)
            index_list = ["key_params"] + list(df_comp.loc[funda_menu[1]].index)
            with st.sidebar:
                param = option_menu(sub_choose, index_list,
                                    styles={"container": {"padding": "5!important"},
                                            "icon": {"color": "orange", "font-size": "25px"},
                                            "nav-link": {"font-size": "16px", "text-align": "left", "margin": "0px",
                                                         "--hover-color": color_hover},
                                            "nav-link-selected": {"background-color": color_background}},
                                    menu_icon="cast", default_index=0, orientation="vertical")
            if param == "key_params":
                st.dataframe(df_comp.loc[funda_menu[1]].style.format(formatter="{:.1f}"))
                go_bar_line(df_comp.loc[funda_menu[1]], "RESERVES", color_bar, color_line)
                go_bar_line(df_comp.loc[funda_menu[1]], "BORROWINGS", color_bar, color_line)
                go_bar_line(df_comp.loc[funda_menu[1]], "CAPITAL WORK IN PROGRESS", color_bar, color_line)
                go_bar_line(df_comp.loc[funda_menu[1]], "CASH & BANK", color_bar, color_line)
            else:
                if st.checkbox("Sequential_Growth_%"):
                    go_bar_line(df_comp.loc[funda_menu[1]], param, color_bar, color_line)
                else:
                    go_bar(df_comp.loc[funda_menu[1]], param, color_bar, color_line,comp_Name)
        if sub_choose == funda_menu[2]:
            index_list = ["key_params"] + list(df_comp.loc[funda_menu[2]].index)
            # param = st.selectbox("Select  column", df_comp.loc[funda_menu[3]].index)
            with st.sidebar:
                param = option_menu(sub_choose, index_list,
                                    styles={"container": {"padding": "5!important"},
                                            "icon": {"color": "orange", "font-size": "25px"},
                                            "nav-link": {"font-size": "16px", "text-align": "left", "margin": "0px",
                                                         "--hover-color": color_hover},
                                            "nav-link-selected": {"background-color": color_background}},
                                    menu_icon="cast", default_index=0, orientation="vertical")
            if param == "key_params":
                st.dataframe(df_comp.loc[funda_menu[2]].style.format(formatter="{:.1f}"))
                go_group_bar(df_comp.loc[funda_menu[2]], "cash_flows", color_bar, color_line)
            else:
                if st.checkbox("Sequential_Growth_%"):
                    go_bar_line(df_comp.loc[funda_menu[2]], param, color_bar, color_line)
                else:
                    go_bar(df_comp.loc[funda_menu[2]], param, color_bar, color_line,comp_Name)
    if len(uploaded_file)==2:
        book1 = openpyxl.load_workbook(uploaded_file[0])
        qtr1_pnl, df1 = get_tables(book1[tabs[-1]], uploaded_file[0])  # send a sheet(not whole workbook)
        book2 = openpyxl.load_workbook(uploaded_file[1])
        qtr2_pnl, df2 = get_tables(book2[tabs[-1]], uploaded_file[1])  # send a sheet(not whole workbook)
        comp1_Name = uploaded_file[0].name.split('.')[0]
        comp2_Name = uploaded_file[1].name.split('.')[0]
        book1 = openpyxl.load_workbook(uploaded_file[0])
        qtr1_pnl, df1 = get_tables(book1[tabs[-1]], uploaded_file[0])  # send a sheet(not whole workbook)
        book2 = openpyxl.load_workbook(uploaded_file[1])
        qtr2_pnl, df2 = get_tables(book2[tabs[-1]], uploaded_file[1])  # send a sheet(not whole workbook)
        col1, col2 = st.columns([0.7, 0.3])
        with col1:
            main_menu = st.selectbox("Chose", funda_keys)
        with col2:
            sub_menu = st.selectbox("SubChose", list(df1.loc[main_menu].index))
        new_df = [df1.loc[main_menu].loc[sub_menu], df2.loc[main_menu].loc[sub_menu]]
        df_new = pd.concat(new_df, keys=[comp1_Name, comp2_Name], axis=1)
        group_bar(df_new, sub_menu)
        st.write(df_new)
#Custom CSS to remove header,footer, hamburger icon
hide_st_style = """
                <style>
                MainMenu {visibility: hidden;} 
                footer {visibility: hidden;}
                </style>
                """
st.markdown(hide_st_style,unsafe_allow_html=True)