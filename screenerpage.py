from selenium.webdriver.common.by import By
import streamlit as st
import glob
import os
import time
import datetime
import random
import urllib.parse
from urllib.parse import urlparse
import openpyxl
from openpyxl.utils import get_column_letter

import fundamentals
import processdriver
import pandas as pd
import create_database
import variables 
import config

# Credentials fetched from Config/Secrets
usermail, pswrd = config.Config.get_screener_credentials()

download_excel_xpath = "/html/body/main/div[3]/div[1]/form/button"
login_button_xpath = "/html/body/main/div[2]/div[2]/form/button"
follow_button_path = "/html/body/main/div[3]/div[1]/form/div/button"
BSE_Code_xpath = "/html/body/main/div[3]/div[2]/a[2]/span"
NSE_Code_xpath = "/html/body/main/div[3]/div[2]/a[3]/span" 
comp_fullname_xpath = "/html/body/main/div[3]/div[1]/div/h1" 
# sector_xpath = "/html/body/main/section[3]/div[1]/div[1]/p/a[1]"
sector_xpath = "/html/body/main/section[3]/div[1]/div[1]/p[1]/a[3]" #"/html/body/main/section[3]/div[1]/div[1]/p[1]/a[2]"
# industry_xpath = "/html/body/main/section[3]/div[1]/div[1]/p/a[2]"
industry_xpath = "/html/body/main/section[3]/div[1]/div[1]/p[1]/a[4]"

# global bsecodenum_codename
# global bsecodename_codenum
# bsecodenum_codename,bsecodename_codenum,bsecodenum_fullname,bsecodename_fullname,bsefullname_codenum,bsefullname_codename = nse_bse_search.bsecodenum_bsecodename()
# # This gets us the BSE NAME from the DAILY BHAVCOPY THAT WE ARE DOWNLOADING
# bsesccode_scname,bsescname_sccode = nse_bse_search.bseSCNAME_SCCODE()

global driver

with st.sidebar:
    # Use path from session state (set in Admin Portal) or default
    path_download = st.session_state.get('path_download', 'C:/Users/Sahaveer/Downloads/')
    path_bhav = './bhavcopy/'
    path_csv = './bhavcopy/csv/'
    path_to_save = os.path.join(path_download, "Results/")
    type_file = "*.xlsx"

def recently_downloaded_file(path_download, type_file):
    sendfile = ''
    # grabs the last created file in a specified folder
    pick_xlsx = glob.glob(os.path.join(path_download, type_file))
    
    if not pick_xlsx:
        return None
        
    last_created_file = max(pick_xlsx, key=os.path.getctime)
    send_file = last_created_file.replace("\\","/")
    return send_file

def is_404(driver):
    """Check if the current page is a 404 error page."""
    try:
        # Screener usually has "Page not found" in title or h1 for 404s
        if "Page not found" in driver.title or "404" in driver.title:
            return True
        h1_elements = driver.find_elements(by=By.TAG_NAME, value="h1")
        for h1 in h1_elements:
            if "Page not found" in h1.text or "404" in h1.text:
                return True
    except:
        pass
    return False

def login_screener(driver):
    driver.get('https://www.screener.in/login/')
    #if driver.find_elements_by_id("id_username"):
    if driver.find_elements(by=By.ID, value="id_username"):
        pass
    else:
        return
    #useridform = driver.find_element_by_id('id_username')
    useridform = driver.find_element(by=By.ID,value='id_username')
    pswrdform = driver.find_element(by=By.ID,value='id_password')
    #pswrdform = driver.find_element_by_id('id_password')
    useridform.send_keys(usermail)
    pswrdform.send_keys(pswrd)
    #if driver.find_elements_by_xpath(login_button_xpath):
    if driver.find_elements(by=By.XPATH, value=login_button_xpath):
        #driver.find_element_by_xpath(login_button_xpath).click()
        driver.find_element(by=By.XPATH, value=login_button_xpath).click()
    else:
        st.error("seems like Screener has changed the Xpath")


def month_diff(date1, date2):
    return (date1.year - date2.year) * 12 + (date1.month - date2.month)

def validate_dates_reverse(df):
# Convert columns to datetime
    cols_dates = pd.to_datetime(df.columns)    
    # Calculate the month difference for all consecutive dates
    initial_diff = month_diff(cols_dates[-1], cols_dates[-2])  # Calculate initial difference
    # st.info(f"initial difference is {initial_diff}")
    # Start reverse validation from the last date
    valid_dates = [cols_dates[-1]]
    if initial_diff in [3, 6, 12]:
        for i in range(len(cols_dates) - 2, -1, -1):
            current_date = cols_dates[i]
            last_valid_date = valid_dates[-1]
            # Calculate the difference in months between current date and the last valid date
            date_diff_months = month_diff(last_valid_date, current_date)
            # st.info(date_diff_months)
            # Allow differences of 3 (Quarterly), 6 (Half-Yearly), or 12 (Yearly) months
            if date_diff_months == initial_diff:
                valid_dates.append(current_date)
            else:
                st.warning(f"Missing or irregular date detected at {current_date}")
                break  # Stop when the pattern breaks
            # last_valid_date = cols_dates[i]

        # Reverse the list to maintain chronological order
        valid_dates.reverse()
        st.info(f"we are considering only {valid_dates}")
        # Create a new DataFrame only with valid dates
        valid_df = df[valid_dates]
        
        # Check if this data is valid
        # if len(valid_dates) == len(pd.date_range(start=valid_dates[0], end=valid_dates[-1], freq=frequency)):
        #     st.success(f"{frequency} data is sequential and valid")
        # else:
        #     st.warning(f"Some data points are missing or out of sequence")
        return valid_df
    else:
        return pd.DataFrame()


def scrape(driver):
    comp_dict = {}
    comp_dict['code_names'] = {}
    # get BSE code and link
    if driver.find_elements(by=By.XPATH, value=BSE_Code_xpath):
        bse_elements = driver.find_elements(by=By.XPATH, value=BSE_Code_xpath)
        bse_link_elements = driver.find_elements(by=By.XPATH, value=BSE_Code_xpath + "/..")
        
        bse_list = [element.text for element in bse_elements]
        split_bse_list =  bse_list[0].split(': ')
        comp_dict['code_names']['BSE'] = split_bse_list[-1]
        
        if bse_link_elements:
            comp_dict['bse_link'] = bse_link_elements[0].get_attribute("href")

    # get NSE code and link
    if driver.find_elements(by=By.XPATH, value=NSE_Code_xpath):
        nse_elements = driver.find_elements(by=By.XPATH, value=NSE_Code_xpath)
        nse_link_elements = driver.find_elements(by=By.XPATH, value=NSE_Code_xpath + "/..")
        
        nse_list = [element.text for element in nse_elements]
        # Print or use the extracted texts
        split_nse_list = nse_list[0].split(': ')
        # comp_dict['codes_dict'][split_nse_list[0]] = split_nse_list[1]
        comp_dict['code_names']['NSE'] = split_nse_list[-1]
        
        if nse_link_elements:
            comp_dict['nse_link'] = nse_link_elements[0].get_attribute("href")
    

    # get value of comp_fullname_xpath 
    if driver.find_elements(by=By.XPATH, value=comp_fullname_xpath):
        fullname_elements = driver.find_elements(by=By.XPATH, value=comp_fullname_xpath)
        # Extract the actual text from each WebElement
        fullname_list = [element.text for element in fullname_elements]
        # Print or use the extracted texts
        # st.success(fullname_list)
        comp_dict['comp_fullname'] = fullname_list[0]
        # st.success(f"{comp_dict['comp_fullname']}")
        
    
    #finding SECTOR
    if driver.find_elements(by=By.XPATH, value=sector_xpath):
        sector_elements = driver.find_elements(by=By.XPATH, value=sector_xpath)
        st.success(sector_elements)
        # Extract the actual text from each WebElement
        sector_list = [element.text for element in sector_elements]
        comp_dict['sector'] = sector_list[0]
        # Print or use the extracted texts
        # st.success(f"Sector : {sector_list[0]}")
    #finding Industry
    if driver.find_elements(by=By.XPATH, value=industry_xpath):
        industry_elements = driver.find_elements(by=By.XPATH, value=industry_xpath)
        st.success(industry_elements)
        #extract the actual text from each webelement
        industry_list = [element.text for element in industry_elements]
        comp_dict['industry'] = industry_list[0]
        # Print or use the extracted texts
        # st.success(f"Industry : {industry_list[0]}")

    comp_dict['consolidated_available'] = False
    comp_dict['standalone_available'] = False
    st.success(comp_dict)
    return comp_dict 

# lets get all industries and sectors from screener site for stocks already available in database
def scrape_all_listed(available_in_db:list)->None:
    driver = processdriver.getedgedriver()
    if driver is None:
        st.error("Scraping is not available in this environment.")
        return
    # LOGIN SCREENER SITE
    if driver.find_elements(by=By.XPATH, value='/html/body/nav/div[1]/div/div[1]/div/div[3]/div[1]'):
        pass
    else:
        login_screener(driver)

    for code in available_in_db:
        comp_dict ={}
        comp_dict['consolidated_available'] = False
        comp_dict['standalone_available'] = False        
        try:
            driver.get('https://www.screener.in/company/' + code + '/consolidated/')
            time.sleep(random.uniform(1, 5))
            #time.sleep(2)
            if driver.find_elements(by=By.XPATH, value=download_excel_xpath):
                parsed = urlparse(driver.current_url)
                path_parts = parsed.path.split('/')
                st.info(f"trying in consolidated {path_parts}")
                if len(path_parts) >= 2:
                    compcode_url = path_parts[2]  # gets code from within the screener_url                
                    #scrape few data from screener site 
                    comp_dict = scrape(driver)
                    comp_dict["consolidated_available"] = True 
            
        except Exception as HTTPError:
            st.error(f"HTTPError in consolidated {HTTPError}")
        
        try:
            driver.get('https://www.screener.in/company/' + code)
            time.sleep(random.uniform(1, 3))
            
            if is_404(driver):
                st.warning(f"🚫 {code} returned 404. Removing from listed list.")
                create_database.add_to_avoid_list(code)
                continue

            # time.sleep(2)
            # st.info(f"{variables.metadata['no_download_link']}")
            if driver.find_elements(by=By.XPATH, value=download_excel_xpath):
                parsed = urlparse(driver.current_url)
                path_parts = parsed.path.split('/')
                st.info(f"trying in standalone {path_parts}")
                if len(path_parts) >= 2:
                    compcode_url = path_parts[2]  # gets code from within the screener_url
                    if comp_dict['consolidated_available'] == False:
                        comp_dict = scrape(driver)
                        comp_dict['standalone_available'] = True
        except Exception as HTTPError:
            st.error(f"HTTPError for {code} in standalone {HTTPError}")
        
        st.success(comp_dict)
        #lets save these in MONGODB database
        if comp_dict['consolidated_available'] == True or comp_dict["standalone_available"] == True:
            if 'industry' in comp_dict.keys():
                create_database.insert_list(id_value=comp_dict['industry'],list_data= [code])
            if 'sector' in comp_dict.keys():
                create_database.insert_list(id_value=comp_dict['sector'],list_data= [code])

            if 'NSE' in comp_dict['code_names'].keys():
                search_in_db = comp_dict['code_names']['NSE']
            elif 'NSE - SME' in comp_dict['code_names'].keys():
                search_in_db = comp_dict['code_names']['NSE - SME']
            elif 'BSE' in  comp_dict['code_names'].keys():
                search_in_db = comp_dict['code_names']['BSE']
            elif 'BSE - SME' in  comp_dict['code_names'].keys():
                search_in_db = comp_dict['code_names']['BSE - SME']
            else:
                search_in_db = "NONE-NONE"
            if search_in_db != "NONE-NONE":
                create_database.insert_dict(col = create_database.company_metadata_col,id_value = search_in_db, save_within_document="comp_metadata", dict=comp_dict,task="REPLACE")


def df_to_dict(df):
    #convert timestamp in df.columns to string
    df.columns = pd.to_datetime(df.columns, format='%d-%m-%Y')
    df.columns = df.columns.astype(str)
    df_t = df.transpose()
    # st.dataframe(df_t)
    df_dict = df_t.to_dict(orient="index")
    # st.success(f"DF obtained from df_to_dict FUNC in screenerpage.py")
    # st.success(df_dict)
    return df_dict

def nested_df_to_dict(df):
    nested_dict = {}
    df.columns = df.columns.astype(str)
    for (section, item), row in df.iterrows():
        if section not in nested_dict:
            nested_dict[section] = {}
        nested_dict[section][item] = row.to_dict()
    return nested_dict

# def write_df_to_database(df,col,comp_code,key1,key2):
#     col = company_metadata_col
#     #lets try saving DataFrames as Dict in MONGODB database        
#     # st.dataframe(df)
#     #convert timestamp in df.columns to string
#     df.columns = pd.to_datetime(df.columns, format='%d-%m-%Y')
#     df.columns = df.columns.astype(str)
#     df_t = df.transpose()
#     # st.dataframe(df_t)
#     #convert dataframe to dict    
#     df_dict = df_t.to_dict(orient="index")
#     # st.success(df_dict)    
#     # pprint.pprint(qtr_pnl_dict)
#     if "results" in create_database.col.keys():
#         results = create_database.col["results"]
#         results[keys1][key2] = df_dict
#     else:
#         results = {key1 : {key2 : df_dict}}
#     create_database.insert_dict(col=create_database.col, id_value=comp_code, save_within_document="results", dict=results,task="REPLACE")


def read_database_to_get_df(id_value)->dict:
    #check if data is available in Database
    col = create_database.company_metadata_col
    if col.count_documents({"_id": id_value}):
        #documents of this id
        data = col.find_one({"_id": id_value})
        if "Yearly Consolidated" in data.keys():
            yr_cons_dict = data["Yearly Consolidated"]
            # st.success(dict)
            yr_cons = pd.DataFrame(yr_cons_dict)
            yr_cons.columns = pd.to_datetime(yr_cons.columns, format='%d-%m-%Y')
            # st.dataframe(yr_cons)    
        else:
            yr_cons_dict = {}
            yr_cons = pd.DataFrame()

        if "Yearly Standalone" in data.keys():
            yr_std_dict = data["Yearly Standalone"]
            # st.success(dict)
            yr_std = pd.DataFrame(yr_std_dict)
            # yr_std.columns = pd.to_datetime(yr_std.columns, format='%Y-%m-%d')
            yr_std.columns = pd.to_datetime(yr_std.columns, format='%d-%m-%Y')
            
            # st.dataframe(yr_std)
        else:
            yr_std_dict = {}
            yr_std = pd.DataFrame()
        if "Quarterly Consolidated" in data.keys():
            qtr_cons_dict = data["Quarterly Consolidated"]
            # st.success(dict)
            qtr_cons = pd.DataFrame(qtr_cons_dict)
            # qtr_cons.columns = pd.to_datetime(qtr_cons.columns, format='%Y-%m-%d')
            qtr_cons.columns = pd.to_datetime(qtr_cons.columns, format='%d-%m-%Y')
            # st.dataframe(qtr_cons)
        else:
            qtr_cons_dict = {}
            qtr_cons = pd.DataFrame()
        if "Quarterly Standalone" in data.keys():
            qtr_std_dict = data["Quarterly Standalone"]
            # st.success(dict)
            qtr_std = pd.DataFrame(qtr_std_dict)
            # qtr_std.columns = pd.to_datetime(qtr_std.columns, format='%Y-%m-%d')
            qtr_std.columns = pd.to_datetime(qtr_std.columns, format='%d-%m-%Y')
            # st.dataframe(qtr_std)
        else:
            qtr_std_dict = {}
            qtr_std = pd.DataFrame()
    # return yr_cons,yr_std,qtr_cons,qtr_std
    return yr_cons_dict,yr_std_dict,qtr_cons_dict,qtr_std_dict



# along with EXCEL SHEET DOWNLOAD AND PROCESSING, it TAKES FEW DATA FROM THE SCREENER SITE as well
def search_screener1(driver,code:str):
    code = str(code)
    st.success(f"Trying {code} in screenerpage.search_screener1 FUNC")
    NSE_ifnot_BSE = []
    consolidated_available = False
    standalone_available = False
    any_download_found = False
    # LOGIN SCREENER SITE
    if driver.find_elements(by=By.XPATH, value='/html/body/nav/div[1]/div/div[1]/div/div[3]/div[1]'):
        pass
    else:
        login_screener(driver)
        #link_consolidated = 'https://www.screener.in/company/'+code+'/consolidated/'
        #link_standalone =  'https://www.screener.in/company/'+code+'/'
        #st.info(f"Inside the search_screener function, we are searching for {code_list[i]}")
                    # FIRST SEARCH IS IN CONSOLIDATED
 
 
    # DOwnloads excel sheet, grabs feew necessary data and saves in Database
    
    # STANDALONE RESULTS
    driver.get('https://www.screener.in/company/' + code)
    time.sleep(random.uniform(1, 3))
    
    if is_404(driver):
        st.error(f"404 Error: Stock {code} not found on Screener.")
        raise ValueError(f"404 Error for {code}")

    # st.info(f"{variables.metadata['no_download_link']}")
    if driver.find_elements(by=By.XPATH, value=download_excel_xpath):
        parsed = urlparse(driver.current_url)
        path_parts = parsed.path.split('/')
        # st.info(f"tryingf in consolidated {path_parts}")
        if len(path_parts) >= 2:
            qtr_pnl2 = pd.DataFrame()
            df_comp = pd.DataFrame()
            # compcode_url = path_parts[2]  # gets code from within the screener_url
            #scrape few data from screener site 
            comp_dict = scrape(driver)
            comp_dict['standalone_available'] = True
            any_download_found = True
            driver.find_element(by=By.XPATH, value=download_excel_xpath).click()
            
            # Wait for file to appear (polling for 10 seconds)
            latest_file = None
            for _ in range(10):
                time.sleep(1)
                latest_file = recently_downloaded_file(path_download, type_file)
                if latest_file:
                    break
            
            if not latest_file:
                st.error(f"Could not find the downloaded file in {path_download}. Please check your sidebar path.")
                return

            book = openpyxl.load_workbook(latest_file)
            qtr_pnl2, df_comp = fundamentals.get_tables(book[fundamentals.tabs[-1]],
                                                    latest_file)  # send a sheet(not whole workbook)

            #standalone yearly data
            yrly_df1 = pd.DataFrame()
            df_comp2= pd.DataFrame()
            if not df_comp.eq(0).all().all():
                df_comp2 = df_comp
                # st.success(f" The processed DF we get from Standalone get_tables is :")
                # st.dataframe(df_comp2)
                pnl1, balancesht1,cashflow = fundamentals.develop_yearly(df_comp)            
                yrly_df1 = pd.concat([pnl1, balancesht1, df_comp2.loc["CASH FLOW", :]], axis=0)
                # st.success(f" The processed DF we get from Standalone develop_yearly is :")
                # st.dataframe(yrly_df1)

                # if len(df_comp.columns)==1:
                #     df_comp2 = df_comp
                # else:
                #     df_comp2 = validate_dates_reverse(df_comp)
                # # print(df_comp2)            
                # if not df_comp2.empty:
                #     pnl1, balancesht1,cashflow = fundamentals.develop_yearly(df_comp2)
                #     yrly_df1 = pd.concat([pnl1, balancesht1, df_comp2.loc["CASH FLOW", :]], axis=0)
            # st.dataframe(df_comp2)
            # dict_3 = df_to_dict(yrly_df1)
            dict_3 =nested_df_to_dict(df_comp2)
            # STANDALONE QUARTERLY
            qtr_df2 = pd.DataFrame() 
            if not qtr_pnl2.eq(0).all().all():   
                qtr_df2 = qtr_pnl2
                qtr_pnl2 = fundamentals.develop_quarterly(qtr_pnl2)

                # st.success(f" The processed Standalone Quarterly DF we get from develop_quarterly is:")
                # st.dataframe(qtr_pnl2)

                # if len(qtr_pnl2.columns)==1: 
                #     qtr_df2 = qtr_pnl2
                # else:          
                #     qtr_df2 = validate_dates_reverse(qtr_pnl2)
                # if not qtr_df2.empty:
                #     # st.success("Quarterly Standalone dates are equal")
                #     qtr_pnl2 = fundamentals.develop_quarterly(qtr_df2)
            dict_4 = df_to_dict(qtr_df2)
            # st.success(f"standalone_available is {standalone_available}")
            # ANALYSE FUNDAMENTAL DATA
            #Analyse Dataframes
            metadata = {}
            if not df_comp2.empty and df_comp2 is not None and isinstance(df_comp2,pd.DataFrame) and not qtr_df2.empty and qtr_df2 is not None and isinstance(qtr_df2,pd.DataFrame):
                metadata = fundamentals.analyse_df(pnl1, balancesht1, qtr_pnl2)
            else:
                if not qtr_df2.empty and qtr_df2 is not None and isinstance(qtr_df2,pd.DataFrame):
                    metadata = fundamentals.analyse_Q_df(qtr_pnl2)
                elif not df_comp2.empty and df_comp2 is not None and isinstance(df_comp2,pd.DataFrame):
                    metadata = fundamentals.analyse_Y_df(pnl1, balancesht1)
                
            final_dict3 = {"metadata":metadata}
            
            #lets save in id_value where the dict key starts with NSE
            # Get values where keys start with "NSE"
            NSE_ifnot_BSE = [value for key, value in comp_dict['code_names'].items() if key.startswith("NSE")]
            codenames_list = list(comp_dict['code_names'].values())
            # If there are no values that start with "NSE", take all available values
            if not NSE_ifnot_BSE:
                NSE_ifnot_BSE = codenames_list
            if len(NSE_ifnot_BSE)==0:
                NSE_ifnot_BSE = codenames_list
            # st.error(NSE_ifnot_BSE)
            if 'industry' in comp_dict.keys():
                create_database.insert_list(id_value=comp_dict['industry'],list_data= [NSE_ifnot_BSE[-1]])
            # st.success(f"{NSE_ifnot_BSE[-1]} is saved in {comp_dict['industry']}")
            final_dict3["STANDALONE"] = {}
            final_dict3["STANDALONE"]["YEARLY"]=dict_3
            final_dict3["STANDALONE"]["QUARTERLY"] = dict_4

            #lets try Saving The Dataframe in MongoDB
            final_dict3['code_names'] = codenames_list
            final_dict3['Code'] = NSE_ifnot_BSE[-1]

            #save the scraped data from site
            final_dict3['comp_metadata'] = comp_dict

            if len(metadata)>=1:
                # Saves the company in the INDUSTRY_COL 
                if len(metadata['tags'])>1:
                    for each in metadata['tags']:
                        create_database.insert_list(list_data= [NSE_ifnot_BSE[-1]], id_value = each)
                        # st.success(f"Saved Company in {each} TAG")

            variables.metadata[NSE_ifnot_BSE[-1]] = final_dict3
            create_database.create_doc(col=create_database.comp_metadata_col, id_value=NSE_ifnot_BSE[-1], dict=final_dict3)            
            st.success(f"{NSE_ifnot_BSE[-1]} among {codenames_list} is saved as \n{final_dict3}")

    else:
        st.error("No STANDALONE Data Available")
        #lets save v few data for such stocks
        comp_dict = scrape(driver)
        # Get values where keys start with "NSE"
        codenames_list = list(comp_dict['code_names'].values())
        if len(codenames_list)>=1:
            NSE_ifnot_BSE = [value for key, value in comp_dict['code_names'].items() if key.startswith("NSE")]

            # If there are no values that start with "NSE", take all available values
            if not NSE_ifnot_BSE:
                NSE_ifnot_BSE = codenames_list
                # NSE_ifnot_BSE = codenames_list[-1]
            final_dict3 = {}
            final_dict3['code_names'] = codenames_list
            final_dict3['Code'] = codenames_list[-1]
            final_dict3['comp_metadata'] = comp_dict
            create_database.create_doc(col=create_database.comp_metadata_col, id_value=NSE_ifnot_BSE[-1], dict=final_dict3)
            st.success(f"{NSE_ifnot_BSE[-1]} among {codenames_list} is saved as \n{final_dict3}")

    # Consolidated - 
    standalone_url = driver.current_url
    driver.get('https://www.screener.in/company/' + code + '/consolidated/')
    time.sleep(random.uniform(1, 5))
    
    # Check if we were redirected back to Standalone (common if no consolidated data exists)
    if driver.current_url == standalone_url:
        st.info(f"Skipping consolidated download for {code} (Same as Standalone)")
        return

    # if download button is available
    if driver.find_elements(by=By.XPATH, value=download_excel_xpath):
                                                                            #get the code from the opened url
        parsed = urlparse(driver.current_url)
        path_parts = parsed.path.split('/')
        # st.info(f"trying in consolidated {path_parts}")
        if len(path_parts) >= 2:
            qtr_pnl = pd.DataFrame()
            df_comp = pd.DataFrame()
            # compcode_url = path_parts[2]  # gets code from within the screener_url            
                                                        #scrape few data like Industry Sector Company Fullname NSEname BSEname from screener site 
            comp_dict = scrape(driver)              #returns scraped 
            comp_dict['consolidated_available'] = True
            any_download_found = True

            # excel button, click it                
            driver.find_element(by=By.XPATH, value=download_excel_xpath).click()
            
            # Wait for file to appear (polling for 10 seconds)
            latest_file = None
            for _ in range(10):
                time.sleep(1)
                latest_file = recently_downloaded_file(path_download, type_file)
                if latest_file:
                    break
            
            if not latest_file:
                st.error(f"Could not find the downloaded file in {path_download}. Please check your sidebar path.")
                return

            book = openpyxl.load_workbook(latest_file)

            #get dataframes from the opened excel sheet
            try:
                qtr_pnl, df_comp = fundamentals.get_tables(book[fundamentals.tabs[-1]],
                                                        latest_file)  # send a sheet(not whole workbook)
                # YEARLY  DATA
                yrly_df1 = pd.DataFrame()
                df_comp1 = pd.DataFrame()
                if not df_comp.eq(0).all().all():
                    # # we will check all the dates, validate it i.e. it will consider only till it is in sequence
                    # if len(df_comp.columns)==1:
                    #     df_comp1 = df_comp
                    # else:
                    #     df_comp1 = validate_dates_reverse(df_comp)
                    # st.dataframe(df_comp)
                    df_comp1 = df_comp
                    # st.success(f" The processed DF we get from CONSOLIDATED get_tables is :")
                    # st.dataframe(df_comp1)
                    if not df_comp1.empty:
                        pnl, balancesht,cashflow = fundamentals.develop_yearly(df_comp1)
                        yrly_df1 = pd.concat([pnl, balancesht, df_comp1.loc["CASH FLOW", :]], axis=0)                                  
                        # st.success(f"The DF we get from CONSOLIDATED fundamentals.develop_yearly is:")
                        # st.dataframe(yrly_df1)
                # st.dataframe(df_comp1)
                # dict_1 = df_to_dict(yrly_df1)
                dict_1 =nested_df_to_dict(df_comp1)

                #Quarterly Data
                qtr_pnl1 = pd.DataFrame()
                if not qtr_pnl.eq(0).all().all():            
                    qtr_df = qtr_pnl
                    # st.success(f"The Quarterly DF we get from CONSOLIDATED get_tables:")
                    # st.dataframe(qtr_df)                
                    qtr_pnl1 = fundamentals.develop_quarterly(qtr_df)
                    # st.success(f"The Quarterly DF we get from CONSOLIDATED fundamentals.develop_quarterly is:")
                    # st.dataframe(qtr_pnl1)
                    
                    # if len(qtr_pnl.columns)==1:
                    #     qtr_df = qtr_pnl
                    # else:
                    #     qtr_df = validate_dates_reverse(qtr_pnl)
                    # if not qtr_df.empty:
                    #     # st.success("Quarterly Consolidated dates are equal")
                    #     # Quarterly_data_consolidated = "valid"
                    #     qtr_pnl1 = fundamentals.develop_quarterly(qtr_df)
                dict_2 = df_to_dict(qtr_df)
                
                # ANALYSE FUNDAMENTAL DATA
                #Analyse Dataframes
                metadata = {}
                if not df_comp1.empty and df_comp1 is not None and isinstance(df_comp1,pd.DataFrame) and not qtr_pnl1.empty and qtr_pnl1 is not None and isinstance(qtr_pnl1,pd.DataFrame):                    
                    metadata = fundamentals.analyse_df(pnl, balancesht, qtr_pnl1)                
                else:
                    if not qtr_pnl1.empty and qtr_pnl1 is not None and isinstance(qtr_pnl1,pd.DataFrame):
                        metadata = fundamentals.analyse_Q_df(qtr_pnl1)
                    elif not df_comp1.empty and df_comp1 is not None and isinstance(df_comp1,pd.DataFrame):
                        metadata = fundamentals.analyse_Y_df(pnl, balancesht)
                
                final_dict1 = {"metadata" : metadata}

                #make a list from the dictionary values
                codenames_list = list(comp_dict['code_names'].values())
                
                #lets save in id_value where the dict key starts with NSE
                # Get values where keys start with "NSE"
                NSE_ifnot_BSE = [value for key, value in comp_dict['code_names'].items() if key.startswith("NSE")]
                # If there are no values that start with "NSE", take all available values
                if not NSE_ifnot_BSE:
                    NSE_ifnot_BSE = codenames_list
                # st.error(NSE_ifnot_BSE)
                if 'industry' in comp_dict.keys():
                    create_database.insert_list(id_value=comp_dict['industry'],list_data= [NSE_ifnot_BSE[-1]])                                                                                                                                                                
                    st.success(f"{NSE_ifnot_BSE[-1]} among {codenames_list} is saved in {comp_dict['industry']}")
                final_dict1["CONSOLIDATED"] = {}
                final_dict1["CONSOLIDATED"]["YEARLY"]=dict_1
                # Saves the company in the INDUSTRY_COL 
                if len(metadata['tags'])>=1:
                    for each in metadata['tags']:
                        create_database.insert_list(list_data = [NSE_ifnot_BSE[-1]], id_value= each)
                        st.success(f"{NSE_ifnot_BSE[-1]} among {codenames_list} is saved in {each} as well")

                #lets try Saving The Dataframe in MongoDB
                final_dict1['code_names'] = codenames_list
                final_dict1['Code'] = NSE_ifnot_BSE[-1]

                #save the scraped data from site
                final_dict1['comp_metadata'] = comp_dict
                final_dict1["CONSOLIDATED"]["QUARTERLY"] = dict_2
                create_database.create_doc(col=create_database.comp_metadata_col, id_value=NSE_ifnot_BSE[-1], dict=final_dict1)
                st.success(f"{NSE_ifnot_BSE[-1]} among {codenames_list} is saved as \n{final_dict1}")
                # st.success(f"{NSE_ifnot_BSE[-1]} among {codenames_list} is saved as \n{final_dict1}")
                variables.metadata[NSE_ifnot_BSE[-1]] = final_dict1
                # variables.metadata[final_dict1['Code']] = final_dict1
                # create_database.insert_dict(col=create_database.company_metadata_col, id_value=codenames_list[0], save_within_document="results", dict=final_dict2,task="REPLACE")         
                # st.success(f"Saved entire Metadata in Database as {NSE_ifnot_BSE[-1]}")
            except Exception as e:
                st.error(f"Error: {e}")
    
    if not any_download_found:
        st.warning(f"⚠️ No download button available for {code}. This script might be an ETF or illiquid stock.")
        raise ValueError(f"No Data Available for {code}")
    else:
        st.error("No Data Available for Consolidated")        #lets save v few data for such stocks
        comp_dict = scrape(driver)
        # Get values where keys start with "NSE"
        codenames_list = list(comp_dict['code_names'].values())

        if len(codenames_list)>=1:
            final_dict1 = {}
            NSE_ifnot_BSE = [value for key, value in comp_dict['code_names'].items() if key.startswith("NSE")]
            # If there are no values that start with "NSE", take all available values
            if not NSE_ifnot_BSE:
                NSE_ifnot_BSE = codenames_list
            final_dict1={}
            final_dict1['code_names'] = codenames_list
            final_dict1['Code'] = codenames_list[-1]
            final_dict1['comp_metadata'] = comp_dict            
            create_database.create_doc(col=create_database.comp_metadata_col, id_value=NSE_ifnot_BSE[-1], dict=final_dict1)
            st.success(f"{NSE_ifnot_BSE[-1]} among {codenames_list} is saved as \n{final_dict1}")


def main():
    last_date_of_quarter="2024-12-31"
    driver = processdriver.getedgedriver()
    if driver is None:
        st.error("Scraping is not available in this environment.")
        return
    # 2. Scraper Configuration
    codes = []
    # Fetch master list from MongoDB (Avoids local file dependency)
    alllisted_list = create_database.get_all_listed_stocks()

    if st.button("SCAN ALL LISTED STOCKS"):
        codes = alllisted_list
    
    compCode = st.text_input(label="Give a CODE",value="")
    # Split by comma and space, and handle the quotes separately
    codes_with_quotes = re.findall(r"'(.*?)'", compCode)
    codes_without_quotes = re.split(r',\s*|\s+', compCode)
    # Combine the results and remove empty strings
    code_list = [code.strip() for code in codes_with_quotes + codes_without_quotes if code.strip()]
    if st.button("Download Data to DB"):
        if len(code_list)>0:
            codes = code_list

    if len(codes)>0:
        for code in codes:
            if create_database.comp_metadata_col.count_documents({"code_names": code}):
                doc_is = create_database.comp_metadata_col.find_one({"code_names":code})
                if doc_is is not None:
                    # st.success(doc_is)
                    if "CONSOLIDATED" in doc_is.keys():
                        # get the last key value saved in the dict of doc_is["CONSOLIDATED"]['QUARTERLY']
                        listed_dict_keys = list(doc_is['CONSOLIDATED']['QUARTERLY'])
                        if len(listed_dict_keys) > 0:
                            last_quarter_announced = listed_dict_keys[-1]
                    elif "STANDALONE" in doc_is.keys():
                        # get the last key value saved in the dict of doc_is["CONSOLIDATED"]['QUARTERLY']
                        listed_dict_keys = list(doc_is['STANDALONE']['QUARTERLY'])
                        if len(listed_dict_keys)>0:
                            last_quarter_announced = listed_dict_keys[-1]
                    # st.success(last_quarter_announced)
                    # st.success(type(last_quarter_announced))
                    if last_quarter_announced != last_date_of_quarter and (datetime.datetime.now() - doc_is['timestamp']).days>1:
                        search_screener1(driver,code)
                    else:
                        st.success(f"We already got LATEST RESULTS for {code} : {last_quarter_announced}")
                else:
                    search_screener1(driver,code)
            else:
                search_screener1(driver,code)

    # simple_search = st.text_input(label="Simple Funda data", value="RELIANCE")
    # if st.button("Search DB"):

if __name__ == '__main__':
    main()

#END of the CODE