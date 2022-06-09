import datetime
from datetime import date
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import plotly.express as px
from plotly.subplots import make_subplots
import plotly.figure_factory as ff
from streamlit_option_menu import option_menu
import streamlit.components.v1 as html
from PIL import Image
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

tabs = ['Profit & Loss', 'Quarters','Balance Sheet', 'Cash Flow' ,'Data Sheet']
table = ['PROFIT & LOSS', 'Quarters', 'BALANCE SHEET' ]
# lets define the first and last table keys to extract the exact table size
DataSheet_Key_Values = ['PROFIT & LOSS', 'Dividend Amount', 'Quarters', 'Operating Profit', 'BALANCE SHEET', 'Cash & Bank',
                        'CASH FLOW:', 'Net Cash Flow']


results_saved_at = 'C:/Users/sahaveer/OneDrive/Documents/bhavcopy/'
xlsx_to_read = 'Cravatex'
type_of_file_chart = '.xlsx'
logo = Image.open(r'./image/logo.png')
st.title('iTimes')
def go_bar(df, row_name):
    #fig = ff.create_table(df)
    #st.plotly_chart(fig)
    #st.write(df)                                                   # this is givng data conversion error sometimes
    fig = go.Figure(data = [go.Bar(x = df.columns,y = df.loc[row_name],
                                   text = df.loc[row_name], textposition = 'auto')])
    # go.bar has another attribute - hovertext = ['27% market share', '24% market share', '19% market share']
    fig.update_traces(marker_color=color_bar, marker_line_color='rgb(8,48,107)',
                      marker_line_width=1.5, opacity=1, texttemplate='%{text:.2s}', textposition='outside')
    fig.update_layout(height=600,width=900,
                      title={'text': row_name + ' Report',
                             'y': 0.9,'x': 0.5,'xanchor': 'center','yanchor': 'top'},
                      xaxis_tickfont_size=14, xaxis_tickangle=-45,
                      xaxis=dict(showgrid=False),
                      yaxis=dict(showgrid=False, title='INR (cr)',titlefont_size=16,tickfont_size=14,),
                      legend=dict(x=0,y=1.0,bgcolor='rgba(255, 255, 255, 0)',bordercolor='rgba(255, 255, 255, 0)'),bargap=0.15)
    st.plotly_chart(fig)

def go_bar_line(df,row_name):
    temp_df = df.loc[row_name]
    df_sales_qoq = (temp_df.pct_change() * 100)
    df_sales_qoq.name = row_name + '_QoQ'
    df2 = df.append(df_sales_qoq)
    # Create figure with secondary y-axis
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace( go.Bar(x=df2.columns, y=df2.loc[row_name], name=row_name, textposition = 'auto',text = df.loc[row_name]),
                   secondary_y=False)
    fig.update_traces(marker_color=color_bar, marker_line_color='rgb(8,48,107)',
                      marker_line_width=1.5, opacity=1, texttemplate='%{text:.2s}', textposition='outside')
    fig.add_trace( go.Scatter(x=df2.columns, y=df2.iloc[-1], name=row_name + " QoQ", text = df.iloc[-1]),
                   secondary_y=True)
    # Add figure title
    fig.update_layout( height=600, width=900,
                       title={'text': row_name + ' Report',
                              'y': 0.9, 'x': 0.5, 'xanchor': 'center', 'yanchor': 'top'},
                       xaxis_tickfont_size=14, xaxis_tickangle=-45,
                       yaxis=dict(title='INR (cr)', titlefont_size=16, tickfont_size=14,),
                       legend=dict(yanchor="top",y=0.99,xanchor="left",x=0.01),
                       bargap = 0.15)
    # Set y-axes titles
    fig.update_yaxes(title_text="<b>" + row_name + "</b> in cr ", secondary_y=False,showgrid=False)
    fig.update_yaxes(title_text="<b>" + row_name + " QoQ</b> in % ", secondary_y=True,showgrid=False)
    st.plotly_chart(fig)

def go_group_bar(df, row_name):
    bar_list = list(df.index)
    fig = go.Figure(data=[ go.Bar(name=bar_list[0], x=df.columns, y=df.iloc[0], textposition = 'auto', text = df.iloc[0]),
                           go.Bar(name=bar_list[1],  x=df.columns, y=df.iloc[1], textposition = 'auto', text = df.iloc[1]),
                           go.Bar(name=bar_list[2], x=df.columns, y=df.iloc[2], textposition='auto', text=df.iloc[2]),
                           go.Bar(name=bar_list[3], x=df.columns, y=df.iloc[3], textposition='auto', text=df.iloc[3])
                           ])
    # Change the bar mode
    fig.update_layout(barmode='group', bargroupgap=0.1,
                      height=600, width=900,
                      title={'text': row_name + ' Report',
                             'y': 0.9, 'x': 0.5, 'xanchor': 'center', 'yanchor': 'top'},
                      xaxis_tickfont_size=14, xaxis_tickangle=-45,
                      yaxis=dict(title='INR (cr)', titlefont_size=16, tickfont_size=14, ),
                      legend=dict(yanchor="top",y=0.99,xanchor="left",x=0.01),)   # this barmode = 'group | stack | 'relative''
    #fig.update_layout(barmode='stack', xaxis={'categoryorder':'category ascending'}) # WHILE USING STACK
    st.plotly_chart(fig)

def get_tables(datasht,file):
    for i in range(1,datasht.max_row+1) :
        if datasht['A'+str(i)].value == DataSheet_Key_Values[0] :
            pnl_start_row = i
        if datasht['A' + str(i)].value == DataSheet_Key_Values[1]:
            pnl_end_row = i-1
        if datasht['A' + str(i)].value == DataSheet_Key_Values[2]:
            quarterly_start_row = i
        if datasht['A' + str(i)].value == DataSheet_Key_Values[3]:
            quarterly_end_row = i-1
        if datasht['A' + str(i)].value == DataSheet_Key_Values[4]:
            BS_start_row = i
        if datasht['A' + str(i)].value == DataSheet_Key_Values[5]:
            BS_end_row = i-1
        if datasht['A' + str(i)].value == DataSheet_Key_Values[6]:
            cash_start_row = i
        if datasht['A' + str(i)].value == DataSheet_Key_Values[7]:
            cash_end_row = i-1
    reqd_cols = "A :" + str(get_column_letter(datasht.max_column))

    if pnl_start_row is not None and pnl_end_row is not None :
        pnl = pd.read_excel(file, index_col=0, sheet_name=tabs[-1], header=pnl_start_row,usecols=reqd_cols,
                            nrows=pnl_end_row-pnl_start_row )
    if quarterly_start_row is not None and quarterly_end_row is not None :
        qtr_pnl = pd.read_excel(file, index_col=0, sheet_name=tabs[-1], header=quarterly_start_row, usecols=reqd_cols,
                                nrows=quarterly_end_row - quarterly_start_row)
    if BS_start_row is not None and BS_end_row is not None:
        balancesht = pd.read_excel(file, index_col=0, sheet_name=tabs[-1], header=BS_start_row,usecols=reqd_cols,
                                   nrows=BS_end_row-BS_start_row )
    if cash_start_row is not None and cash_end_row is not None:
        cashflow = pd.read_excel(file, index_col=0, sheet_name=tabs[-1], header=cash_start_row,usecols=reqd_cols,
                                   nrows=cash_end_row-cash_start_row )
    return pnl,qtr_pnl, balancesht,cashflow
color_hover = "#0D7292"
color_background = "#32B7D6"
with st.sidebar:
    #tab1_color = st.sidebar.color_picker("tab1",value="#1799D0")
    #tab2_color = st.sidebar.color_picker("tab2",value="#4789A4")
    main_menu = option_menu("Main Menu", ["About", "Fundamental Charts", "BhavCopy", "Contact"],
                         icons=['house', 'file-slides', 'app-indicator', 'person lines fill'],
                         menu_icon="list", default_index=0,
                         styles={
                             "container": {"padding": "5!important"},
                             "icon": {"color": "orange", "font-size": "25px"},
                             "nav-link": {"font-size": "16px", "text-align": "left", "margin": "0px",
                                          "--hover-color": color_hover},
                             "nav-link-selected": {"background-color": color_background}})

if main_menu == "About":
    col1, col2 = st.columns( [0.8, 0.2])
    with col1:
        st.markdown(""" <style> .font {
        font-size:35px ; font-family: 'Cooper Black'; color: #FF9633;} 
        </style> """, unsafe_allow_html=True)
        st.markdown('<p class="font">About the Creator</p>', unsafe_allow_html=True)
    with col2:               # To display brand log
        st.image(logo, width=130)
    st.write("We @iTimes are trying to create basic DIY fundamental analysis. \n\n We shall try bringing you here bse announcements, news, amibroker eod data here")

elif main_menu == "Fundamental Charts":
    # Add a file uploader to allow users to upload their csv file
    st.markdown(""" <style> .font {
        font-size:25px ; font-family: 'Cooper Black'; color: #FF9633;} 
        </style> """, unsafe_allow_html=True)
    st.markdown('<p class="font">Upload your xlsx/xlsm from screener.in </p>',
                unsafe_allow_html=True)  # use st.markdown() with CSS style to create a nice-formatted header/text
    col1, col2 = st.columns([0.8, 0.2])
    with col1:
        uploaded_file = st.file_uploader("", type=['xlsx', 'xlsm'])  # Only accepts xlsx,xlsm file format
    with col2:
        color_bar = st.color_picker("Bar",value = "#0f7eec")
        color_line = st.color_picker("Line",value = "#D60A10")
    if uploaded_file is not None:
        book = openpyxl.load_workbook(uploaded_file)
        pnl, qtr_pnl, balance_sht, cash_flow = get_tables(book[tabs[-1]],uploaded_file)  # send a sheet(not whole workbook)
        sub_choose = option_menu("Fundamentals", ["Yearly PnL", "Quarterly PnL", "Balance Sheet", "Cash Flow"],
                                 icons=['house', 'file-slides', 'app-indicator', 'person lines fill'],
                                 styles={"container": {"padding": "5!important"},
                                         "icon": {"color": "orange", "font-size": "25px"},
                                         "nav-link": {"font-size": "16px", "text-align": "left", "margin": "0px","--hover-color": color_hover},
                                         "nav-link-selected": {"background-color": color_background}},
                                 menu_icon="cast", default_index=0, orientation="horizontal")
        if sub_choose == "Yearly PnL":
            index_list = ["key_params"] + list(pnl.index)
            #param = st.selectbox("Select  column", pnl.index)
            with st.sidebar:
                param = option_menu(sub_choose, index_list,
                                    styles={"container": {"padding": "5!important"},
                                            "icon": {"color": "orange", "font-size": "25px"},
                                            "nav-link": {"font-size": "16px", "text-align": "left", "margin": "0px",
                                                         "--hover-color": color_hover},
                                            "nav-link-selected": {"background-color": color_background}},
                                    menu_icon="cast", default_index=0, orientation="vertical")
            if param == "key_params":
                go_bar_line(pnl,"Sales")
                go_bar_line(pnl,"Profit before tax")
                go_bar_line(pnl,"Net profit")
            else:
                if st.checkbox("QoQ Growth"):
                    go_bar_line(pnl, param)
                else:
                    go_bar(pnl, param)

        if sub_choose == "Quarterly PnL":
            #param = st.selectbox("Select  column", qtr_pnl.index)
            index_list = ["key_params"] + list(qtr_pnl.index)
            with st.sidebar:
                param = option_menu(sub_choose, index_list,
                                    styles={"container": {"padding": "5!important"},"icon": {"color": "orange", "font-size": "25px"},
                                            "nav-link": {"font-size": "16px", "text-align": "left", "margin": "0px",
                                                         "--hover-color": color_hover},
                                            "nav-link-selected": {"background-color": color_background}},
                                    menu_icon="cast", default_index=0, orientation="vertical")
            if param == "key_params":
                go_bar_line(qtr_pnl, "Sales")
                go_bar_line(qtr_pnl, "Profit before tax")
                go_bar_line(qtr_pnl, "Net profit")
            else:
                if st.checkbox("QoQ Growth"):
                    go_bar_line(qtr_pnl, param)
                else:
                    go_bar(qtr_pnl, param)

        if sub_choose == "Balance Sheet":
            #param = st.selectbox("Select  column", balance_sht.index)
            index_list = ["key_params"] + list(balance_sht.index)
            with st.sidebar:
                param = option_menu(sub_choose, index_list,
                                    styles={"container": {"padding": "5!important"},"icon": {"color": "orange", "font-size": "25px"},
                                            "nav-link": {"font-size": "16px", "text-align": "left", "margin": "0px",
                                                         "--hover-color": color_hover},
                                            "nav-link-selected": {"background-color": color_background}},
                                    menu_icon="cast", default_index=0, orientation="vertical")
            if param == "key_params":
                go_bar_line(balance_sht, "Reserves")
                go_bar_line(balance_sht, "Borrowings")
                go_bar_line(balance_sht, "Capital Work in Progress")
                go_bar_line(balance_sht, "Cash & Bank")
            else:
                if st.checkbox("QoQ Growth"):
                    go_bar_line(balance_sht, param)
                else:
                    go_bar(balance_sht, param)
        if sub_choose == "Cash Flow":
            index_list = ["key_params"] + list(cash_flow.index)
            #param = st.selectbox("Select  column", cash_flow.index)
            with st.sidebar:
                param = option_menu(sub_choose, index_list,
                                    styles={"container": {"padding": "5!important"},
                                            "icon": {"color": "orange", "font-size": "25px"},
                                            "nav-link": {"font-size": "16px", "text-align": "left", "margin": "0px",
                                                         "--hover-color": color_hover},
                                            "nav-link-selected": {"background-color": color_background}},
                                    menu_icon="cast", default_index=0, orientation="vertical")
            if param == "key_params":
                go_group_bar(cash_flow,"cash_flows")
            else:
                if st.checkbox("QoQ Growth"):
                    go_bar_line(cash_flow, param)
                else:
                    go_bar(cash_flow, param)

elif main_menu == "BhavCopy":
    st.markdown("### Site is in progress \n Shall be launched asap")
    my_date = st.date_input("Select date", value=date.today(),
                            min_value=datetime.date(1990, 1, 1))
    ddmmmyyyy = my_date.strftime("%d%b%Y")
    driver = webdriver.Edge(r"C://Users/sahaveer/PycharmProjects/onlystocks/msedgedriver.exe")
    if st.button("Download"):
        EOD.eod_date(driver,ddmmmyyy)

elif main_menu == 'Contact':
    col1, col2 = st.columns([0.8, 0.2])
    with col1:  # To display the header text using css style
        st.markdown(""" <style> .font {
            font-size:35px ; font-family: 'Cooper Black'; color: #FF9633;} 
            </style> """, unsafe_allow_html=True)
        st.markdown("Contact us through \n ## [telegram](https://t.me/itimesalgo/) \n ## [Twitter](https://twitter.com/itimesalgo)")
        #st.markdown('<p class="font">Contact us through [telegram](https://t.me/itimesalgo/). </p>', unsafe_allow_html=True)
    with col2:  # To display brand log
        st.image(logo, width=130)
    st.write("We sincerely appreciates your suggestions and contribution to improvise our iTimes community.")

#REFERENCE :
#FLASK : https://www.datasciencelearner.com/how-to-create-a-bar-chart-from-a-dataframe-in-python/#:~:text=There%20is%20also%20another%20method%20to%20create%20a,y-axis%20values%20you%20want%20to%20draw%20the%20bar.

#Streamlit Basics : https://www.datacamp.com/tutorial/streamlit#on-windows-

# https://towardsdatascience.com/make-dataframes-interactive-in-streamlit-c3d0c4f84ccb#:~:text=When%20building%20data%20apps%20using%20Streamlit%2C%20sometimes%20you,displayed%20in%20the%20app%20looks%20plain%20and%20static.

#https://towardsdatascience.com/create-a-bar-chart-race-animation-app-using-streamlit-and-raceplotly-e44495249f11